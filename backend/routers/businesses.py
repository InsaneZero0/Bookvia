"""
Auto-extracted router from server.py refactoring.
"""
from fastapi import APIRouter, HTTPException, Depends, Request, BackgroundTasks, status, File, UploadFile
from fastapi.security import HTTPAuthorizationCredentials
from fastapi.responses import Response
from typing import List, Optional, Dict, Any
from datetime import datetime, timezone, timedelta
from pydantic import BaseModel, ConfigDict, EmailStr
import logging
import os
import re
import random
import uuid
import json
import math
import pytz
from bson import ObjectId

from core.database import db
from core.config import ENV, BASE_URL, ADMIN_EMAIL
from core.security import (
    TokenData, create_token, decode_token,
    hash_password, verify_password,
    generate_totp_secret, verify_totp, generate_totp_qr
)
from core.dependencies import (
    security, get_current_user, require_auth, require_business, require_admin
)
from core.helpers import (
    generate_id, generate_slug, calculate_bayesian_rating,
    is_user_blacklisted, send_sms, create_notification,
    create_audit_log, create_business_activity,
    amount_to_cents, cents_to_amount,
    create_ledger_entry, create_transaction_ledger_entries,
    calculate_business_ledger_summary
)
from models.enums import (
    UserRole, AppointmentStatus, BusinessStatus, PaymentStatus,
    TransactionStatus, LedgerDirection, LedgerAccount, LedgerEntryStatus,
    SettlementStatus, AuditAction,
    PLATFORM_FEE_PERCENT, HOLD_EXPIRATION_MINUTES, MIN_DEPOSIT_AMOUNT,
    SUBSCRIPTION_PRICE_MXN, SUBSCRIPTION_PRICE_USD, SUBSCRIPTION_TRIAL_DAYS,
    VISIBLE_BUSINESS_FILTER, DEFAULT_MANAGER_PERMISSIONS, visible_business_filter_now
)
from models.schemas import *

logger = logging.getLogger(__name__)

import stripe as stripe_lib
from core.stripe_config import STRIPE_API_KEY, get_or_create_stripe_price
stripe_lib.api_key = STRIPE_API_KEY
if "sk_test_emergent" in (STRIPE_API_KEY or ""):
    stripe_lib.api_base = "https://integrations.emergentagent.com/stripe"
from services.cloudinary_service import upload_image, delete_image as cloudinary_delete, is_configured as cloudinary_configured, validate_image
from services.storage import init_storage, put_object, get_object, generate_upload_path, ALLOWED_IMAGE_TYPES, ALLOWED_IMAGE_EXTENSIONS, MAX_FILE_SIZE

router = APIRouter(prefix="/businesses", tags=["Businesses"])


def validate_schedule_blocks(schedule: dict):
    """Validate that schedule blocks don't overlap within a day."""
    for day_key, day_schedule in schedule.items():
        if not day_schedule.is_available or not day_schedule.blocks:
            continue
        blocks = sorted(day_schedule.blocks, key=lambda b: b.start_time)
        for i in range(len(blocks) - 1):
            if blocks[i].end_time > blocks[i + 1].start_time:
                raise HTTPException(
                    status_code=400,
                    detail=f"Overlapping blocks on day {day_key}: {blocks[i].end_time} > {blocks[i+1].start_time}"
                )


@router.get("/my/dashboard-summary")
async def get_dashboard_summary(token_data: TokenData = Depends(require_business)):
    """Quick dashboard summary: today, this week, comparison."""
    business = await db.businesses.find_one({"user_id": token_data.user_id}, {"_id": 0, "id": 1})
    if not business:
        raise HTTPException(status_code=404, detail="Business not found")
    bid = business["id"]
    now = datetime.now(timezone.utc)
    today = now.strftime("%Y-%m-%d")
    week_start = (now - timedelta(days=now.weekday())).strftime("%Y-%m-%d")
    prev_week_start = (now - timedelta(days=now.weekday() + 7)).strftime("%Y-%m-%d")
    month_start = now.replace(day=1).strftime("%Y-%m-%d")
    thirty_days_ago = (now - timedelta(days=30)).strftime("%Y-%m-%d")

    # Today
    today_bookings = await db.bookings.count_documents({"business_id": bid, "date": today})
    today_completed = await db.bookings.count_documents({"business_id": bid, "date": today, "status": "completed"})
    today_revenue_pipe = [
        {"$match": {"business_id": bid, "date": today, "deposit_paid": True}},
        {"$group": {"_id": None, "t": {"$sum": "$deposit_amount"}}}
    ]
    today_rev = await db.bookings.aggregate(today_revenue_pipe).to_list(1)
    today_revenue = today_rev[0]["t"] if today_rev else 0

    # This week
    week_bookings = await db.bookings.count_documents({"business_id": bid, "date": {"$gte": week_start}})
    week_rev_pipe = [
        {"$match": {"business_id": bid, "date": {"$gte": week_start}, "deposit_paid": True}},
        {"$group": {"_id": None, "t": {"$sum": "$deposit_amount"}}}
    ]
    week_rev = await db.bookings.aggregate(week_rev_pipe).to_list(1)
    week_revenue = week_rev[0]["t"] if week_rev else 0

    # Previous week (for comparison)
    prev_week_bookings = await db.bookings.count_documents({
        "business_id": bid, "date": {"$gte": prev_week_start, "$lt": week_start}
    })

    # This month
    month_bookings = await db.bookings.count_documents({"business_id": bid, "date": {"$gte": month_start}})
    month_rev_pipe = [
        {"$match": {"business_id": bid, "date": {"$gte": month_start}, "deposit_paid": True}},
        {"$group": {"_id": None, "t": {"$sum": "$deposit_amount"}}}
    ]
    month_rev = await db.bookings.aggregate(month_rev_pipe).to_list(1)
    month_revenue = month_rev[0]["t"] if month_rev else 0

    # Pending reviews (unresponded)
    new_reviews = await db.reviews.count_documents({
        "business_id": bid,
        "created_at": {"$gte": (now - timedelta(days=7)).isoformat()}
    })

    # Client retention: unique clients this month vs last month
    this_month_clients_pipe = [
        {"$match": {"business_id": bid, "date": {"$gte": month_start}}},
        {"$group": {"_id": "$user_id"}}
    ]
    this_month_clients = len(await db.bookings.aggregate(this_month_clients_pipe).to_list(1000))

    week_change = round(((week_bookings - prev_week_bookings) / max(prev_week_bookings, 1)) * 100) if prev_week_bookings else 0

    # Phase 14: Profile views + conversion + top services (last 30d)
    views_30d = await db.profile_views.count_documents({
        "business_id": bid, "date": {"$gte": thirty_days_ago}
    })
    bookings_30d = await db.bookings.count_documents({
        "business_id": bid, "date": {"$gte": thirty_days_ago}
    })
    conversion_pct = round((bookings_30d / views_30d) * 100, 1) if views_30d > 0 else 0.0

    top_services_pipe = [
        {"$match": {"business_id": bid, "date": {"$gte": thirty_days_ago},
                    "service_id": {"$ne": None}}},
        {"$group": {"_id": "$service_id", "count": {"$sum": 1},
                    "revenue": {"$sum": {"$ifNull": ["$deposit_amount", 0]}}}},
        {"$sort": {"count": -1}},
        {"$limit": 3},
    ]
    top_rows = await db.bookings.aggregate(top_services_pipe).to_list(3)
    top_services = []
    for r in top_rows:
        svc = await db.services.find_one({"id": r["_id"]}, {"_id": 0, "name": 1}) or {}
        top_services.append({
            "service_id": r["_id"],
            "name": svc.get("name", "-"),
            "bookings": r["count"],
            "revenue": round(r.get("revenue", 0), 2),
        })

    return {
        "today": {"bookings": today_bookings, "completed": today_completed, "revenue": round(today_revenue, 2)},
        "week": {"bookings": week_bookings, "revenue": round(week_revenue, 2), "change_pct": week_change},
        "month": {"bookings": month_bookings, "revenue": round(month_revenue, 2), "unique_clients": this_month_clients},
        "new_reviews": new_reviews,
        "profile_views_30d": views_30d,
        "bookings_30d": bookings_30d,
        "conversion_pct": conversion_pct,
        "top_services": top_services,
    }


# ========================== PHASE 14: PROFILE VIEW TRACKING ==========================


async def _track_profile_view(
    business_id: str,
    request: Request,
    current_user: Optional[TokenData],
    owner_user_id: Optional[str],
) -> None:
    """Fire-and-forget best-effort tracker.

    Deduped per (business, viewer, day) using an idempotent upsert. Owners
    visiting their own profile do not count. Catches all errors so a
    tracking failure never breaks the public read endpoint.
    """
    try:
        if current_user and current_user.user_id == owner_user_id:
            return
        if current_user:
            viewer_key = current_user.user_id
        else:
            # Behind k8s ingress `request.client.host` rotates across upstream
            # hops; use the first X-Forwarded-For entry when present so anon
            # dedup is stable per calendar day.
            xff = (request.headers.get("x-forwarded-for", "") if request else "").split(",")
            viewer_key = (xff[0].strip() if xff and xff[0].strip() else None) or (
                request.client.host if request and request.client else "anonymous"
            )
        day = datetime.now(timezone.utc).strftime("%Y-%m-%d")
        await db.profile_views.update_one(
            {"business_id": business_id, "viewer_key": viewer_key, "date": day},
            {"$setOnInsert": {
                "business_id": business_id,
                "viewer_key": viewer_key,
                "date": day,
                "first_seen_at": datetime.now(timezone.utc).isoformat(),
            }},
            upsert=True,
        )
    except Exception:
        logger.debug("profile-view tracker failed", exc_info=True)


@router.get("/my/profile-completion")
async def get_profile_completion(token_data: TokenData = Depends(require_business)):
    """Checklist + percentage for the Business Dashboard progress banner.

    Each item has `done: bool`, `label`, `action_path` so the frontend can
    wire contextual CTAs straight to the right tab.
    """
    business = await db.businesses.find_one({"user_id": token_data.user_id}, {"_id": 0})
    if not business:
        raise HTTPException(status_code=404, detail="Business not found")
    bid = business["id"]

    has_cover = bool(business.get("cover_image_url")) or bool(business.get("logo_url"))
    photos_count = await db.business_photos.count_documents({"business_id": bid})
    services_priced = await db.services.count_documents({
        "business_id": bid, "active": {"$ne": False},
        "$or": [{"price": {"$gt": 0}}, {"price_from": {"$gt": 0}}],
    })
    workers_count = await db.workers.count_documents({"business_id": bid, "active": True})
    hours_set = bool(business.get("business_hours")) and any(
        (v or {}).get("open") and (v or {}).get("close")
        for v in (business.get("business_hours") or {}).values() if isinstance(v, dict)
    )
    docs_verified = bool(business.get("documents_verified", False))
    description_set = bool((business.get("description") or "").strip())

    items = [
        {"key": "cover",       "done": has_cover,               "label_es": "Foto de portada o logo",             "label_en": "Cover photo or logo",        "action_path": "photos"},
        {"key": "photos",      "done": photos_count >= 3,       "label_es": "Al menos 3 fotos del lugar",         "label_en": "At least 3 gallery photos",  "action_path": "photos"},
        {"key": "services",    "done": services_priced >= 1,    "label_es": "Servicios con precio",               "label_en": "Services with price",        "action_path": "services"},
        {"key": "description", "done": description_set,         "label_es": "Descripcion del negocio",            "label_en": "Business description",       "action_path": "overview"},
        {"key": "hours",       "done": hours_set,               "label_es": "Horarios de atencion",               "label_en": "Operating hours",            "action_path": "overview"},
        {"key": "team",        "done": workers_count >= 1,      "label_es": "Al menos 1 miembro del equipo",      "label_en": "At least 1 team member",     "action_path": "team"},
        {"key": "kyc",         "done": docs_verified,           "label_es": "Documentos KYC verificados",         "label_en": "KYC documents verified",     "action_path": "subscription"},
    ]
    done_count = sum(1 for i in items if i["done"])
    pct = round((done_count / len(items)) * 100)

    return {
        "percentage": pct,
        "done_count": done_count,
        "total_count": len(items),
        "items": items,
        "is_complete": pct == 100,
    }


@router.get("/my/reports")
async def get_business_reports(
    period: str = "month",
    token_data: TokenData = Depends(require_business)
):
    """Get business reports/analytics. period: week, month, quarter, year"""
    business = await db.businesses.find_one({"user_id": token_data.user_id}, {"_id": 0})
    if not business:
        raise HTTPException(status_code=404, detail="Business not found")
    
    business_id = business["id"]
    now = datetime.now(timezone.utc)
    
    period_days = {"week": 7, "month": 30, "quarter": 90, "year": 365}
    days = period_days.get(period, 30)
    start_date = now - timedelta(days=days)
    prev_start = start_date - timedelta(days=days)
    
    # Current period bookings
    current_bookings = await db.bookings.find(
        {"business_id": business_id, "created_at": {"$gte": start_date.isoformat()}},
        {"_id": 0}
    ).to_list(10000)
    
    # Previous period bookings (for comparison)
    prev_bookings = await db.bookings.find(
        {"business_id": business_id, "created_at": {"$gte": prev_start.isoformat(), "$lt": start_date.isoformat()}},
        {"_id": 0}
    ).to_list(10000)
    
    # Current period transactions (paid only)
    current_transactions = await db.transactions.find(
        {"business_id": business_id, "status": "paid", "created_at": {"$gte": start_date.isoformat()}},
        {"_id": 0}
    ).to_list(10000)
    
    prev_transactions = await db.transactions.find(
        {"business_id": business_id, "status": "paid", "created_at": {"$gte": prev_start.isoformat(), "$lt": start_date.isoformat()}},
        {"_id": 0}
    ).to_list(10000)
    
    # === BOOKING STATS ===
    completed = len([b for b in current_bookings if b.get("status") == "completed"])
    confirmed = len([b for b in current_bookings if b.get("status") == "confirmed"])
    cancelled = len([b for b in current_bookings if b.get("status") == "cancelled"])
    total_bookings = len(current_bookings)
    prev_total = len(prev_bookings)
    cancel_rate = round((cancelled / total_bookings * 100) if total_bookings > 0 else 0, 1)
    
    # === REVENUE ===
    current_revenue = sum(t.get("amount_total", 0) for t in current_transactions)
    prev_revenue = sum(t.get("amount_total", 0) for t in prev_transactions)
    revenue_change = round(((current_revenue - prev_revenue) / prev_revenue * 100) if prev_revenue > 0 else 0, 1)
    bookings_change = round(((total_bookings - prev_total) / prev_total * 100) if prev_total > 0 else 0, 1)
    
    # === REVENUE BY DAY (for chart) ===
    revenue_by_day = {}
    for t in current_transactions:
        day = t.get("created_at", "")[:10]
        if day:
            revenue_by_day[day] = revenue_by_day.get(day, 0) + t.get("amount_total", 0)
    
    bookings_by_day = {}
    for b in current_bookings:
        day = b.get("date", b.get("created_at", "")[:10])
        if day:
            bookings_by_day[day] = bookings_by_day.get(day, 0) + 1
    
    # Build daily chart data
    daily_chart = []
    for i in range(min(days, 30)):
        d = (now - timedelta(days=days - 1 - i)).strftime("%Y-%m-%d")
        daily_chart.append({
            "date": d,
            "revenue": round(revenue_by_day.get(d, 0), 2),
            "bookings": bookings_by_day.get(d, 0)
        })
    
    # === TOP SERVICES ===
    service_counts = {}
    service_revenue = {}
    for b in current_bookings:
        sid = b.get("service_id")
        if sid:
            service_counts[sid] = service_counts.get(sid, 0) + 1
            if b.get("status") in ("completed", "confirmed"):
                service_revenue[sid] = service_revenue.get(sid, 0) + (b.get("total_amount") or b.get("deposit_amount") or 0)
    
    service_ids = list(set(list(service_counts.keys()) + list(service_revenue.keys())))
    services_map = {}
    if service_ids:
        svcs = await db.services.find({"id": {"$in": service_ids}}, {"_id": 0, "id": 1, "name": 1}).to_list(100)
        services_map = {s["id"]: s["name"] for s in svcs}
    
    top_services = sorted(
        [{"name": services_map.get(sid, "Servicio"), "bookings": service_counts.get(sid, 0), "revenue": round(service_revenue.get(sid, 0), 2)}
         for sid in service_ids],
        key=lambda x: x["bookings"], reverse=True
    )[:5]
    
    # === TOP CLIENTS ===
    client_visits = {}
    client_spent = {}
    for b in current_bookings:
        uid = b.get("user_id")
        if uid and b.get("status") in ("completed", "confirmed"):
            client_visits[uid] = client_visits.get(uid, 0) + 1
            client_spent[uid] = client_spent.get(uid, 0) + (b.get("total_amount") or b.get("deposit_amount") or 0)
    
    top_client_ids = sorted(client_visits.keys(), key=lambda x: client_visits[x], reverse=True)[:5]
    top_clients = []
    for uid in top_client_ids:
        u = await db.users.find_one({"id": uid}, {"_id": 0, "full_name": 1, "email": 1})
        top_clients.append({
            "name": u.get("full_name", "Cliente") if u else "Cliente",
            "email": u.get("email", "") if u else "",
            "visits": client_visits[uid],
            "total_spent": round(client_spent.get(uid, 0), 2)
        })
    
    # === PEAK HOURS ===
    hour_counts = {}
    for b in current_bookings:
        t = b.get("time", "")
        if t:
            hour = t.split(":")[0]
            hour_counts[hour] = hour_counts.get(hour, 0) + 1
    
    peak_hours = sorted(
        [{"hour": f"{h}:00", "bookings": c} for h, c in hour_counts.items()],
        key=lambda x: x["bookings"], reverse=True
    )[:6]
    
    # === PEAK DAYS ===
    day_names_es = ["Lun", "Mar", "Mie", "Jue", "Vie", "Sab", "Dom"]
    day_counts = {i: 0 for i in range(7)}
    for b in current_bookings:
        d = b.get("date", "")
        if d:
            try:
                weekday = datetime.strptime(d, "%Y-%m-%d").weekday()
                day_counts[weekday] += 1
            except ValueError:
                pass
    
    peak_days = [{"day": day_names_es[i], "bookings": day_counts[i]} for i in range(7)]
    
    # === CANCELLATION BREAKDOWN ===
    cancelled_by_user = len([b for b in current_bookings if b.get("status") == "cancelled" and b.get("cancelled_by") == "user"])
    cancelled_by_business = len([b for b in current_bookings if b.get("status") == "cancelled" and b.get("cancelled_by") == "business"])
    
    return {
        "period": period,
        "summary": {
            "total_bookings": total_bookings,
            "completed": completed,
            "confirmed": confirmed,
            "cancelled": cancelled,
            "cancel_rate": cancel_rate,
            "revenue": round(current_revenue, 2),
            "revenue_change": revenue_change,
            "bookings_change": bookings_change,
            "cancelled_by_user": cancelled_by_user,
            "cancelled_by_business": cancelled_by_business,
        },
        "daily_chart": daily_chart,
        "top_services": top_services,
        "top_clients": top_clients,
        "peak_hours": peak_hours,
        "peak_days": peak_days,
    }





@router.get("/my/client-history/{user_id}")
async def get_client_history(user_id: str, token_data: TokenData = Depends(require_business)):
    """Get client history within this business"""
    business = await db.businesses.find_one({"user_id": token_data.user_id}, {"_id": 0, "id": 1})
    if not business:
        raise HTTPException(status_code=404, detail="Business not found")
    
    bookings = await db.bookings.find(
        {"business_id": business["id"], "user_id": user_id},
        {"_id": 0, "id": 1, "date": 1, "time": 1, "status": 1, "service_id": 1, "total_amount": 1, "deposit_amount": 1, "cancelled_by": 1}
    ).sort("date", -1).to_list(50)
    
    # Get service names
    service_ids = list(set(b.get("service_id") for b in bookings if b.get("service_id")))
    services_map = {}
    if service_ids:
        for s in await db.services.find({"id": {"$in": service_ids}}, {"_id": 0, "id": 1, "name": 1}).to_list(100):
            services_map[s["id"]] = s["name"]
    
    completed = [b for b in bookings if b.get("status") in ("completed", "confirmed")]
    cancelled = [b for b in bookings if b.get("status") == "cancelled"]
    total_spent = sum(b.get("total_amount") or b.get("deposit_amount") or 0 for b in completed)
    
    history = []
    for b in bookings[:5]:
        history.append({
            "date": b.get("date", ""),
            "time": b.get("time", ""),
            "service_name": services_map.get(b.get("service_id"), ""),
            "status": b.get("status", ""),
            "amount": b.get("total_amount") or b.get("deposit_amount") or 0,
            "cancelled_by": b.get("cancelled_by"),
        })
    
    return {
        "total_visits": len(completed),
        "total_cancelled": len(cancelled),
        "total_spent": round(total_spent, 2),
        "first_visit": bookings[-1]["date"] if bookings else None,
        "last_visit": bookings[0]["date"] if bookings else None,
        "history": history,
    }


# ========================== PHASE 15: MINI-CRM (MY CLIENTS) ==========================


@router.get("/my/clients")
async def list_my_clients(
    q: str = "",
    tag: Optional[str] = None,  # "vip" | "new" | "noshow" | "inactive"
    sort: str = "recent",       # "recent" | "visits" | "spent" | "name"
    page: int = 1,
    limit: int = 50,
    token_data: TokenData = Depends(require_business),
):
    """Mini-CRM: aggregated client list with last visit, total visits,
    total spent, no-show count and a private note. Powers the new
    "Mis Clientes" tab on the Business Dashboard.

    Rules:
      * only clients who have at least one booking with THIS business
      * clients without user_id (walk-ins booked by recepcion) are still
        surfaced using their stored name/phone/email
      * `tag=vip` means 5+ completed visits, `tag=new` means 1 visit only,
        `tag=noshow` means >=2 no-shows, `tag=inactive` >= 90d since last
    """
    business = await db.businesses.find_one({"user_id": token_data.user_id}, {"_id": 0, "id": 1})
    if not business:
        raise HTTPException(status_code=404, detail="Business not found")
    bid = business["id"]

    pipeline = [
        {"$match": {"business_id": bid}},
        {"$addFields": {
            "_client_key": {
                "$ifNull": [
                    "$user_id",
                    {"$concat": [
                        {"$ifNull": ["$client_phone", ""]},
                        "|",
                        {"$toLower": {"$ifNull": ["$client_name", ""]}},
                    ]},
                ]
            },
            "_is_completed": {"$in": ["$status", ["completed", "confirmed"]]},
            "_is_noshow": {"$eq": ["$status", "no_show"]},
            "_paid_amount": {
                "$cond": [
                    {"$in": ["$status", ["completed", "confirmed"]]},
                    {"$ifNull": ["$total_amount", {"$ifNull": ["$deposit_amount", 0]}]},
                    0,
                ]
            },
        }},
        {"$group": {
            "_id": "$_client_key",
            "user_id": {"$first": "$user_id"},
            "fallback_name": {"$first": "$client_name"},
            "fallback_phone": {"$first": "$client_phone"},
            "fallback_email": {"$first": "$client_email"},
            "total_bookings": {"$sum": 1},
            "total_visits": {"$sum": {"$cond": ["$_is_completed", 1, 0]}},
            "noshow_count": {"$sum": {"$cond": ["$_is_noshow", 1, 0]}},
            "total_spent": {"$sum": "$_paid_amount"},
            "first_visit": {"$min": "$date"},
            "last_visit": {"$max": "$date"},
        }},
        {"$match": {"_id": {"$ne": None}}},
    ]
    rows = await db.bookings.aggregate(pipeline).to_list(5000)

    # Resolve user profile data and attach private note
    user_ids = [r["user_id"] for r in rows if r.get("user_id")]
    users_map = {}
    if user_ids:
        async for u in db.users.find(
            {"id": {"$in": user_ids}},
            {"_id": 0, "id": 1, "full_name": 1, "email": 1, "phone": 1, "profile_image_url": 1, "public_code": 1},
        ):
            users_map[u["id"]] = u

    notes_map = {}
    async for n in db.business_client_notes.find({"business_id": bid}, {"_id": 0}):
        notes_map[n.get("client_key")] = n

    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    ninety_ago = (datetime.now(timezone.utc) - timedelta(days=90)).strftime("%Y-%m-%d")

    clients = []
    for r in rows:
        uid = r.get("user_id")
        profile = users_map.get(uid) or {}
        name = profile.get("full_name") or r.get("fallback_name") or "-"
        email = profile.get("email") or r.get("fallback_email") or ""
        phone = profile.get("phone") or r.get("fallback_phone") or ""
        note = notes_map.get(r["_id"], {})
        tags = []
        if r["total_visits"] >= 5:
            tags.append("vip")
        if r["total_visits"] == 1 and r["noshow_count"] == 0:
            tags.append("new")
        if r["noshow_count"] >= 2:
            tags.append("noshow")
        if r.get("last_visit") and r["last_visit"] < ninety_ago:
            tags.append("inactive")

        clients.append({
            "client_key": r["_id"],
            "user_id": uid,
            "name": name,
            "email": email,
            "phone": phone,
            "public_code": profile.get("public_code") or None,
            "avatar_url": profile.get("profile_image_url"),
            "is_registered": bool(uid),
            "total_bookings": r["total_bookings"],
            "total_visits": r["total_visits"],
            "noshow_count": r["noshow_count"],
            "total_spent": round(float(r.get("total_spent") or 0), 2),
            "first_visit": r.get("first_visit"),
            "last_visit": r.get("last_visit"),
            "days_since_last": None if not r.get("last_visit") else (
                (datetime.strptime(today, "%Y-%m-%d") - datetime.strptime(r["last_visit"], "%Y-%m-%d")).days
            ),
            "private_note": note.get("note", ""),
            "note_updated_at": note.get("updated_at"),
            "tags": tags,
        })

    # Search filter (name / phone / email / public_code contains q)
    if q:
        ql = q.lower().strip()
        clients = [c for c in clients if
                   ql in (c["name"] or "").lower()
                   or ql in (c["phone"] or "").lower()
                   or ql in (c["email"] or "").lower()
                   or ql in (c.get("public_code") or "").lower()]

    # Tag filter
    if tag:
        clients = [c for c in clients if tag in c["tags"]]

    # Sorting
    if sort == "recent":
        clients.sort(key=lambda c: c["last_visit"] or "", reverse=True)
    elif sort == "visits":
        clients.sort(key=lambda c: c["total_visits"], reverse=True)
    elif sort == "spent":
        clients.sort(key=lambda c: c["total_spent"], reverse=True)
    elif sort == "name":
        clients.sort(key=lambda c: (c["name"] or "").lower())

    total = len(clients)
    page = max(1, page)
    limit = max(1, min(limit, 200))
    start = (page - 1) * limit
    page_rows = clients[start:start + limit]

    # Aggregate KPIs for the header
    vip_count = sum(1 for c in clients if "vip" in c["tags"])
    new_count = sum(1 for c in clients if "new" in c["tags"])
    inactive_count = sum(1 for c in clients if "inactive" in c["tags"])

    return {
        "total": total,
        "page": page,
        "limit": limit,
        "kpis": {
            "total_clients": total,
            "vip": vip_count,
            "new": new_count,
            "inactive": inactive_count,
        },
        "items": page_rows,
    }


class ClientNoteUpdate(BaseModel):
    note: str


@router.put("/my/clients/{client_key:path}/note")
async def update_client_note(
    client_key: str,
    payload: ClientNoteUpdate,
    token_data: TokenData = Depends(require_business),
):
    """Upsert a private note for a specific client (max 500 chars)."""
    business = await db.businesses.find_one({"user_id": token_data.user_id}, {"_id": 0, "id": 1})
    if not business:
        raise HTTPException(status_code=404, detail="Business not found")

    note = (payload.note or "").strip()[:500]
    now_iso = datetime.now(timezone.utc).isoformat()
    await db.business_client_notes.update_one(
        {"business_id": business["id"], "client_key": client_key},
        {"$set": {"note": note, "updated_at": now_iso},
         "$setOnInsert": {"created_at": now_iso}},
        upsert=True,
    )
    return {"ok": True, "note": note, "updated_at": now_iso}


@router.post("/my/clients/export")
async def export_my_clients(token_data: TokenData = Depends(require_business)):
    """Return plain CSV of the entire client list for the business (data
    portability — Fase LFPDPPP). The frontend downloads the file."""
    from fastapi.responses import Response as _Response
    import csv as _csv
    import io as _io

    result = await list_my_clients(token_data=token_data, limit=1000)
    buf = _io.StringIO()
    w = _csv.writer(buf)
    w.writerow(["bookvia_code", "name", "email", "phone", "total_visits", "total_spent",
                "last_visit", "noshow_count", "tags", "private_note"])
    for c in result["items"]:
        w.writerow([
            c.get("public_code") or "",
            c["name"], c["email"], c["phone"], c["total_visits"],
            c["total_spent"], c["last_visit"] or "",
            c["noshow_count"], ",".join(c["tags"]), c["private_note"],
        ])
    return _Response(
        content=buf.getvalue(),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=clientes.csv"},
    )


@router.get("/my/clients/lookup")
async def lookup_client_by_code(
    code: str,
    token_data: TokenData = Depends(require_business),
):
    """Lookup a single client by their public Bookvia code (CL-XXXXX).
    Returns aggregated stats scoped to THIS business only — walk-in /
    unregistered clients are excluded (no user_id means no code)."""
    from services.public_code import normalize_public_code, is_valid_public_code

    normalized = normalize_public_code(code)
    if not is_valid_public_code(normalized) or not normalized.startswith("CL-"):
        raise HTTPException(status_code=400, detail="Código inválido. Formato esperado: CL-XXXXX")

    business = await db.businesses.find_one({"user_id": token_data.user_id}, {"_id": 0, "id": 1})
    if not business:
        raise HTTPException(status_code=404, detail="Business not found")

    user = await db.users.find_one(
        {"public_code": normalized},
        {"_id": 0, "id": 1, "full_name": 1, "email": 1, "phone": 1,
         "profile_image_url": 1, "public_code": 1},
    )
    if not user:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")

    # Aggregate this business's history with the client
    bookings = await db.bookings.find(
        {"business_id": business["id"], "user_id": user["id"]},
        {"_id": 0, "id": 1, "date": 1, "status": 1, "total_amount": 1, "deposit_amount": 1},
    ).sort("date", -1).to_list(500)

    total_visits = sum(1 for b in bookings if b.get("status") in ("completed", "confirmed"))
    noshow_count = sum(1 for b in bookings if b.get("status") == "no_show")
    total_spent = sum(
        float(b.get("total_amount") or b.get("deposit_amount") or 0)
        for b in bookings if b.get("status") in ("completed", "confirmed")
    )
    last_visit = bookings[0]["date"] if bookings else None
    has_history = len(bookings) > 0

    return {
        "found": True,
        "has_history_with_you": has_history,
        "public_code": user["public_code"],
        "name": user.get("full_name") or "-",
        "email": user.get("email"),
        "phone": user.get("phone"),
        "avatar_url": user.get("profile_image_url"),
        "total_bookings": len(bookings),
        "total_visits": total_visits,
        "noshow_count": noshow_count,
        "total_spent": round(total_spent, 2),
        "last_visit": last_visit,
    }




@router.get("/my/reports/export")
async def export_business_reports(
    period: str = "month",
    token_data: TokenData = Depends(require_business)
):
    """Export business reports as Excel file"""
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import io
    
    business = await db.businesses.find_one({"user_id": token_data.user_id}, {"_id": 0})
    if not business:
        raise HTTPException(status_code=404, detail="Business not found")
    
    business_id = business["id"]
    now = datetime.now(timezone.utc)
    period_days = {"week": 7, "month": 30, "quarter": 90, "year": 365}
    days = period_days.get(period, 30)
    start_date = now - timedelta(days=days)
    period_labels = {"week": "7 dias", "month": "30 dias", "quarter": "90 dias", "year": "1 ano"}
    
    bookings = await db.bookings.find(
        {"business_id": business_id, "created_at": {"$gte": start_date.isoformat()}},
        {"_id": 0}
    ).to_list(10000)
    
    transactions = await db.transactions.find(
        {"business_id": business_id, "status": "paid", "created_at": {"$gte": start_date.isoformat()}},
        {"_id": 0}
    ).to_list(10000)
    
    # Build lookup maps
    service_ids = list(set(b.get("service_id") for b in bookings if b.get("service_id")))
    user_ids = list(set(b.get("user_id") for b in bookings if b.get("user_id")))
    worker_ids = list(set(b.get("worker_id") for b in bookings if b.get("worker_id")))
    
    services_map, users_map, workers_map = {}, {}, {}
    if service_ids:
        for s in await db.services.find({"id": {"$in": service_ids}}, {"_id": 0, "id": 1, "name": 1}).to_list(100):
            services_map[s["id"]] = s["name"]
    if user_ids:
        for u in await db.users.find({"id": {"$in": user_ids}}, {"_id": 0, "id": 1, "full_name": 1, "email": 1}).to_list(500):
            users_map[u["id"]] = u
    if worker_ids:
        for w in await db.workers.find({"id": {"$in": worker_ids}}, {"_id": 0, "id": 1, "name": 1}).to_list(100):
            workers_map[w["id"]] = w["name"]
    
    # Create workbook
    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'), right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'), bottom=Side(style='thin', color='D0D0D0')
    )
    
    def style_header(ws, cols):
        for col_idx, title in enumerate(cols, 1):
            cell = ws.cell(row=1, column=col_idx, value=title)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
    
    # === Sheet 1: Resumen ===
    ws1 = wb.active
    ws1.title = "Resumen"
    completed = len([b for b in bookings if b.get("status") == "completed"])
    confirmed = len([b for b in bookings if b.get("status") == "confirmed"])
    cancelled = len([b for b in bookings if b.get("status") == "cancelled"])
    total_rev = sum(t.get("amount_total", 0) for t in transactions)
    
    summary_data = [
        ("Negocio", business["name"]),
        ("Periodo", period_labels.get(period, period)),
        ("Fecha del reporte", now.strftime("%Y-%m-%d %H:%M")),
        ("", ""),
        ("Total de citas", len(bookings)),
        ("Completadas", completed),
        ("Confirmadas", confirmed),
        ("Canceladas", cancelled),
        ("Tasa de cancelacion", f"{round((cancelled / len(bookings) * 100) if bookings else 0, 1)}%"),
        ("Ingresos totales", f"${round(total_rev, 2)}"),
    ]
    for row_idx, (label, value) in enumerate(summary_data, 1):
        ws1.cell(row=row_idx, column=1, value=label).font = Font(bold=True) if label else Font()
        ws1.cell(row=row_idx, column=2, value=value)
    ws1.column_dimensions['A'].width = 25
    ws1.column_dimensions['B'].width = 30
    
    # === Sheet 2: Citas ===
    ws2 = wb.create_sheet("Citas")
    cols = ["Fecha", "Hora", "Cliente", "Email", "Servicio", "Trabajador", "Estado", "Monto", "Cancelado por"]
    style_header(ws2, cols)
    
    status_map = {"confirmed": "Confirmada", "completed": "Completada", "cancelled": "Cancelada", "pending": "Pendiente"}
    for row_idx, b in enumerate(sorted(bookings, key=lambda x: x.get("date", ""), reverse=True), 2):
        u = users_map.get(b.get("user_id"), {})
        ws2.cell(row=row_idx, column=1, value=b.get("date", "")).border = thin_border
        ws2.cell(row=row_idx, column=2, value=b.get("time", "")).border = thin_border
        ws2.cell(row=row_idx, column=3, value=u.get("full_name", b.get("client_name", ""))).border = thin_border
        ws2.cell(row=row_idx, column=4, value=u.get("email", b.get("client_email", ""))).border = thin_border
        ws2.cell(row=row_idx, column=5, value=services_map.get(b.get("service_id"), "")).border = thin_border
        ws2.cell(row=row_idx, column=6, value=workers_map.get(b.get("worker_id"), "")).border = thin_border
        ws2.cell(row=row_idx, column=7, value=status_map.get(b.get("status"), b.get("status", ""))).border = thin_border
        ws2.cell(row=row_idx, column=8, value=b.get("total_amount") or b.get("deposit_amount") or 0).border = thin_border
        ws2.cell(row=row_idx, column=9, value=b.get("cancelled_by", "")).border = thin_border
    
    for col in ['A','B','C','D','E','F','G','H','I']:
        ws2.column_dimensions[col].width = 18
    
    # === Sheet 3: Servicios ===
    ws3 = wb.create_sheet("Servicios")
    style_header(ws3, ["Servicio", "Total citas", "Ingresos"])
    svc_counts, svc_rev = {}, {}
    for b in bookings:
        sid = b.get("service_id")
        if sid:
            svc_counts[sid] = svc_counts.get(sid, 0) + 1
            if b.get("status") in ("completed", "confirmed"):
                svc_rev[sid] = svc_rev.get(sid, 0) + (b.get("total_amount") or b.get("deposit_amount") or 0)
    for row_idx, sid in enumerate(sorted(svc_counts.keys(), key=lambda x: svc_counts[x], reverse=True), 2):
        ws3.cell(row=row_idx, column=1, value=services_map.get(sid, "Servicio")).border = thin_border
        ws3.cell(row=row_idx, column=2, value=svc_counts[sid]).border = thin_border
        ws3.cell(row=row_idx, column=3, value=round(svc_rev.get(sid, 0), 2)).border = thin_border
    for col in ['A','B','C']:
        ws3.column_dimensions[col].width = 25
    
    # === Sheet 4: Clientes ===
    ws4 = wb.create_sheet("Clientes")
    style_header(ws4, ["Cliente", "Email", "Visitas", "Total gastado"])
    client_visits, client_spent = {}, {}
    for b in bookings:
        uid = b.get("user_id")
        if uid and b.get("status") in ("completed", "confirmed"):
            client_visits[uid] = client_visits.get(uid, 0) + 1
            client_spent[uid] = client_spent.get(uid, 0) + (b.get("total_amount") or b.get("deposit_amount") or 0)
    for row_idx, uid in enumerate(sorted(client_visits.keys(), key=lambda x: client_visits[x], reverse=True), 2):
        u = users_map.get(uid, {})
        ws4.cell(row=row_idx, column=1, value=u.get("full_name", "Cliente")).border = thin_border
        ws4.cell(row=row_idx, column=2, value=u.get("email", "")).border = thin_border
        ws4.cell(row=row_idx, column=3, value=client_visits[uid]).border = thin_border
        ws4.cell(row=row_idx, column=4, value=round(client_spent.get(uid, 0), 2)).border = thin_border
    for col in ['A','B','C','D']:
        ws4.column_dimensions[col].width = 25
    
    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"Reporte_{business['name'].replace(' ', '_')}_{period}_{now.strftime('%Y%m%d')}.xlsx"
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )





@router.get("/by-code/{code}", response_model=BusinessResponse)
async def get_business_by_public_code(code: str):
    """Look up a business by its public Bookvia code (BV-XXXXX). Used for QR codes and short URLs."""
    from services.public_code import normalize_public_code, is_valid_public_code
    normalized = normalize_public_code(code)
    if not is_valid_public_code(normalized):
        raise HTTPException(status_code=400, detail="Codigo invalido")
    business = await db.businesses.find_one(
        {"public_code": normalized, "status": BusinessStatus.APPROVED},
        {"_id": 0, "password_hash": 0}
    )
    if not business:
        raise HTTPException(status_code=404, detail="Negocio no encontrado")
    business.setdefault("description", "")
    business.setdefault("address", "")
    business.setdefault("city", "")
    business.setdefault("state", "")
    business.setdefault("country", "MX")
    business.setdefault("zip_code", "")
    business.setdefault("subscription_status", "none")
    return BusinessResponse(**business)



@router.get("", response_model=List[BusinessResponse])
async def search_businesses(
    request: Request,
    query: Optional[str] = None,
    category_id: Optional[str] = None,
    city: Optional[str] = None,
    country_code: Optional[str] = None,
    min_rating: Optional[float] = None,
    is_home_service: Optional[bool] = None,
    include_pending: bool = False,
    user_lat: Optional[float] = None,
    user_lng: Optional[float] = None,
    sort: Optional[str] = None,
    page: int = 1,
    limit: int = 20,
    current_user: Optional[TokenData] = Depends(get_current_user)
):
    # By default only show approved businesses with active subscription (or trialing)
    # If include_pending=True, also show PENDING (for profile viewing, but no bookings)
    if include_pending:
        filters = {"status": {"$in": [BusinessStatus.APPROVED, BusinessStatus.PENDING]}}
    else:
        filters = visible_business_filter_now()
    
    if country_code:
        filters["country_code"] = country_code.upper()
    if query and not category_id:
        # Also search by matching category names
        matching_cats = await db.categories.find(
            {"$or": [
                {"name_es": {"$regex": query, "$options": "i"}},
                {"name_en": {"$regex": query, "$options": "i"}}
            ]},
            {"_id": 0, "id": 1}
        ).to_list(50)
        matching_cat_ids = [c["id"] for c in matching_cats]
        
        or_conditions = [
            {"name": {"$regex": query, "$options": "i"}},
            {"description": {"$regex": query, "$options": "i"}}
        ]
        if matching_cat_ids:
            or_conditions.append({"category_id": {"$in": matching_cat_ids}})
        filters["$or"] = or_conditions
    elif query:
        filters["$or"] = [
            {"name": {"$regex": query, "$options": "i"}},
            {"description": {"$regex": query, "$options": "i"}}
        ]
    if category_id:
        filters["category_id"] = category_id
    if city:
        # When sorting by nearest, don't filter by city (geographic proximity takes priority)
        if sort != "nearest":
            filters["city"] = {"$regex": city, "$options": "i"}
    if min_rating:
        filters["rating"] = {"$gte": min_rating}
    if is_home_service is not None:
        filters["service_radius_km"] = {"$exists": True, "$ne": None} if is_home_service else {"$in": [None, 0]}
    
    skip = (page - 1) * limit
    # When sorting by nearest, increase limit to show more nearby businesses
    effective_limit = limit * 3 if (sort == "nearest" and user_lat is not None) else limit
    businesses = await db.businesses.find(
        filters,
        {"_id": 0, "password_hash": 0, "clabe": 0, "rfc": 0, "ine_url": 0, "proof_of_address_url": 0}
    ).sort("rating", -1).skip(skip).limit(effective_limit).to_list(effective_limit)
    
    # Add category names and booking availability
    for b in businesses:
        if b.get("category_id"):
            cat = await db.categories.find_one({"id": b["category_id"]})
            if cat:
                b["category_name"] = cat.get("name_es", "")
        # Mark if business can accept bookings (must be approved + subscription active/trialing or legacy + documents verified)
        sub_status = b.get("subscription_status", "none")
        b["can_accept_bookings"] = (
            b.get("status") == BusinessStatus.APPROVED
            and (sub_status in ("active", "trialing", "none", None) or sub_status is None)
            and bool(b.get("documents_verified", False))
        )
    
    # Filter out businesses where the current user is blacklisted
    if current_user:
        blacklisted_biz_ids = set()
        for b in businesses:
            if await is_user_blacklisted(b["id"], user_id=current_user.user_id):
                blacklisted_biz_ids.add(b["id"])
        businesses = [b for b in businesses if b["id"] not in blacklisted_biz_ids]
    
    # Calculate distance and sort by proximity if user location provided
    if user_lat is not None and user_lng is not None:
        import math
        def haversine(lat1, lng1, lat2, lng2):
            R = 6371  # km
            dlat = math.radians(lat2 - lat1)
            dlng = math.radians(lng2 - lng1)
            a = math.sin(dlat/2)**2 + math.cos(math.radians(lat1)) * math.cos(math.radians(lat2)) * math.sin(dlng/2)**2
            return R * 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
        
        for b in businesses:
            if b.get("latitude") and b.get("longitude"):
                b["distance_km"] = round(haversine(user_lat, user_lng, b["latitude"], b["longitude"]), 1)
            else:
                b["distance_km"] = None
        
        if sort == "nearest":
            businesses.sort(key=lambda x: x.get("distance_km") if x.get("distance_km") is not None else 99999)
    
    # Compute next available text based on worker schedules
    biz_ids = [b["id"] for b in businesses]
    if biz_ids:
        all_workers = await db.workers.find(
            {"business_id": {"$in": biz_ids}, "active": True},
            {"_id": 0, "business_id": 1, "schedule": 1}
        ).to_list(500)
        biz_workers_map = {}
        for w in all_workers:
            biz_workers_map.setdefault(w["business_id"], []).append(w)
        
        day_names_es = ["Lun", "Mar", "Mie", "Jue", "Vie", "Sab", "Dom"]
        try:
            now = datetime.now(pytz.timezone("America/Mexico_City"))
        except Exception:
            now = datetime.now(timezone.utc)
        current_weekday = now.weekday()
        current_time = now.strftime("%H:%M")
        
        for b in businesses:
            workers_list = biz_workers_map.get(b["id"], [])
            if not workers_list:
                continue
            # Check if open right now
            is_open = False
            for worker in workers_list:
                ds = worker.get("schedule", {}).get(str(current_weekday), {})
                if ds.get("is_available") and ds.get("blocks"):
                    for block in ds["blocks"]:
                        if block["start_time"] <= current_time < block["end_time"]:
                            is_open = True
                            break
                if is_open:
                    break
            b["is_open_now"] = is_open
            
            # Find next available text
            found = False
            for day_offset in range(7):
                check_day = (current_weekday + day_offset) % 7
                day_str = str(check_day)
                for worker in workers_list:
                    schedule = worker.get("schedule", {})
                    day_schedule = schedule.get(day_str, {})
                    if day_schedule.get("is_available") and day_schedule.get("blocks"):
                        for block in day_schedule["blocks"]:
                            if day_offset == 0:
                                if block["end_time"] > current_time:
                                    b["next_available_text"] = "Hoy disponible"
                                    found = True
                                    break
                            else:
                                if day_offset == 1:
                                    b["next_available_text"] = f"Manana {block['start_time']}"
                                else:
                                    target_day = (current_weekday + day_offset) % 7
                                    b["next_available_text"] = f"{day_names_es[target_day]} {block['start_time']}"
                                found = True
                                break
                        if found:
                            break
                    if found:
                        break
                if found:
                    break
    
    # Fetch top services and min price per business (single bulk query)
    if biz_ids:
        all_services = await db.services.find(
            {"business_id": {"$in": biz_ids}, "active": {"$ne": False}},
            {"_id": 0, "id": 1, "business_id": 1, "name": 1, "price": 1}
        ).sort("price", 1).to_list(2000)
        biz_services_map = {}
        for s in all_services:
            biz_services_map.setdefault(s["business_id"], []).append(s)
        
        for b in businesses:
            services = biz_services_map.get(b["id"], [])
            if services:
                b["top_services"] = [
                    {"name": s.get("name", ""), "price": s.get("price", 0)}
                    for s in services[:3]
                ]
                prices = [s.get("price", 0) for s in services if s.get("price")]
                if prices:
                    b["min_price"] = min(prices)
            else:
                b["top_services"] = []
    
    return [BusinessResponse(**b) for b in businesses]



@router.get("/featured", response_model=List[BusinessResponse])
async def get_featured_businesses(limit: int = 8, country_code: Optional[str] = None, current_user: Optional[TokenData] = Depends(get_current_user)):
    base_filter = {**visible_business_filter_now(), "is_featured": True}
    if country_code:
        base_filter["country_code"] = country_code.upper()
    businesses = await db.businesses.find(
        base_filter,
        {"_id": 0, "password_hash": 0, "clabe": 0, "rfc": 0}
    ).sort("rating", -1).limit(limit).to_list(limit)
    
    # If not enough featured, add top rated
    if len(businesses) < limit:
        existing_ids = [b["id"] for b in businesses]
        more_filter = {**visible_business_filter_now(), "id": {"$nin": existing_ids}}
        if country_code:
            more_filter["country_code"] = country_code.upper()
        more = await db.businesses.find(
            more_filter,
            {"_id": 0, "password_hash": 0, "clabe": 0, "rfc": 0}
        ).sort("rating", -1).limit(limit - len(businesses)).to_list(limit - len(businesses))
        businesses.extend(more)
    
    # Filter out businesses where the current user is blacklisted
    if current_user:
        blacklisted_biz_ids = set()
        for b in businesses:
            if await is_user_blacklisted(b["id"], user_id=current_user.user_id):
                blacklisted_biz_ids.add(b["id"])
        businesses = [b for b in businesses if b["id"] not in blacklisted_biz_ids]
    
    # Compute next available text for featured businesses
    biz_ids = [b["id"] for b in businesses]
    if biz_ids:
        all_workers = await db.workers.find(
            {"business_id": {"$in": biz_ids}, "active": True},
            {"_id": 0, "business_id": 1, "schedule": 1}
        ).to_list(500)
        biz_workers_map = {}
        for w in all_workers:
            biz_workers_map.setdefault(w["business_id"], []).append(w)
        day_names_es = ["Lun", "Mar", "Mie", "Jue", "Vie", "Sab", "Dom"]
        try:
            now_ft = datetime.now(pytz.timezone("America/Mexico_City"))
        except Exception:
            now_ft = datetime.now(timezone.utc)
        current_wd = now_ft.weekday()
        current_tm = now_ft.strftime("%H:%M")
        for b in businesses:
            wl = biz_workers_map.get(b["id"], [])
            if not wl:
                continue
            # Check if open right now
            is_open_ft = False
            for wk in wl:
                ds_now = wk.get("schedule", {}).get(str(current_wd), {})
                if ds_now.get("is_available") and ds_now.get("blocks"):
                    for blk in ds_now["blocks"]:
                        if blk["start_time"] <= current_tm < blk["end_time"]:
                            is_open_ft = True
                            break
                if is_open_ft:
                    break
            b["is_open_now"] = is_open_ft
            
            found = False
            for doff in range(7):
                cd = (current_wd + doff) % 7
                for wk in wl:
                    ds = wk.get("schedule", {}).get(str(cd), {})
                    if ds.get("is_available") and ds.get("blocks"):
                        for blk in ds["blocks"]:
                            if doff == 0 and blk["end_time"] > current_tm:
                                b["next_available_text"] = "Hoy disponible"
                                found = True
                                break
                            elif doff > 0:
                                b["next_available_text"] = f"{'Manana' if doff == 1 else day_names_es[(current_wd + doff) % 7]} {blk['start_time']}"
                                found = True
                                break
                        if found:
                            break
                    if found:
                        break
                if found:
                    break
    
    return [BusinessResponse(**b) for b in businesses]



@router.get("/slug/{slug}", response_model=BusinessResponse)
async def get_business_by_slug(
    slug: str,
    request: Request,
    current_user: Optional[TokenData] = Depends(get_current_user),
):
    """Get business by slug or ID (fallback)."""
    business = await db.businesses.find_one(
        {"slug": slug},
        {"_id": 0, "password_hash": 0, "clabe": 0, "rfc": 0, "ine_url": 0, "proof_of_address_url": 0}
    )
    if not business:
        # Fallback: try finding by ID
        business = await db.businesses.find_one(
            {"id": slug},
            {"_id": 0, "password_hash": 0, "clabe": 0, "rfc": 0, "ine_url": 0, "proof_of_address_url": 0}
        )
    if not business:
        raise HTTPException(status_code=404, detail="Business not found")
    
    # Check blacklist
    if current_user and await is_user_blacklisted(business["id"], user_id=current_user.user_id):
        raise HTTPException(status_code=404, detail="Business not found")
    
    if business.get("category_id"):
        cat = await db.categories.find_one({"id": business["category_id"]})
        if cat:
            business["category_name"] = cat.get("name_es", "")
    
    await _track_profile_view(business["id"], request, current_user, business.get("user_id"))
    return BusinessResponse(**business)



@router.get("/{business_id}", response_model=BusinessResponse)
async def get_business(
    business_id: str,
    request: Request,
    current_user: Optional[TokenData] = Depends(get_current_user),
):
    business = await db.businesses.find_one(
        {"id": business_id},
        {"_id": 0, "password_hash": 0, "clabe": 0, "rfc": 0, "ine_url": 0, "proof_of_address_url": 0}
    )
    if not business:
        raise HTTPException(status_code=404, detail="Business not found")
    # Check blacklist
    if current_user and await is_user_blacklisted(business["id"], user_id=current_user.user_id):
        raise HTTPException(status_code=404, detail="Business not found")
    await _track_profile_view(business["id"], request, current_user, business.get("user_id"))
    return BusinessResponse(**business)


@router.get("/{business_id}/trust-score")
async def get_business_trust_score(business_id: str):
    """Public: composite trust score (rating + completion + strikes) for the business profile badge."""
    business = await db.businesses.find_one({"id": business_id}, {"_id": 0})
    if not business:
        raise HTTPException(status_code=404, detail="Business not found")
    from services.strikes import compute_trust_score
    return compute_trust_score(business)



@router.put("/me", response_model=BusinessResponse)
async def update_my_business(update: BusinessUpdate, token_data: TokenData = Depends(require_business)):
    user = await db.users.find_one({"id": token_data.user_id})
    if not user or not user.get("business_id"):
        raise HTTPException(status_code=404, detail="Business not found")
    
    update_data = {k: v for k, v in update.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="No data to update")
    
    if "deposit_amount" in update_data:
        # If also disabling requires_deposit in the same update, allow zero
        requires_dep_in_update = update_data.get("requires_deposit")
        if requires_dep_in_update is False:
            update_data["deposit_amount"] = 0.0
        else:
            update_data["deposit_amount"] = max(update_data["deposit_amount"], MIN_DEPOSIT_AMOUNT)

    # Enforce unified payout cadence whenever requires_deposit is being turned ON
    if update_data.get("requires_deposit") is True:
        update_data["payout_schedule"] = "monthly_cutoff_20"
    elif update_data.get("requires_deposit") is False:
        update_data["payout_schedule"] = None

    await db.businesses.update_one({"id": user["business_id"]}, {"$set": update_data})
    business = await db.businesses.find_one({"id": user["business_id"]}, {"_id": 0, "password_hash": 0})
    # Apply defaults for legacy documents that may be missing required fields
    business.setdefault("description", "")
    business.setdefault("category_id", "")
    business.setdefault("address", "")
    business.setdefault("city", "")
    business.setdefault("state", "")
    business.setdefault("country", "MX")
    business.setdefault("zip_code", "")
    business.setdefault("subscription_status", "none")
    return BusinessResponse(**business)



@router.get("/me/private-info")
async def get_my_private_info(token_data: TokenData = Depends(require_business)):
    """Get private/legal information only visible to the business owner."""
    if token_data.is_manager:
        raise HTTPException(status_code=403, detail="Solo el dueno puede ver esta informacion")
    
    user = await db.users.find_one({"id": token_data.user_id}, {"_id": 0, "business_id": 1})
    if not user or not user.get("business_id"):
        raise HTTPException(status_code=404, detail="Business not found")
    
    business = await db.businesses.find_one(
        {"id": user["business_id"]},
        {"_id": 0, "ine_url": 1, "rfc": 1, "legal_name": 1, "clabe": 1, 
         "proof_of_address_url": 1, "owner_birth_date": 1,
         "stripe_customer_id": 1, "subscription_status": 1,
         "stripe_subscription_id": 1, "subscription_started_at": 1,
         "name": 1, "email": 1, "phone": 1, "description": 1,
         "notify_email": 1, "notify_sms": 1, "public_code": 1,
         "tax_regime": 1, "tax_regime_certificate_url": 1,
         "commission_terms_accepted_at": 1, "commission_terms_version": 1,
         "commission_terms_hash": 1, "commission_terms_snapshot": 1,
         "requires_deposit": 1, "deposit_amount": 1,
         "cancellation_days": 1, "payout_schedule": 1}
    )
    if not business:
        raise HTTPException(status_code=404, detail="Business not found")
    
    # Get subscription info from Stripe if available
    subscription_info = None
    if business.get("stripe_subscription_id"):
        try:
            sub = stripe_lib.Subscription.retrieve(business["stripe_subscription_id"])
            subscription_info = {
                "status": sub.status,
                "current_period_end": datetime.fromtimestamp(sub.current_period_end, tz=timezone.utc).isoformat() if sub.current_period_end else None,
                "cancel_at_period_end": sub.cancel_at_period_end,
                "trial_end": datetime.fromtimestamp(sub.trial_end, tz=timezone.utc).isoformat() if sub.trial_end else None,
            }
            # Get payment method info (last 4 digits)
            if sub.default_payment_method:
                pm = stripe_lib.PaymentMethod.retrieve(sub.default_payment_method)
                subscription_info["card_brand"] = pm.card.brand if pm.card else None
                subscription_info["card_last4"] = pm.card.last4 if pm.card else None
            elif business.get("stripe_customer_id"):
                pms = stripe_lib.PaymentMethod.list(customer=business["stripe_customer_id"], type="card", limit=1)
                if pms.data:
                    subscription_info["card_brand"] = pms.data[0].card.brand
                    subscription_info["card_last4"] = pms.data[0].card.last4
        except Exception as e:
            logger.warning(f"Failed to fetch subscription info: {e}")
    
    # Also include the bank_proof fields so it loads with the rest of the profile
    business_full = await db.businesses.find_one(
        {"id": user["business_id"]},
        {"_id": 0, "bank_proof_url": 1, "documents_verified": 1,
         "documents_verified_at": 1, "documents_rejection_reason": 1,
         "clabe_changed_at": 1}
    ) or {}

    return {
        "name": business.get("name", ""),
        "email": business.get("email", ""),
        "phone": business.get("phone", ""),
        "description": business.get("description", ""),
        "rfc": business.get("rfc", ""),
        "legal_name": business.get("legal_name", ""),
        "clabe": business.get("clabe", ""),
        "ine_url": business.get("ine_url", ""),
        "proof_of_address_url": business.get("proof_of_address_url", ""),
        "bank_proof_url": business_full.get("bank_proof_url", ""),
        "documents_verified": bool(business_full.get("documents_verified", False)),
        "documents_verified_at": business_full.get("documents_verified_at"),
        "documents_rejection_reason": business_full.get("documents_rejection_reason"),
        "clabe_changed_at": business_full.get("clabe_changed_at"),
        "owner_birth_date": business.get("owner_birth_date", ""),
        "subscription_status": business.get("subscription_status", "none"),
        "subscription_started_at": business.get("subscription_started_at"),
        "subscription_info": subscription_info,
        "notify_email": business.get("notify_email", True),
        "notify_sms": business.get("notify_sms", True),
        "public_code": business.get("public_code"),
        "tax_regime": business.get("tax_regime"),
        "tax_regime_certificate_url": business.get("tax_regime_certificate_url"),
        "commission_terms": {
            "accepted_at": business.get("commission_terms_accepted_at"),
            "version": business.get("commission_terms_version"),
            "hash": business.get("commission_terms_hash"),
            "snapshot": business.get("commission_terms_snapshot"),
        } if business.get("commission_terms_accepted_at") else None,
        "requires_deposit": business.get("requires_deposit", False),
        "deposit_amount": business.get("deposit_amount", 0.0),
        "cancellation_days": business.get("cancellation_days", 1),
        "payout_schedule": business.get("payout_schedule"),
    }


@router.post("/me/commission-terms/accept")
async def accept_commission_terms(
    payload: Dict[str, Any],
    request: Request,
    token_data: TokenData = Depends(require_business),
):
    """Business owner accepts (or re-accepts) the commission terms.
    Persists version + hash + snapshot + acceptance timestamp + IP.

    Body: {version: str, hash: str, snapshot: dict}
    """
    if token_data.is_manager:
        raise HTTPException(status_code=403, detail="Solo el dueno puede aceptar términos de comisiones")

    if not isinstance(payload, dict) or "snapshot" not in payload:
        raise HTTPException(status_code=400, detail="snapshot es obligatorio")
    version = (payload.get("version") or "").strip()
    hash_str = (payload.get("hash") or "").strip().lower()
    snapshot = payload.get("snapshot")
    if not version or not hash_str or not isinstance(snapshot, dict):
        raise HTTPException(status_code=400, detail="version, hash y snapshot son obligatorios")
    if len(hash_str) != 64 or not all(c in "0123456789abcdef" for c in hash_str):
        raise HTTPException(status_code=400, detail="hash debe ser SHA-256 hex (64 chars)")

    user = await db.users.find_one({"id": token_data.user_id}, {"_id": 0, "business_id": 1})
    if not user or not user.get("business_id"):
        raise HTTPException(status_code=404, detail="Business not found")

    now_iso = datetime.now(timezone.utc).isoformat()
    ip = request.headers.get("x-forwarded-for", "").split(",")[0].strip() or (request.client.host if request.client else "")
    ua = request.headers.get("user-agent", "")[:500]

    history_entry = {
        "version": version, "hash": hash_str, "snapshot": snapshot,
        "accepted_at": now_iso, "ip": ip, "user_agent": ua,
    }

    await db.businesses.update_one(
        {"id": user["business_id"]},
        {
            "$set": {
                "commission_terms_version": version,
                "commission_terms_hash": hash_str,
                "commission_terms_snapshot": snapshot,
                "commission_terms_accepted_at": now_iso,
            },
            "$push": {"commission_terms_history": history_entry},
        },
    )

    await create_audit_log(
        admin_id=token_data.user_id, admin_email=token_data.email,
        action="commission_terms_accept", target_type="business",
        target_id=user["business_id"],
        details={"version": version, "hash": hash_str},
        request=request,
    )
    return {"ok": True, "version": version, "hash": hash_str, "accepted_at": now_iso}


class TaxRegimeUpdate(BaseModel):
    tax_regime: str
    tax_regime_certificate_url: Optional[str] = None


@router.put("/me/tax-regime")
async def update_my_tax_regime(
    payload: TaxRegimeUpdate,
    token_data: TokenData = Depends(require_business),
):
    """Business owner updates its tax regime + (optional) certificate URL.
    Used for future Fintech withholding calculations (LISR 113-A)."""
    if token_data.is_manager:
        raise HTTPException(status_code=403, detail="Solo el dueno puede actualizar régimen fiscal")

    valid = {"PF_RESICO", "PF_ACT_EMPRESARIAL", "PF_HONORARIOS",
             "PF_PLATAFORMAS", "PM_GENERAL", "PM_NO_LUCRATIVA",
             "RIF", "OTRO"}
    if payload.tax_regime not in valid:
        raise HTTPException(status_code=400, detail="Régimen fiscal no válido")

    user = await db.users.find_one({"id": token_data.user_id}, {"_id": 0, "business_id": 1})
    if not user or not user.get("business_id"):
        raise HTTPException(status_code=404, detail="Business not found")

    await db.businesses.update_one(
        {"id": user["business_id"]},
        {"$set": {
            "tax_regime": payload.tax_regime,
            "tax_regime_certificate_url": payload.tax_regime_certificate_url,
            "tax_regime_updated_at": datetime.now(timezone.utc).isoformat(),
        }},
    )
    return {"ok": True, "tax_regime": payload.tax_regime}


@router.get("/me/legal-file.pdf")
async def download_my_legal_file(
    request: Request,
    token_data: TokenData = Depends(require_business),
):
    """Download the business owner's legal file (expediente) as a PDF.
    Only the owner (not managers) can download. Audit-logged."""
    if token_data.is_manager:
        raise HTTPException(status_code=403, detail="Solo el dueno puede descargar el expediente")

    user = await db.users.find_one({"id": token_data.user_id}, {"_id": 0, "business_id": 1})
    if not user or not user.get("business_id"):
        raise HTTPException(status_code=404, detail="Business not found")

    from fastapi.responses import Response as _Response
    from services.legal_file_service import generate_business_legal_file

    origin = request.headers.get("origin") or str(request.base_url).rstrip("/")
    result = await generate_business_legal_file(user["business_id"], origin)
    if not result:
        raise HTTPException(status_code=404, detail="Business not found")

    await create_audit_log(
        admin_id=token_data.user_id, admin_email=token_data.email,
        action="legal_file_download", target_type="business",
        target_id=user["business_id"],
        details={"file_id": result["file_id"], "by": "owner"},
        request=request,
    )

    safe_rfc = (result.get("rfc") or "sin_rfc").replace("/", "_")[:20]
    filename = f"expediente_bookvia_{safe_rfc}_{datetime.now(timezone.utc).strftime('%Y%m%d')}.pdf"
    return _Response(
        content=result["pdf_bytes"],
        media_type="application/pdf",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "X-Legal-File-Id": result["file_id"],
            "X-Legal-File-Hash": result["content_hash"],
        },
    )


@router.get("/verificar-expediente/{file_id}")
async def verify_legal_file(file_id: str):
    """Public endpoint — anyone can verify that Bookvia issued a given
    expediente. Returns minimal, non-sensitive data for transparency."""
    record = await db.business_legal_files.find_one(
        {"id": file_id.strip().upper()}, {"_id": 0}
    )
    if not record:
        return {"ok": False, "error": "Expediente no encontrado. Verifica el folio."}
    return {
        "ok": True,
        "file_id": record["id"],
        "issued_at": record.get("issued_at"),
        "legal_name": record.get("legal_name"),
        "rfc_masked": (record.get("rfc") or "")[:4] + "••••" + (record.get("rfc") or "")[-3:],
        "public_code": record.get("public_code"),
        "content_hash": record.get("content_hash"),
        "file_version": record.get("file_version"),
    }


@router.get("/me/settlements")
async def list_my_settlements(
    limit: int = 24,
    token_data: TokenData = Depends(require_business),
):
    """Return the business's past day-20 settlements (most recent first)."""
    user = await db.users.find_one({"id": token_data.user_id}, {"_id": 0, "business_id": 1})
    if not user or not user.get("business_id"):
        raise HTTPException(status_code=404, detail="Business not found")

    rows = await db.settlements.find(
        {"business_id": user["business_id"]},
        {"_id": 0, "id": 1, "period_key": 1, "net_payout": 1, "payout_amount": 1,
         "booking_count": 1, "status": 1, "created_at": 1, "paid_at": 1,
         "transaction_ids": 1},
    ).sort("created_at", -1).limit(max(1, min(limit, 100))).to_list(100)

    # Slim the payload
    items = []
    for r in rows:
        items.append({
            "id": r.get("id"),
            "period_key": r.get("period_key"),
            "net_amount": float(r.get("net_payout") or r.get("payout_amount") or 0),
            "booking_count": r.get("booking_count", 0),
            "transaction_count": len(r.get("transaction_ids") or []),
            "status": r.get("status") or "pending",
            "created_at": r.get("created_at"),
            "paid_at": r.get("paid_at"),
        })
    return {"items": items, "count": len(items)}


@router.get("/me/settlements/{settlement_id}/statement.pdf")
async def download_my_settlement_statement(
    settlement_id: str,
    request: Request,
    token_data: TokenData = Depends(require_business),
):
    """Owner downloads the PDF statement for one of their own settlements."""
    user = await db.users.find_one({"id": token_data.user_id}, {"_id": 0, "business_id": 1})
    if not user or not user.get("business_id"):
        raise HTTPException(status_code=404, detail="Business not found")

    # Ownership check: settlement must belong to this business
    settlement = await db.settlements.find_one(
        {"id": settlement_id, "business_id": user["business_id"]},
        {"_id": 0, "id": 1},
    )
    if not settlement:
        raise HTTPException(status_code=404, detail="Estado de cuenta no encontrado")

    from fastapi.responses import Response as _Response
    from services.payout_statement import generate_payout_statement_pdf

    result = await generate_payout_statement_pdf(settlement_id)
    if not result:
        raise HTTPException(status_code=404, detail="Estado de cuenta no encontrado")

    await create_audit_log(
        admin_id=token_data.user_id, admin_email=token_data.email,
        action="payout_statement_download", target_type="settlement",
        target_id=settlement_id,
        details={"by": "owner", "period_key": result.get("period_key")},
        request=request,
    )

    safe_rfc = (result.get("rfc") or "sin_rfc").replace("/", "_")[:20]
    period = result.get("period_key") or "sinperiodo"
    filename = f"estado_de_cuenta_bookvia_{safe_rfc}_{period}.pdf"
    return _Response(
        content=result["pdf_bytes"], media_type="application/pdf",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "X-Statement-Hash": result["content_hash"],
        },
    )


@router.put("/me/legal-docs")
async def update_my_legal_docs(
    payload: BusinessLegalDocsUpdate,
    token_data: TokenData = Depends(require_business),
):
    """Business owner updates its legal/banking documents.

    Any change to these fields ALWAYS sets `documents_verified=false` and
    notifies admins so the new documentation can be re-reviewed. A CLABE
    change is recorded with `clabe_changed_at`.
    """
    if token_data.is_manager:
        raise HTTPException(status_code=403, detail="Solo el dueno puede actualizar documentos")

    # Fase 10: hard-gate this owner action if T&C are outdated and the grace period passed.
    from routers.terms import require_terms_up_to_date
    await require_terms_up_to_date(token_data.user_id)

    user = await db.users.find_one({"id": token_data.user_id}, {"_id": 0, "business_id": 1})
    if not user or not user.get("business_id"):
        raise HTTPException(status_code=404, detail="Business not found")

    business = await db.businesses.find_one({"id": user["business_id"]}, {"_id": 0})
    if not business:
        raise HTTPException(status_code=404, detail="Business not found")

    data = {k: v for k, v in payload.model_dump().items() if v is not None}
    if not data:
        raise HTTPException(status_code=400, detail="No data to update")

    sensitive_fields = {"rfc", "clabe", "legal_name", "ine_url", "proof_of_address_url", "bank_proof_url"}
    changed_sensitive = any(
        f in data and str(data.get(f, "")).strip() != str(business.get(f, "")).strip()
        for f in sensitive_fields
    )
    now_iso = datetime.now(timezone.utc).isoformat()

    update_set = dict(data)
    if changed_sensitive:
        update_set.update({
            "documents_verified": False,
            "documents_verified_at": None,
            "documents_rejection_reason": None,
            "documents_submitted_at": now_iso,
        })
    if "clabe" in data and str(data["clabe"]).strip() != str(business.get("clabe", "")).strip():
        update_set["clabe_changed_at"] = now_iso

    await db.businesses.update_one({"id": business["id"]}, {"$set": update_set})

    # Notify every admin for review (best-effort)
    if changed_sensitive:
        try:
            admin_users = await db.users.find(
                {"role": UserRole.ADMIN},
                {"_id": 0, "id": 1},
            ).to_list(50)
            for admin in admin_users:
                await create_notification(
                    admin["id"],
                    "Documentos pendientes de revision",
                    f"El negocio {business.get('name','(sin nombre)')} actualizo sus documentos legales o bancarios.",
                    "docs_review",
                    {"business_id": business["id"], "changed_clabe": "clabe" in data},
                )
        except Exception as e:
            logger.warning(f"Failed to notify admins about docs change: {e}")

    return {
        "message": "Documentos actualizados",
        "documents_verified": False if changed_sensitive else bool(business.get("documents_verified", False)),
        "requires_review": changed_sensitive,
    }



@router.get("/me/dashboard")
async def get_business_dashboard(token_data: TokenData = Depends(require_business)):
    user = await db.users.find_one({"id": token_data.user_id})
    if not user or not user.get("business_id"):
        raise HTTPException(status_code=404, detail="Business not found")
    
    business = await db.businesses.find_one({"id": user["business_id"]}, {"_id": 0, "password_hash": 0})
    if not business:
        raise HTTPException(status_code=404, detail="Business document not found")
    
    # Ensure slug exists
    if not business.get("slug"):
        slug = f"{business['name'].lower().replace(' ', '-')}-{business['id'][:8]}"
        business["slug"] = slug
        await db.businesses.update_one({"id": business["id"]}, {"$set": {"slug": slug}})
    
    # Ensure required fields have defaults to prevent Pydantic validation errors
    business.setdefault("description", "")
    business.setdefault("category_id", "")
    business.setdefault("address", "")
    business.setdefault("city", "")
    business.setdefault("state", "")
    business.setdefault("country", "MX")
    business.setdefault("zip_code", "")
    business.setdefault("subscription_status", "none")
    
    # Get stats
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    bid = business["id"]
    
    today_appointments = await db.bookings.count_documents({
        "business_id": bid,
        "date": today,
        "status": AppointmentStatus.CONFIRMED
    })
    
    pending_appointments = await db.bookings.count_documents({
        "business_id": bid,
        "status": AppointmentStatus.CONFIRMED
    })
    
    # This month revenue
    first_of_month = datetime.now(timezone.utc).replace(day=1).strftime("%Y-%m-%d")
    month_bookings = await db.bookings.find({
        "business_id": bid,
        "date": {"$gte": first_of_month},
        "status": AppointmentStatus.COMPLETED
    }).to_list(1000)
    
    month_revenue = sum(b.get("total_amount", 0) for b in month_bookings)
    
    total_appointments = await db.bookings.count_documents({
        "business_id": bid,
        "status": {"$in": [AppointmentStatus.CONFIRMED, AppointmentStatus.COMPLETED, AppointmentStatus.NO_SHOW]}
    })
    
    try:
        biz_response = BusinessResponse(**business).model_dump()
    except Exception as e:
        logger.error(f"BusinessResponse validation error for {bid}: {e}")
        logger.error(f"Business fields: {list(business.keys())}")
        raise HTTPException(status_code=500, detail=f"Error loading business data: {str(e)}")
    
    return {
        "business": biz_response,
        "stats": {
            "today_appointments": today_appointments,
            "pending_appointments": pending_appointments,
            "month_revenue": month_revenue,
            "total_appointments": total_appointments,
            "total_reviews": business.get("review_count", 0),
            "rating": business.get("rating", 0)
        },
        "subscription": {
            "status": business.get("subscription_status", "none"),
            "subscription_id": business.get("stripe_subscription_id"),
            "customer_id": business.get("stripe_customer_id"),
            "started_at": business.get("subscription_started_at"),
            "cancel_requested": business.get("subscription_cancel_requested", False)
        }
    }



@router.get("/me/blacklist", response_model=List[BlacklistResponse])
async def get_blacklist(token_data: TokenData = Depends(require_business)):
    user = await db.users.find_one({"id": token_data.user_id})
    if not user or not user.get("business_id"):
        raise HTTPException(status_code=404, detail="Business not found")
    entries = await db.blacklist.find(
        {"business_id": user["business_id"]}, {"_id": 0}
    ).sort("created_at", -1).to_list(500)
    return [BlacklistResponse(**e) for e in entries]



@router.post("/me/blacklist", response_model=BlacklistResponse)
async def add_to_blacklist(entry: BlacklistEntry, token_data: TokenData = Depends(require_business)):
    user = await db.users.find_one({"id": token_data.user_id})
    if not user or not user.get("business_id"):
        raise HTTPException(status_code=404, detail="Business not found")
    if not entry.email and not entry.phone and not entry.user_id:
        raise HTTPException(status_code=400, detail="At least one identifier (email, phone, or user_id) is required")
    
    # Normalize email
    email_normalized = entry.email.strip().lower() if entry.email else None
    
    # Check if already blacklisted
    or_conditions = []
    if email_normalized:
        or_conditions.append({"email": email_normalized})
    if entry.phone:
        or_conditions.append({"phone": entry.phone})
    if entry.user_id:
        or_conditions.append({"user_id": entry.user_id})
    
    existing = await db.blacklist.find_one({
        "business_id": user["business_id"],
        "$or": or_conditions
    })
    if existing:
        raise HTTPException(status_code=409, detail="This user is already blacklisted")
    
    doc = {
        "id": generate_id(),
        "business_id": user["business_id"],
        "email": email_normalized,
        "phone": entry.phone,
        "user_id": entry.user_id,
        "reason": entry.reason,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.blacklist.insert_one(doc)
    return BlacklistResponse(**{k: v for k, v in doc.items() if k != "_id"})



@router.delete("/me/blacklist/{entry_id}")
async def remove_from_blacklist(entry_id: str, token_data: TokenData = Depends(require_business)):
    user = await db.users.find_one({"id": token_data.user_id})
    if not user or not user.get("business_id"):
        raise HTTPException(status_code=404, detail="Business not found")
    result = await db.blacklist.delete_one({"id": entry_id, "business_id": user["business_id"]})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Entry not found")
    return {"message": "Removed from blacklist"}



@router.post("/my/workers", response_model=WorkerResponse)
async def create_worker(worker: WorkerCreate, token_data: TokenData = Depends(require_business)):
    """Create a new worker for the business"""
    user = await db.users.find_one({"id": token_data.user_id})
    if not user or not user.get("business_id"):
        raise HTTPException(status_code=404, detail="Business not found")
    
    # Default schedule: Mon-Fri 9:00-18:00
    default_schedule = {}
    for day in range(5):  # Mon-Fri
        default_schedule[str(day)] = {
            "is_available": True,
            "blocks": [{"start_time": "09:00", "end_time": "18:00"}]
        }
    for day in range(5, 7):  # Sat-Sun
        default_schedule[str(day)] = {"is_available": False, "blocks": []}
    
    worker_doc = {
        "id": generate_id(),
        "business_id": user["business_id"],
        "name": worker.name,
        "email": worker.email,
        "phone": worker.phone,
        "photo_url": worker.photo_url,
        "bio": worker.bio,
        "service_ids": worker.service_ids,
        "schedule": default_schedule,
        "exceptions": [],  # New: list of WorkerException
        "active": True,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "deactivated_at": None
    }
    
    await db.workers.insert_one(worker_doc)
    return WorkerResponse(**worker_doc)



@router.get("/my/workers", response_model=List[WorkerResponse])
async def get_business_workers(
    include_inactive: bool = False,
    token_data: TokenData = Depends(require_auth)
):
    """Get all workers for the authenticated business"""
    user = await db.users.find_one({"id": token_data.user_id})
    if not user or not user.get("business_id"):
        raise HTTPException(status_code=400, detail="Business ID required")
    business_id = user["business_id"]
    
    filters = {"business_id": business_id}
    if not include_inactive:
        filters["active"] = True
    
    workers = await db.workers.find(filters, {"_id": 0}).to_list(100)
    for w in workers:
        w["has_manager_pin"] = bool(w.get("manager_pin_hash"))
    return [WorkerResponse(**w) for w in workers]



@router.get("/{business_id}/workers", response_model=List[WorkerResponse])
async def get_workers_by_business(business_id: str, service_id: Optional[str] = None, include_inactive: bool = False):
    """Get workers for a specific business (public endpoint), optionally filtered by service."""
    filters = {"business_id": business_id}
    if not include_inactive:
        filters["active"] = True
    workers = await db.workers.find(filters, {"_id": 0}).to_list(100)
    # Filter by service_id: include workers who have the service in their service_ids or have no service_ids set (legacy)
    if service_id:
        workers = [w for w in workers if not w.get("service_ids") or service_id in w.get("service_ids", [])]
    for w in workers:
        w["has_manager_pin"] = bool(w.get("manager_pin_hash"))
    return [WorkerResponse(**w) for w in workers]



@router.get("/my/workers/{worker_id}", response_model=WorkerResponse)
async def get_worker(worker_id: str, token_data: TokenData = Depends(require_business)):
    """Get a specific worker"""
    user = await db.users.find_one({"id": token_data.user_id})
    worker = await db.workers.find_one({"id": worker_id, "business_id": user.get("business_id")}, {"_id": 0})
    if not worker:
        raise HTTPException(status_code=404, detail="Worker not found")
    worker["has_manager_pin"] = bool(worker.get("manager_pin_hash"))
    return WorkerResponse(**worker)



@router.put("/my/workers/{worker_id}/services")
async def update_worker_services(worker_id: str, request: Request, token_data: TokenData = Depends(require_business)):
    """Update which services a worker can perform."""
    user = await db.users.find_one({"id": token_data.user_id})
    if not user or not user.get("business_id"):
        raise HTTPException(status_code=404, detail="Business not found")
    
    worker = await db.workers.find_one({"id": worker_id, "business_id": user["business_id"]})
    if not worker:
        raise HTTPException(status_code=404, detail="Worker not found")
    
    body = await request.json()
    service_ids = body.get("service_ids", [])
    
    await db.workers.update_one(
        {"id": worker_id},
        {"$set": {"service_ids": service_ids}}
    )
    
    return {"message": "Worker services updated", "service_ids": service_ids}




@router.put("/my/workers/{worker_id}", response_model=WorkerResponse)
async def update_worker(
    worker_id: str,
    update: WorkerUpdate,
    token_data: TokenData = Depends(require_business)
):
    """Update worker basic info"""
    user = await db.users.find_one({"id": token_data.user_id})
    worker = await db.workers.find_one({"id": worker_id, "business_id": user.get("business_id")})
    if not worker:
        raise HTTPException(status_code=404, detail="Worker not found")
    
    update_data = {k: v for k, v in update.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="No data to update")
    
    await db.workers.update_one({"id": worker_id}, {"$set": update_data})
    updated = await db.workers.find_one({"id": worker_id}, {"_id": 0})
    return WorkerResponse(**updated)



@router.delete("/my/workers/{worker_id}")
async def delete_worker(worker_id: str, token_data: TokenData = Depends(require_business)):
    """Soft delete a worker (marks as inactive, preserves history)"""
    user = await db.users.find_one({"id": token_data.user_id})
    worker = await db.workers.find_one({"id": worker_id, "business_id": user.get("business_id")})
    if not worker:
        raise HTTPException(status_code=404, detail="Worker not found")
    
    # Soft delete: mark as inactive but preserve all data
    await db.workers.update_one(
        {"id": worker_id},
        {"$set": {
            "active": False,
            "deactivated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    return {"message": "Worker deactivated", "worker_id": worker_id}




@router.post("/my/workers/{worker_id}/photo")
async def upload_worker_photo(worker_id: str, file: UploadFile = File(...), token_data: TokenData = Depends(require_business)):
    """Upload or replace a worker's profile photo."""
    user = await db.users.find_one({"id": token_data.user_id})
    if not user or not user.get("business_id"):
        raise HTTPException(status_code=404, detail="Business not found")

    worker = await db.workers.find_one({"id": worker_id, "business_id": user["business_id"]})
    if not worker:
        raise HTTPException(status_code=404, detail="Worker not found")

    data = await file.read()
    ok, err = validate_image(file.filename, file.content_type, len(data))
    if not ok:
        raise HTTPException(status_code=400, detail=err)

    if cloudinary_configured():
        old_pid = worker.get("photo_public_id")
        if old_pid:
            cloudinary_delete(old_pid)
        try:
            result = upload_image(data, "business_gallery", f"workers/{worker_id}")
            photo_url = result["secure_url"]
            public_id = result["public_id"]
        except Exception as e:
            logger.error(f"Worker photo upload failed: {e}")
            raise HTTPException(status_code=500, detail="Failed to upload photo")
    else:
        path = generate_upload_path(user["business_id"], f"worker_{worker_id}_{file.filename}")
        content_type = "image/jpeg" if file.filename.lower().endswith(("jfif", "jpg", "jpeg")) else file.content_type
        try:
            result = put_object(path, data, content_type)
            base_url = os.environ.get("BASE_URL", "")
            photo_url = f"{base_url}/api/files/{result['path']}"
            public_id = path
        except Exception as e:
            logger.error(f"Worker photo upload failed: {e}")
            raise HTTPException(status_code=500, detail="Failed to upload photo")

    await db.workers.update_one(
        {"id": worker_id},
        {"$set": {"photo_url": photo_url, "photo_public_id": public_id}}
    )

    return {"secure_url": photo_url, "public_id": public_id}





@router.put("/my/workers/{worker_id}/schedule")
async def update_worker_schedule(
    worker_id: str,
    schedule_update: WorkerScheduleUpdate,
    token_data: TokenData = Depends(require_business)
):
    """Update worker schedule with validation for overlapping blocks"""
    user = await db.users.find_one({"id": token_data.user_id})
    worker = await db.workers.find_one({"id": worker_id, "business_id": user.get("business_id")})
    if not worker:
        raise HTTPException(status_code=404, detail="Worker not found")
    
    # Validate no overlapping blocks
    validate_schedule_blocks(schedule_update.schedule)
    
    # Convert to dict for storage
    schedule_dict = {}
    for day, day_schedule in schedule_update.schedule.items():
        schedule_dict[day] = {
            "is_available": day_schedule.is_available,
            "blocks": [{"start_time": b.start_time, "end_time": b.end_time} for b in day_schedule.blocks]
        }
    
    await db.workers.update_one({"id": worker_id}, {"$set": {"schedule": schedule_dict}})
    updated = await db.workers.find_one({"id": worker_id}, {"_id": 0})
    return WorkerResponse(**updated)



@router.post("/my/workers/{worker_id}/exceptions")
async def add_worker_exception(
    worker_id: str,
    exception_data: WorkerExceptionAdd,
    token_data: TokenData = Depends(require_business)
):
    """Add an exception (vacation/block) to a worker's schedule"""
    user = await db.users.find_one({"id": token_data.user_id})
    worker = await db.workers.find_one({"id": worker_id, "business_id": user.get("business_id")})
    if not worker:
        raise HTTPException(status_code=404, detail="Worker not found")
    
    exception = exception_data.exception
    
    # Validate dates
    try:
        start = datetime.strptime(exception.start_date, "%Y-%m-%d")
        end = datetime.strptime(exception.end_date, "%Y-%m-%d")
        if end < start:
            raise HTTPException(status_code=400, detail="end_date cannot be before start_date")
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")
    
    # Validate times if provided
    if exception.start_time and exception.end_time:
        try:
            st = datetime.strptime(exception.start_time, "%H:%M")
            et = datetime.strptime(exception.end_time, "%H:%M")
            if et <= st:
                raise HTTPException(status_code=400, detail="end_time must be after start_time")
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid time format. Use HH:MM")
    
    exception_dict = {
        "id": generate_id(),
        "start_date": exception.start_date,
        "end_date": exception.end_date,
        "start_time": exception.start_time,
        "end_time": exception.end_time,
        "reason": exception.reason,
        "exception_type": exception.exception_type,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.workers.update_one(
        {"id": worker_id},
        {"$push": {"exceptions": exception_dict}}
    )
    
    return {"message": "Exception added", "exception_id": exception_dict["id"]}



@router.delete("/my/workers/{worker_id}/exceptions/{exception_id}")
async def remove_worker_exception(
    worker_id: str,
    exception_id: str,
    token_data: TokenData = Depends(require_business)
):
    """Remove an exception from a worker's schedule"""
    user = await db.users.find_one({"id": token_data.user_id})
    worker = await db.workers.find_one({"id": worker_id, "business_id": user.get("business_id")})
    if not worker:
        raise HTTPException(status_code=404, detail="Worker not found")
    
    await db.workers.update_one(
        {"id": worker_id},
        {"$pull": {"exceptions": {"id": exception_id}}}
    )
    
    return {"message": "Exception removed"}




@router.post("/me/pin")
async def set_owner_pin(data: PinCreate, token_data: TokenData = Depends(require_business)):
    """Set or update the owner's security PIN (4-6 digits)"""
    if not data.pin.isdigit() or not (4 <= len(data.pin) <= 6):
        raise HTTPException(status_code=400, detail="PIN must be 4-6 digits")
    
    user = await db.users.find_one({"id": token_data.user_id})
    pin_hash = hash_password(data.pin)
    await db.businesses.update_one(
        {"id": user["business_id"]},
        {"$set": {"owner_pin_hash": pin_hash}}
    )
    return {"message": "PIN created successfully"}



@router.post("/me/pin/verify")
async def verify_owner_pin(data: PinVerify, token_data: TokenData = Depends(require_business)):
    """Verify the owner's PIN"""
    user = await db.users.find_one({"id": token_data.user_id})
    business = await db.businesses.find_one({"id": user["business_id"]})
    pin_hash = business.get("owner_pin_hash")
    if not pin_hash:
        raise HTTPException(status_code=400, detail="No PIN configured")
    if not verify_password(data.pin, pin_hash):
        raise HTTPException(status_code=401, detail="Incorrect PIN")
    return {"verified": True}



@router.get("/me/pin/status")
async def get_pin_status(token_data: TokenData = Depends(require_business)):
    """Check if owner has a PIN configured"""
    user = await db.users.find_one({"id": token_data.user_id})
    business = await db.businesses.find_one({"id": user["business_id"]})
    return {"has_pin": bool(business.get("owner_pin_hash"))}



@router.put("/my/workers/{worker_id}/manager")
async def designate_manager(worker_id: str, data: ManagerDesignate, token_data: TokenData = Depends(require_business)):
    """Designate a worker as manager with custom permissions"""
    user = await db.users.find_one({"id": token_data.user_id})
    business_id = user.get("business_id")
    
    worker = await db.workers.find_one({"id": worker_id, "business_id": business_id})
    if not worker:
        raise HTTPException(status_code=404, detail="Worker not found")
    
    # Check max managers (limit 2)
    current_managers = await db.workers.count_documents({"business_id": business_id, "is_manager": True, "id": {"$ne": worker_id}})
    if current_managers >= 2:
        raise HTTPException(status_code=400, detail="Maximum 2 managers allowed")
    
    # Merge provided permissions with defaults
    permissions = {**DEFAULT_MANAGER_PERMISSIONS, **data.permissions}
    
    await db.workers.update_one(
        {"id": worker_id},
        {"$set": {
            "is_manager": True,
            "manager_permissions": permissions,
            "manager_designated_at": datetime.now(timezone.utc).isoformat(),
            "manager_designated_by": token_data.user_id,
        }}
    )
    
    updated = await db.workers.find_one({"id": worker_id}, {"_id": 0})
    updated["has_manager_pin"] = bool(updated.get("manager_pin_hash"))
    
    # Log activity
    await create_business_activity(
        business_id, token_data, "designate_admin", "worker", worker_id,
        {"worker_name": worker.get("name", ""), "permissions": permissions}
    )
    
    return WorkerResponse(**updated)



@router.put("/my/workers/{worker_id}/manager/permissions")
async def update_manager_permissions(worker_id: str, data: ManagerPermissionsUpdate, token_data: TokenData = Depends(require_business)):
    """Update a manager's permissions"""
    user = await db.users.find_one({"id": token_data.user_id})
    worker = await db.workers.find_one({"id": worker_id, "business_id": user.get("business_id"), "is_manager": True})
    if not worker:
        raise HTTPException(status_code=404, detail="Manager not found")
    
    permissions = {**worker.get("manager_permissions", DEFAULT_MANAGER_PERMISSIONS), **data.permissions}
    await db.workers.update_one({"id": worker_id}, {"$set": {"manager_permissions": permissions}})
    
    # Log activity
    await create_business_activity(
        user.get("business_id"), token_data, "update_permissions", "worker", worker_id,
        {"worker_name": worker.get("name", ""), "permissions": permissions}
    )
    
    return {"message": "Permissions updated"}



@router.delete("/my/workers/{worker_id}/manager")
async def remove_manager(worker_id: str, token_data: TokenData = Depends(require_business)):
    """Remove manager role from a worker"""
    user = await db.users.find_one({"id": token_data.user_id})
    worker = await db.workers.find_one({"id": worker_id, "business_id": user.get("business_id")})
    if not worker:
        raise HTTPException(status_code=404, detail="Worker not found")
    
    await db.workers.update_one(
        {"id": worker_id},
        {"$set": {"is_manager": False}, "$unset": {"manager_pin_hash": "", "manager_permissions": "", "manager_designated_at": "", "manager_designated_by": ""}}
    )
    
    # Log activity
    await create_business_activity(
        user.get("business_id"), token_data, "remove_admin", "worker", worker_id,
        {"worker_name": worker.get("name", "")}
    )
    
    return {"message": "Manager role removed"}



@router.post("/my/workers/{worker_id}/manager/pin")
async def set_manager_pin(worker_id: str, data: PinCreate, token_data: TokenData = Depends(require_business)):
    """Set or reset a manager's PIN (can be called by owner or manager themselves)"""
    if not data.pin.isdigit() or not (4 <= len(data.pin) <= 6):
        raise HTTPException(status_code=400, detail="PIN must be 4-6 digits")
    
    user = await db.users.find_one({"id": token_data.user_id})
    worker = await db.workers.find_one({"id": worker_id, "business_id": user.get("business_id"), "is_manager": True})
    if not worker:
        raise HTTPException(status_code=404, detail="Manager not found")
    
    pin_hash = hash_password(data.pin)
    await db.workers.update_one({"id": worker_id}, {"$set": {"manager_pin_hash": pin_hash}})
    return {"message": "Manager PIN set successfully"}



@router.get("/my/activity-log")
async def get_business_activity_log(
    page: int = 1,
    limit: int = 30,
    actor_type: Optional[str] = None,
    action: Optional[str] = None,
    token_data: TokenData = Depends(require_business)
):
    """Get business activity log (owner only)"""
    if token_data.is_manager:
        raise HTTPException(status_code=403, detail="Solo el dueño puede ver el historial de actividad")
    
    user = await db.users.find_one({"id": token_data.user_id})
    business_id = user.get("business_id")
    
    query = {"business_id": business_id}
    if actor_type:
        query["actor_type"] = actor_type
    if action:
        query["action"] = action
    
    total = await db.business_activity_logs.count_documents(query)
    skip = (page - 1) * limit
    
    logs = await db.business_activity_logs.find(
        query, {"_id": 0}
    ).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    
    return {"logs": logs, "total": total, "page": page, "pages": (total + limit - 1) // limit if total > 0 else 1}




@router.post("/me/subscribe")
async def create_subscription_checkout(request: Request, token_data: TokenData = Depends(require_business)):
    """Create a Stripe Checkout Session for the monthly subscription."""
    user = await db.users.find_one({"id": token_data.user_id})
    if not user or not user.get("business_id"):
        raise HTTPException(status_code=404, detail="Business not found")
    
    business = await db.businesses.find_one({"id": user["business_id"]})
    if not business:
        raise HTTPException(status_code=404, detail="Business not found")
    
    # Check if already subscribed
    if business.get("stripe_subscription_id"):
        raise HTTPException(status_code=400, detail="Already subscribed")
    
    price_id = await get_or_create_stripe_price(country_code=business.get("country_code", "MX"))
    if not price_id:
        raise HTTPException(status_code=500, detail="Payment configuration error")
    
    body = await request.json()
    origin_url = body.get("origin_url", os.environ.get("BASE_URL", ""))
    success_url = f"{origin_url}/business/subscription/success?session_id={{CHECKOUT_SESSION_ID}}"
    cancel_url = f"{origin_url}/business/dashboard"
    
    try:
        # Create or get Stripe Customer
        customer_id = business.get("stripe_customer_id")
        if not customer_id:
            customer = stripe_lib.Customer.create(
                email=business.get("email"),
                name=business.get("name"),
                metadata={"business_id": business["id"]}
            )
            customer_id = customer.id
            await db.businesses.update_one(
                {"id": business["id"]},
                {"$set": {"stripe_customer_id": customer_id}}
            )
        
        # Create Checkout Session for subscription
        session = stripe_lib.checkout.Session.create(
            customer=customer_id,
            payment_method_types=["card"],
            mode="subscription",
            line_items=[{"price": price_id, "quantity": 1}],
            subscription_data={"trial_period_days": SUBSCRIPTION_TRIAL_DAYS},
            success_url=success_url,
            cancel_url=cancel_url,
            metadata={"business_id": business["id"]}
        )
        
        # Record the transaction (currency depends on country)
        is_usd = (business.get("country_code") or "MX").upper() != "MX"
        sub_amount = SUBSCRIPTION_PRICE_USD if is_usd else SUBSCRIPTION_PRICE_MXN
        sub_currency = "usd" if is_usd else "mxn"
        await db.payment_transactions.insert_one({
            "id": generate_id(),
            "business_id": business["id"],
            "session_id": session.id,
            "type": "subscription",
            "amount": sub_amount,
            "currency": sub_currency,
            "payment_status": "pending",
            "created_at": datetime.now(timezone.utc).isoformat()
        })
        
        return {"url": session.url, "session_id": session.id}
    except Exception as e:
        logger.error(f"Stripe subscription error: {e}")
        raise HTTPException(status_code=500, detail=str(e))




@router.get("/me/subscription/status")
async def get_subscription_status(session_id: str = None, token_data: TokenData = Depends(require_business)):
    """Check subscription status."""
    user = await db.users.find_one({"id": token_data.user_id})
    if not user or not user.get("business_id"):
        raise HTTPException(status_code=404, detail="Business not found")
    
    business = await db.businesses.find_one({"id": user["business_id"]})
    
    # If session_id provided, verify the checkout session
    if session_id:
        try:
            session = stripe_lib.checkout.Session.retrieve(session_id)
            if session.payment_status in ("paid", "no_payment_required") or session.status == "complete":
                subscription_id = session.subscription
                # Update business with subscription info
                await db.businesses.update_one(
                    {"id": user["business_id"]},
                    {"$set": {
                        "stripe_subscription_id": subscription_id,
                        "stripe_customer_id": session.customer,
                        "subscription_status": "trialing",
                        "subscription_started_at": datetime.now(timezone.utc).isoformat()
                    }}
                )
                # Update transaction
                await db.payment_transactions.update_one(
                    {"session_id": session_id},
                    {"$set": {"payment_status": "completed", "subscription_id": subscription_id}}
                )
                return {
                    "status": "active",
                    "subscription_status": "trialing",
                    "subscription_id": subscription_id,
                    "trial": True,
                    "message": "Subscription activated with 30-day free trial"
                }
        except Exception as e:
            logger.error(f"Stripe session check error: {e}")
    
    # Return current subscription status with details from Stripe
    sub_id = business.get("stripe_subscription_id") if business else None
    sub_status = business.get("subscription_status", "none") if business else "none"
    
    result = {
        "status": sub_status if sub_id else "none",
        "subscription_status": sub_status,
        "subscription_id": sub_id,
        "trial": sub_status == "trialing",
        "current_period_end": None,
        "cancel_at_period_end": False,
        "country_code": (business.get("country_code") if business else None) or "MX"
    }
    
    # Fetch live details from Stripe if subscription exists
    if sub_id:
        try:
            sub = stripe_lib.Subscription.retrieve(sub_id)
            result["current_period_end"] = datetime.fromtimestamp(sub.current_period_end, tz=timezone.utc).isoformat() if sub.current_period_end else None
            result["cancel_at_period_end"] = sub.cancel_at_period_end
            # Sync status from Stripe
            stripe_status = sub.status  # trialing, active, past_due, canceled, unpaid
            if stripe_status != sub_status:
                await db.businesses.update_one(
                    {"id": user["business_id"]},
                    {"$set": {"subscription_status": stripe_status}}
                )
                result["status"] = stripe_status
                result["subscription_status"] = stripe_status
                result["trial"] = stripe_status == "trialing"
        except Exception as e:
            logger.error(f"Stripe subscription fetch error: {e}")
    
    return result




@router.post("/me/subscription/cancel")
async def cancel_subscription(token_data: TokenData = Depends(require_business)):
    """Cancel the business subscription at the end of the current period."""
    user = await db.users.find_one({"id": token_data.user_id})
    if not user or not user.get("business_id"):
        raise HTTPException(status_code=404, detail="Business not found")
    
    business = await db.businesses.find_one({"id": user["business_id"]})
    if not business or not business.get("stripe_subscription_id"):
        raise HTTPException(status_code=400, detail="No active subscription")
    
    try:
        # Cancel at end of period (not immediately)
        stripe_lib.Subscription.modify(
            business["stripe_subscription_id"],
            cancel_at_period_end=True
        )
        
        await db.businesses.update_one(
            {"id": business["id"]},
            {"$set": {"subscription_cancel_requested": True}}
        )
        
        return {"message": "Subscription will be canceled at the end of the current billing period"}
    except Exception as e:
        logger.error(f"Stripe cancel error: {e}")
        raise HTTPException(status_code=500, detail="Error canceling subscription")


@router.post("/me/subscription/billing-portal")
async def create_billing_portal_session(request: Request, token_data: TokenData = Depends(require_business)):
    """Phase D — Generate a Stripe Customer Portal session so the business
    can update its payment method, view invoices, etc."""
    user = await db.users.find_one({"id": token_data.user_id})
    if not user or not user.get("business_id"):
        raise HTTPException(status_code=404, detail="Business not found")
    business = await db.businesses.find_one({"id": user["business_id"]})
    if not business or not business.get("stripe_customer_id"):
        raise HTTPException(status_code=400, detail="No Stripe customer found")
    origin = request.headers.get("origin") or os.environ.get("FRONTEND_URL", "https://bookvia.app")
    return_url = f"{origin.rstrip('/')}/business/finance"
    try:
        session = stripe_lib.billing_portal.Session.create(
            customer=business["stripe_customer_id"],
            return_url=return_url,
        )
        return {"url": session.url}
    except Exception as e:
        logger.error(f"Stripe billing portal error: {e}")
        raise HTTPException(status_code=500, detail="Error creating billing portal session")




@router.get("/me/closures")
async def get_business_closures(token_data: TokenData = Depends(require_business)):
    """Get all closure dates for the authenticated business."""
    user = await db.users.find_one({"id": token_data.user_id})
    if not user or not user.get("business_id"):
        raise HTTPException(status_code=404, detail="Business not found")
    closures = await db.business_closures.find(
        {"business_id": user["business_id"], "is_deleted": False},
        {"_id": 0}
    ).to_list(500)
    return closures



@router.post("/me/closures")
async def add_business_closure(data: ClosureDateCreate, token_data: TokenData = Depends(require_business)):
    """Mark a date as closed."""
    user = await db.users.find_one({"id": token_data.user_id})
    if not user or not user.get("business_id"):
        raise HTTPException(status_code=404, detail="Business not found")

    # Check if date already marked
    existing = await db.business_closures.find_one({
        "business_id": user["business_id"], "date": data.date, "is_deleted": False
    })
    if existing:
        raise HTTPException(status_code=400, detail="Date already marked as closed")

    closure = {
        "id": generate_id(),
        "business_id": user["business_id"],
        "date": data.date,
        "reason": data.reason,
        "is_deleted": False,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.business_closures.insert_one(closure)
    del closure["_id"]
    return closure



@router.delete("/me/closures/{date}")
async def remove_business_closure(date: str, token_data: TokenData = Depends(require_business)):
    """Remove a closure date (re-open)."""
    user = await db.users.find_one({"id": token_data.user_id})
    if not user or not user.get("business_id"):
        raise HTTPException(status_code=404, detail="Business not found")
    result = await db.business_closures.update_one(
        {"business_id": user["business_id"], "date": date, "is_deleted": False},
        {"$set": {"is_deleted": True}}
    )
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Closure not found")
    return {"message": "Closure removed"}



@router.get("/me/hours")
async def get_business_hours(token_data: TokenData = Depends(require_business)):
    """Get business opening hours."""
    user = await db.users.find_one({"id": token_data.user_id})
    if not user or not user.get("business_id"):
        raise HTTPException(status_code=404, detail="Business not found")
    business = await db.businesses.find_one({"id": user["business_id"]}, {"_id": 0, "business_hours": 1})
    hours = business.get("business_hours") if business else None
    if not hours:
        # Return default: Mon-Fri 9-18, Sat-Sun closed
        hours = {}
        for d in range(5):
            hours[str(d)] = {"is_open": True, "open_time": "09:00", "close_time": "18:00"}
        for d in range(5, 7):
            hours[str(d)] = {"is_open": False, "open_time": "09:00", "close_time": "18:00"}
    return hours


@router.put("/me/hours")
async def update_business_hours(hours: Dict[str, Any], token_data: TokenData = Depends(require_business)):
    """Update business opening hours. Keys: 0-6 (Mon-Sun). Values: {is_open, open_time, close_time}."""
    user = await db.users.find_one({"id": token_data.user_id})
    if not user or not user.get("business_id"):
        raise HTTPException(status_code=404, detail="Business not found")
    # Validate
    for day_key, day_val in hours.items():
        if day_key not in [str(i) for i in range(7)]:
            raise HTTPException(status_code=400, detail=f"Invalid day key: {day_key}")
        if not isinstance(day_val, dict) or "is_open" not in day_val:
            raise HTTPException(status_code=400, detail=f"Invalid data for day {day_key}")
    await db.businesses.update_one(
        {"id": user["business_id"]},
        {"$set": {"business_hours": hours}}
    )
    # Also update all workers' schedules to match
    workers = await db.workers.find(
        {"business_id": user["business_id"], "active": True},
        {"_id": 0, "id": 1}
    ).to_list(100)
    for w in workers:
        new_schedule = {}
        for day_key, day_val in hours.items():
            if day_val.get("is_open"):
                new_schedule[day_key] = {
                    "is_available": True,
                    "blocks": [{"start_time": day_val.get("open_time", "09:00"), "end_time": day_val.get("close_time", "18:00")}]
                }
            else:
                new_schedule[day_key] = {"is_available": False, "blocks": []}
        await db.workers.update_one({"id": w["id"]}, {"$set": {"schedule": new_schedule}})
    return {"message": "Hours updated", "business_hours": hours}




@router.post("/me/photos")
async def upload_business_photo(file: UploadFile = File(...), token_data: TokenData = Depends(require_business)):
    """Upload a gallery photo for the authenticated business."""
    user = await db.users.find_one({"id": token_data.user_id})
    if not user or not user.get("business_id"):
        raise HTTPException(status_code=404, detail="Business not found")

    data = await file.read()
    ok, err = validate_image(file.filename, file.content_type, len(data))
    if not ok:
        raise HTTPException(status_code=400, detail=err)

    business_id = user["business_id"]

    # Check photo limit (max 10)
    current_photos = await db.business_photos.count_documents({"business_id": business_id, "is_deleted": {"$ne": True}})
    if current_photos >= 10:
        raise HTTPException(status_code=400, detail="Maximum 10 photos allowed")

    # Try Cloudinary first, fallback to Emergent Storage
    if cloudinary_configured():
        try:
            result = upload_image(data, "business_gallery", business_id)
            photo_url = result["secure_url"]
            public_id = result["public_id"]
        except Exception as e:
            logger.error(f"Cloudinary upload failed: {e}")
            raise HTTPException(status_code=500, detail="Failed to upload photo")
    else:
        # Fallback: Emergent Object Storage (preview env)
        ext = file.filename.split(".")[-1].lower() if "." in file.filename else "jpg"
        content_type = file.content_type
        if ext in ("jfif", "jpg", "jpeg", "pjpeg"):
            content_type = "image/jpeg"
        path = generate_upload_path(business_id, file.filename)
        try:
            result = put_object(path, data, content_type)
            base_url = os.environ.get("BASE_URL", "")
            photo_url = f"{base_url}/api/files/{result['path']}"
            public_id = path
        except Exception as e:
            logger.error(f"Storage upload failed: {e}")
            raise HTTPException(status_code=500, detail="Failed to upload photo")

    # Store reference in DB
    file_record = {
        "id": generate_id(),
        "business_id": business_id,
        "url": photo_url,
        "public_id": public_id,
        "storage": "cloudinary" if cloudinary_configured() else "emergent",
        "original_filename": file.filename,
        "content_type": file.content_type,
        "is_deleted": False,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.business_photos.insert_one(file_record)

    # Add to business photos array
    await db.businesses.update_one(
        {"id": business_id},
        {"$push": {"photos": photo_url}}
    )

    return {
        "id": file_record["id"],
        "url": photo_url,
        "public_id": public_id,
        "original_filename": file.filename
    }




@router.delete("/me/photos/{photo_id}")
async def delete_business_photo(photo_id: str, token_data: TokenData = Depends(require_business)):
    """Delete a business photo."""
    user = await db.users.find_one({"id": token_data.user_id})
    if not user or not user.get("business_id"):
        raise HTTPException(status_code=404, detail="Business not found")

    photo = await db.business_photos.find_one({"id": photo_id, "business_id": user["business_id"], "is_deleted": False})
    if not photo:
        raise HTTPException(status_code=404, detail="Photo not found")

    # Delete from cloud storage
    if photo.get("storage") == "cloudinary" and photo.get("public_id"):
        cloudinary_delete(photo["public_id"])

    await db.business_photos.update_one({"id": photo_id}, {"$set": {"is_deleted": True}})

    # Remove from business photos array - handle legacy photos with storage_path instead of url
    photo_url = photo.get("url") or photo.get("storage_path")
    if photo_url:
        # For legacy photos, construct URL from storage_path
        if not photo_url.startswith("http"):
            base_url = os.environ.get("BASE_URL", "")
            photo_url = f"{base_url}/api/files/{photo_url}"
        await db.businesses.update_one(
            {"id": user["business_id"]},
            {"$pull": {"photos": photo_url}}
        )

    return {"message": "Photo deleted"}




@router.get("/me/photos")
async def get_business_photos(token_data: TokenData = Depends(require_business)):
    """Get all photos for the authenticated business."""
    user = await db.users.find_one({"id": token_data.user_id})
    if not user or not user.get("business_id"):
        raise HTTPException(status_code=404, detail="Business not found")

    photos = await db.business_photos.find(
        {"business_id": user["business_id"], "is_deleted": False},
        {"_id": 0}
    ).to_list(100)

    # Normalize: ensure all photos have 'url' field (handle legacy photos with storage_path)
    base_url = os.environ.get("BASE_URL", "")
    for photo in photos:
        if "url" not in photo and "storage_path" in photo:
            photo["url"] = f"{base_url}/api/files/{photo['storage_path']}"

    return photos




@router.post("/me/logo")
async def upload_business_logo(file: UploadFile = File(...), token_data: TokenData = Depends(require_business)):
    """Upload or replace the business logo."""
    user = await db.users.find_one({"id": token_data.user_id})
    if not user or not user.get("business_id"):
        raise HTTPException(status_code=404, detail="Business not found")

    data = await file.read()
    ok, err = validate_image(file.filename, file.content_type, len(data))
    if not ok:
        raise HTTPException(status_code=400, detail=err)

    business_id = user["business_id"]
    business = await db.businesses.find_one({"id": business_id})

    if cloudinary_configured():
        # Delete old logo if exists
        old_public_id = business.get("logo_public_id") if business else None
        if old_public_id:
            cloudinary_delete(old_public_id)

        try:
            result = upload_image(data, "business_logo", business_id)
            logo_url = result["secure_url"]
            public_id = result["public_id"]
        except Exception as e:
            logger.error(f"Logo upload failed: {e}")
            raise HTTPException(status_code=500, detail="Failed to upload logo")
    else:
        # Fallback: Emergent storage
        path = generate_upload_path(business_id, f"logo_{file.filename}")
        content_type = "image/jpeg" if file.filename.lower().endswith(("jfif", "jpg", "jpeg")) else file.content_type
        try:
            result = put_object(path, data, content_type)
            base_url = os.environ.get("BASE_URL", "")
            logo_url = f"{base_url}/api/files/{result['path']}"
            public_id = path
        except Exception as e:
            logger.error(f"Logo upload failed: {e}")
            raise HTTPException(status_code=500, detail="Failed to upload logo")

    await db.businesses.update_one(
        {"id": business_id},
        {"$set": {"logo_url": logo_url, "logo_public_id": public_id}}
    )

    return {"secure_url": logo_url, "public_id": public_id}




