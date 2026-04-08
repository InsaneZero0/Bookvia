from fastapi import FastAPI, APIRouter, HTTPException, Depends, Request, BackgroundTasks, status, File, UploadFile
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import Response
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict, EmailStr, validator
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone, timedelta
import jwt
import bcrypt
import pyotp
import qrcode
import io
import base64
import random
import re
import pytz
from bson import ObjectId
from enum import Enum

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# JWT Config
JWT_SECRET = os.environ.get('JWT_SECRET', 'bookvia-secret-key-change-in-production')
JWT_ALGORITHM = 'HS256'
JWT_EXPIRATION_HOURS = 24

# SMS Config (mock by default)
SMS_PROVIDER = os.environ.get('SMS_PROVIDER', 'mock')  # 'mock' or 'twilio'
TWILIO_ACCOUNT_SID = os.environ.get('TWILIO_ACCOUNT_SID', '')
TWILIO_AUTH_TOKEN = os.environ.get('TWILIO_AUTH_TOKEN', '')
TWILIO_PHONE_NUMBER = os.environ.get('TWILIO_PHONE_NUMBER', '')

# Stripe Config
STRIPE_API_KEY = os.environ.get('STRIPE_API_KEY', 'sk_test_emergent')

# Admin Config (from environment - NEVER hardcode)
ADMIN_EMAIL = os.environ.get('ADMIN_EMAIL')
ADMIN_INITIAL_PASSWORD = os.environ.get('ADMIN_INITIAL_PASSWORD')

# Environment
ENV = os.environ.get('ENV', 'development')

# Create the main app
app = FastAPI(title="Bookvia API", version="1.0.0")

# Create routers
api_router = APIRouter(prefix="/api")
auth_router = APIRouter(prefix="/auth", tags=["Authentication"])
users_router = APIRouter(prefix="/users", tags=["Users"])
businesses_router = APIRouter(prefix="/businesses", tags=["Businesses"])
services_router = APIRouter(prefix="/services", tags=["Services"])
bookings_router = APIRouter(prefix="/bookings", tags=["Bookings"])
reviews_router = APIRouter(prefix="/reviews", tags=["Reviews"])
categories_router = APIRouter(prefix="/categories", tags=["Categories"])
payments_router = APIRouter(prefix="/payments", tags=["Payments"])
admin_router = APIRouter(prefix="/admin", tags=["Admin"])
notifications_router = APIRouter(prefix="/notifications", tags=["Notifications"])
finance_router = APIRouter(prefix="/business/finance", tags=["Business Finance"])
settlements_router = APIRouter(prefix="/settlements", tags=["Settlements"])

security = HTTPBearer(auto_error=False)

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ========================== HEALTH CHECK ==========================

@api_router.get("/health", tags=["System"])
async def health_check():
    """Health check endpoint with configuration status"""
    # Check database connection
    try:
        await db.command("ping")
        db_status = "connected"
    except Exception as e:
        db_status = f"error: {str(e)}"
    
    # Check service configurations
    twilio_configured = all([
        os.environ.get("TWILIO_ACCOUNT_SID"),
        os.environ.get("TWILIO_AUTH_TOKEN"),
        os.environ.get("TWILIO_PHONE_NUMBER")
    ])
    
    resend_configured = bool(os.environ.get("RESEND_API_KEY"))
    
    stripe_key = os.environ.get("STRIPE_API_KEY", "")
    stripe_status = "not configured"
    if stripe_key:
        stripe_status = "live" if stripe_key.startswith("sk_live_") else "test"
    
    return {
        "status": "healthy" if db_status == "connected" else "degraded",
        "version": "1.0.0",
        "environment": ENV,
        "database": db_status,
        "config": {
            "sms": "twilio" if twilio_configured else "mock",
            "email": "resend" if resend_configured else "mock",
            "stripe": stripe_status,
            "base_url": os.environ.get("BASE_URL", "auto-detect")
        }
    }

# ========================== ENUMS ==========================

class UserRole(str, Enum):
    USER = "user"
    BUSINESS = "business"
    ADMIN = "admin"

class AppointmentStatus(str, Enum):
    HOLD = "hold"  # Slot reservado, esperando pago (30 min)
    CONFIRMED = "confirmed"  # Pago completado
    COMPLETED = "completed"  # Cita finalizada
    CANCELLED = "cancelled"  # Cancelada por cliente o negocio
    NO_SHOW = "no_show"  # Cliente no asistió
    EXPIRED = "expired"  # Hold expirado sin pago

class BusinessStatus(str, Enum):
    PENDING = "pending"
    APPROVED = "approved"
    SUSPENDED = "suspended"
    REJECTED = "rejected"

class PaymentStatus(str, Enum):
    PENDING = "pending"
    COMPLETED = "completed"
    REFUNDED = "refunded"
    FAILED = "failed"

class TransactionStatus(str, Enum):
    CREATED = "created"  # Checkout creado / hold activo
    PAID = "paid"  # Confirmado por webhook
    REFUND_PARTIAL = "refund_partial"  # Cliente cancela >24h, se devuelve anticipo-fee
    REFUND_FULL = "refund_full"  # Negocio cancela, se devuelve 100% al cliente
    NO_SHOW_PAYOUT = "no_show_payout"  # Cliente no asiste, negocio recibe anticipo-fee
    BUSINESS_CANCEL_FEE = "business_cancel_fee"  # Negocio cancela, se le cobra 8%
    EXPIRED = "expired"  # Hold expirado sin pago

# Platform fee constant
PLATFORM_FEE_PERCENT = 0.08  # 8%
HOLD_EXPIRATION_MINUTES = 30
MIN_DEPOSIT_AMOUNT = 50.0  # MXN

# Visibility filter: businesses visible to public must be approved + subscription active/trialing
# Also include businesses without subscription_status field (legacy/backwards compat)
VISIBLE_BUSINESS_FILTER = {
    "status": BusinessStatus.APPROVED,
    "$or": [
        {"subscription_status": {"$in": ["active", "trialing"]}},
        {"subscription_status": {"$exists": False}},
        {"subscription_status": None},
        {"subscription_status": "none"},
    ]
}

# Ledger Enums
class LedgerDirection(str, Enum):
    DEBIT = "debit"
    CREDIT = "credit"

class LedgerAccount(str, Enum):
    BUSINESS_REVENUE = "business_revenue"
    PLATFORM_FEE = "platform_fee"
    REFUND = "refund"
    PENALTY = "penalty"
    PAYOUT = "payout"

class LedgerEntryStatus(str, Enum):
    POSTED = "posted"
    REVERSED = "reversed"

class SettlementStatus(str, Enum):
    PENDING = "pending"
    PAID = "paid"
    HELD = "held"
    FAILED = "failed"

class AuditAction(str, Enum):
    BUSINESS_APPROVE = "business_approve"
    BUSINESS_REJECT = "business_reject"
    BUSINESS_SUSPEND = "business_suspend"
    USER_SUSPEND = "user_suspend"
    REVIEW_DELETE = "review_delete"
    BUSINESS_FEATURE = "business_feature"
    PAYMENT_HOLD = "payment_hold"
    PAYMENT_RELEASE = "payment_release"
    ADMIN_LOGIN = "admin_login"
    ADMIN_2FA_SETUP = "admin_2fa_setup"
    PAYOUT_HOLD = "payout_hold"
    PAYOUT_RELEASE = "payout_release"
    SETTLEMENT_GENERATE = "settlement_generate"
    SETTLEMENT_MARK_PAID = "settlement_mark_paid"

# ========================== MODELS ==========================

class TokenData(BaseModel):
    user_id: str
    role: UserRole
    email: str
    worker_id: Optional[str] = None
    is_manager: bool = False

# User Models
class UserCreate(BaseModel):
    email: EmailStr
    password: str
    full_name: str
    phone: str
    country: Optional[str] = None
    city: Optional[str] = None
    birth_date: Optional[str] = None
    gender: Optional[str] = None
    photo_url: Optional[str] = None
    preferred_language: str = "es"

class UserLogin(BaseModel):
    email: EmailStr
    password: str

class UserResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    email: str
    full_name: str
    phone: str
    phone_verified: bool = False
    email_verified: bool = False
    birth_date: Optional[str] = None
    gender: Optional[str] = None
    photo_url: Optional[str] = None
    role: str = "user"
    business_id: Optional[str] = None  # For business users
    active_appointments_count: int = 0
    cancellation_count: int = 0
    suspended_until: Optional[str] = None
    favorites: List[str] = []
    preferred_language: str = "es"
    totp_enabled: bool = False
    created_at: str

class UserUpdate(BaseModel):
    full_name: Optional[str] = None
    phone: Optional[str] = None
    birth_date: Optional[str] = None
    gender: Optional[str] = None
    photo_url: Optional[str] = None
    preferred_language: Optional[str] = None

# Phone Verification
class PhoneVerifyRequest(BaseModel):
    phone: str

class PhoneVerifyConfirm(BaseModel):
    phone: str
    code: str

# Business Models
class BusinessCreate(BaseModel):
    name: str
    email: EmailStr
    password: str
    phone: str
    description: str
    category_id: str
    address: str
    city: str
    state: str
    country: str = "MX"
    zip_code: str
    latitude: Optional[float] = None
    longitude: Optional[float] = None
    timezone: str = "America/Mexico_City"  # IANA timezone (required)
    # Legal documents
    ine_url: Optional[str] = None
    rfc: str
    proof_of_address_url: Optional[str] = None
    clabe: str
    legal_name: str
    # Owner info
    owner_birth_date: Optional[str] = None
    # Business settings
    requires_deposit: bool = False
    deposit_amount: float = 50.0
    cancellation_days: int = 1
    payout_schedule: Optional[str] = "monthly"  # triday, biweekly, monthly
    min_time_between_appointments: int = 0  # minutes (buffer between appointments)
    service_radius_km: Optional[float] = None  # for home service
    plan_type: str = "basic"  # basic, premium

class BusinessResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    name: str
    email: str
    phone: str
    phone_verified: bool = False
    description: str
    category_id: str
    category_name: Optional[str] = None
    address: str
    city: str
    city_slug: Optional[str] = None
    state: str
    country: str
    country_code: str = "MX"
    zip_code: str
    latitude: Optional[float] = None
    longitude: Optional[float] = None
    timezone: str = "America/Mexico_City"
    status: str = "pending"
    rating: float = 0.0
    review_count: int = 0
    completed_appointments: int = 0
    badges: List[str] = []
    requires_deposit: bool = False
    deposit_amount: float = 50.0
    cancellation_days: int = 1
    payout_schedule: Optional[str] = "monthly"
    min_time_between_appointments: int = 0
    photos: List[str] = []
    logo_url: Optional[str] = None
    slug: Optional[str] = None
    created_at: str
    is_featured: bool = False
    plan_type: str = "basic"
    trial_ends_at: Optional[str] = None
    can_accept_bookings: bool = True
    subscription_status: str = "none"
    stripe_customer_id: Optional[str] = None
    stripe_subscription_id: Optional[str] = None
    logo_url: Optional[str] = None
    logo_public_id: Optional[str] = None
    distance_km: Optional[float] = None
    next_available_text: Optional[str] = None

class BusinessUpdate(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None
    address: Optional[str] = None
    city: Optional[str] = None
    state: Optional[str] = None
    zip_code: Optional[str] = None
    latitude: Optional[float] = None
    longitude: Optional[float] = None
    timezone: Optional[str] = None
    requires_deposit: Optional[bool] = None
    deposit_amount: Optional[float] = None
    cancellation_days: Optional[int] = None
    payout_schedule: Optional[str] = None
    min_time_between_appointments: Optional[int] = None
    service_radius_km: Optional[float] = None
    photos: Optional[List[str]] = None
    logo_url: Optional[str] = None

# Worker Models

class ScheduleBlock(BaseModel):
    """A time block within a day (e.g., morning shift, lunch break)"""
    start_time: str  # "09:00"
    end_time: str    # "14:00"

class DaySchedule(BaseModel):
    """Schedule for a single day with multiple blocks"""
    is_available: bool = True
    blocks: List[ScheduleBlock] = []  # Multiple blocks per day (e.g., 09:00-14:00, 16:00-20:00)

class WorkerException(BaseModel):
    """Exception (vacation/block) with date range support"""
    start_date: str  # "2024-01-15"
    end_date: str    # "2024-01-15" (same day) or "2024-01-20" (range)
    start_time: Optional[str] = None  # "09:00" - if null, full day
    end_time: Optional[str] = None    # "12:00" - if null, full day
    reason: Optional[str] = None      # "Vacaciones", "Cita médica", etc.
    exception_type: str = "block"     # "vacation" | "block"

class WorkerCreate(BaseModel):
    name: str
    email: Optional[str] = None
    phone: Optional[str] = None
    photo_url: Optional[str] = None
    bio: Optional[str] = None
    service_ids: List[str] = []

class WorkerUpdate(BaseModel):
    name: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    photo_url: Optional[str] = None
    bio: Optional[str] = None
    service_ids: Optional[List[str]] = None

class WorkerResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    business_id: str
    name: str
    email: Optional[str] = None
    phone: Optional[str] = None
    photo_url: Optional[str] = None
    bio: Optional[str] = None
    service_ids: List[str] = []
    schedule: Dict[str, Any] = {}  # {"0": {"is_available": true, "blocks": [{"start_time": "09:00", "end_time": "18:00"}]}}
    exceptions: List[Dict[str, Any]] = []  # WorkerException as dict
    active: bool = True
    is_manager: bool = False
    manager_permissions: Optional[Dict[str, bool]] = None
    manager_designated_at: Optional[str] = None
    has_manager_pin: bool = False
    created_at: Optional[str] = None
    deactivated_at: Optional[str] = None

# ============================== Manager / PIN Models ==============================

DEFAULT_MANAGER_PERMISSIONS = {
    "complete_bookings": True,
    "reschedule_bookings": True,
    "cancel_bookings": False,
    "block_clients": False,
    "view_client_data": False,
    "edit_services": False,
    "view_reports": False,
    "view_today_bookings": True,
    "view_confirmed_bookings": True,
    "view_agenda": True,
    "view_team": False,
    "edit_photos": False,
    "edit_description": False,
    "edit_schedule": False,
    "edit_contact": False,
}

class PinCreate(BaseModel):
    pin: str  # 4-6 digits

class PinVerify(BaseModel):
    pin: str

class ManagerLogin(BaseModel):
    business_email: EmailStr
    worker_id: str
    pin: str

class ManagerDesignate(BaseModel):
    permissions: Dict[str, bool] = {}

class ManagerPermissionsUpdate(BaseModel):
    permissions: Dict[str, bool]

class WorkerScheduleUpdate(BaseModel):
    """Update schedule for multiple days"""
    schedule: Dict[str, DaySchedule]  # {"0": DaySchedule, "1": DaySchedule, ...}

class WorkerExceptionAdd(BaseModel):
    """Add an exception (vacation/block)"""
    exception: WorkerException

# Service Models
class ServiceCreate(BaseModel):
    name: str
    description: Optional[str] = None
    duration_minutes: int = 60
    price: float
    category_id: Optional[str] = None
    is_home_service: bool = False
    allowed_worker_ids: List[str] = []  # Empty = all workers can perform this service

class ServiceUpdate(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None
    duration_minutes: Optional[int] = None
    price: Optional[float] = None
    category_id: Optional[str] = None
    is_home_service: Optional[bool] = None
    allowed_worker_ids: Optional[List[str]] = None

class ServiceResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    business_id: str
    name: str
    description: Optional[str] = None
    duration_minutes: int
    price: float
    category_id: Optional[str] = None
    is_home_service: bool = False
    allowed_worker_ids: List[str] = []
    active: bool = True

# Booking Models
class BookingCreate(BaseModel):
    business_id: str
    service_id: str
    worker_id: Optional[str] = None
    date: str  # "2024-01-15"
    time: str  # "10:00"
    notes: Optional[str] = None
    is_home_service: bool = False
    address: Optional[str] = None
    # Business-created booking fields
    skip_payment: bool = False
    client_name: Optional[str] = None
    client_email: Optional[str] = None
    client_phone: Optional[str] = None
    client_info: Optional[str] = None

class BookingResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    user_id: str
    business_id: str
    service_id: str
    worker_id: str
    date: str
    time: str
    end_time: str
    duration_minutes: int = 60
    status: str
    notes: Optional[str] = None
    is_home_service: bool = False
    address: Optional[str] = None
    deposit_amount: Optional[float] = None
    deposit_paid: bool = False
    payment_id: Optional[str] = None
    transaction_id: Optional[str] = None
    stripe_session_id: Optional[str] = None
    hold_expires_at: Optional[str] = None
    total_amount: Optional[float] = None
    created_at: str
    confirmed_at: Optional[str] = None
    cancelled_at: Optional[str] = None
    cancelled_by: Optional[str] = None
    cancellation_reason: Optional[str] = None
    # Populated fields
    business_name: Optional[str] = None
    service_name: Optional[str] = None
    worker_name: Optional[str] = None
    user_name: Optional[str] = None
    user_email: Optional[str] = None
    user_phone: Optional[str] = None
    # Client data (for business-created bookings)
    client_name: Optional[str] = None
    client_email: Optional[str] = None
    client_phone: Optional[str] = None
    client_info: Optional[str] = None
    # Computed fields
    can_cancel: bool = True
    hours_until_appointment: Optional[float] = None
    has_review: bool = False
    business_slug: Optional[str] = None

# Review Models
class ReviewCreate(BaseModel):
    business_id: str
    booking_id: str
    rating: int  # 1-5
    comment: Optional[str] = None

class ReviewResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    user_id: str
    business_id: str
    booking_id: str
    rating: int
    comment: Optional[str] = None
    created_at: str
    user_name: Optional[str] = None
    user_photo: Optional[str] = None

# Category Models
class CategoryCreate(BaseModel):
    name_es: str
    name_en: str
    slug: str
    icon: str
    image_url: Optional[str] = None
    parent_id: Optional[str] = None

class CategoryResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    name_es: str
    name_en: str
    slug: str
    icon: str
    image_url: Optional[str] = None
    parent_id: Optional[str] = None
    business_count: int = 0

# Blacklist Models
class BlacklistEntry(BaseModel):
    email: Optional[str] = None
    phone: Optional[str] = None
    user_id: Optional[str] = None
    reason: Optional[str] = None

class BlacklistResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    business_id: str
    email: Optional[str] = None
    phone: Optional[str] = None
    user_id: Optional[str] = None
    reason: Optional[str] = None
    created_at: str

# Payment Models
class PaymentCreate(BaseModel):
    amount: float
    currency: str = "MXN"
    booking_id: Optional[str] = None
    subscription_type: Optional[str] = None  # monthly, monthly_deposit

class PaymentResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    user_id: Optional[str] = None
    business_id: Optional[str] = None
    booking_id: Optional[str] = None
    amount: float
    currency: str
    status: str
    stripe_session_id: Optional[str] = None
    payment_type: str  # deposit, subscription, refund
    created_at: str

# Transaction Models (Bookvia-specific)
class TransactionCreate(BaseModel):
    booking_id: str
    amount: float
    currency: str = "MXN"

class TransactionResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    booking_id: str
    user_id: str
    business_id: str
    stripe_session_id: Optional[str] = None
    stripe_payment_intent_id: Optional[str] = None
    amount_total: float
    fee_amount: float  # 8% platform fee
    payout_amount: float  # Amount for business (amount - fee)
    currency: str = "MXN"
    status: str  # TransactionStatus enum
    refund_amount: Optional[float] = None
    refund_reason: Optional[str] = None
    cancelled_by: Optional[str] = None  # "user" or "business"
    created_at: str
    updated_at: Optional[str] = None
    paid_at: Optional[str] = None
    
# Deposit checkout request
class DepositCheckoutRequest(BaseModel):
    booking_id: str

# Cancel booking request
class CancelBookingRequest(BaseModel):
    reason: Optional[str] = None

# ========================== LEDGER MODELS ==========================

class LedgerEntryResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    transaction_id: str
    booking_id: Optional[str] = None
    business_id: str
    direction: str  # DEBIT | CREDIT
    account: str  # business_revenue, platform_fee, refund, penalty, payout
    amount_cents: int  # Integer to avoid decimal errors
    amount: float  # Decimal for display
    currency: str = "MXN"
    country: str = "MX"
    entry_status: str = "posted"  # posted | reversed
    description: Optional[str] = None
    related_appointment_id: Optional[str] = None
    created_by: str = "system"  # system | admin
    created_at: str

# ========================== SETTLEMENT MODELS ==========================

class SettlementResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    business_id: str
    business_name: Optional[str] = None
    period_key: str  # "2026-02" or "MX-2026-02"
    period_start: str
    period_end: str
    # Calculated from ledger
    gross_paid: float  # sum of business_revenue CREDIT
    total_fees: float  # sum of platform_fee DEBIT
    total_refunds: float  # sum of refund DEBIT
    total_penalties: float  # sum of penalty DEBIT (BUSINESS_CANCEL_FEE)
    net_payout: float  # gross_paid - fees - refunds - penalties
    currency: str = "MXN"
    country: str = "MX"
    status: str  # pending, paid, held, failed
    held_reason: Optional[str] = None
    idempotency_key: str  # job run identifier
    payout_reference: Optional[str] = None
    paid_at: Optional[str] = None
    created_at: str
    updated_at: Optional[str] = None

class SettlementMarkPaidRequest(BaseModel):
    payout_reference: str

class BusinessFinanceSummary(BaseModel):
    gross_revenue: float
    total_fees: float
    total_refunds: float
    total_penalties: float
    net_earnings: float
    pending_payout: float
    paid_payout: float
    held_payout: float
    next_settlement_date: Optional[str] = None
    currency: str = "MXN"

class PayoutHoldRequest(BaseModel):
    hold: bool
    reason: Optional[str] = None

# Admin 2FA
class Admin2FASetup(BaseModel):
    password: str

class Admin2FAVerify(BaseModel):
    code: str

class AdminLogin(BaseModel):
    email: EmailStr
    password: str
    totp_code: str

# Audit Log Models
class AuditLogResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    admin_id: str
    admin_email: str
    action: str
    target_type: str  # business, user, review, payment
    target_id: str
    details: Optional[Dict[str, Any]] = None
    ip_address: Optional[str] = None
    user_agent: Optional[str] = None
    created_at: str

# Notification Models
class NotificationResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    user_id: str
    title: str
    message: str
    type: str  # booking, reminder, system
    read: bool = False
    created_at: str
    data: Optional[Dict[str, Any]] = None

# Search Models
class SearchQuery(BaseModel):
    query: Optional[str] = None
    category_id: Optional[str] = None
    city: Optional[str] = None
    date: Optional[str] = None
    min_rating: Optional[float] = None
    is_home_service: Optional[bool] = None
    page: int = 1
    limit: int = 20

# ========================== HELPERS ==========================

def generate_id() -> str:
    return str(uuid.uuid4())

def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

def verify_password(password: str, hashed: str) -> bool:
    return bcrypt.checkpw(password.encode('utf-8'), hashed.encode('utf-8'))

def create_token(user_id: str, role: str, email: str, worker_id: str = None, is_manager: bool = False) -> str:
    payload = {
        "user_id": user_id,
        "role": role,
        "email": email,
        "exp": datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRATION_HOURS)
    }
    if worker_id:
        payload["worker_id"] = worker_id
        payload["is_manager"] = True
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

def decode_token(token: str) -> Optional[TokenData]:
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        return TokenData(**payload)
    except jwt.ExpiredSignatureError:
        return None
    except jwt.InvalidTokenError:
        return None

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)) -> Optional[TokenData]:
    if not credentials:
        return None
    return decode_token(credentials.credentials)

async def require_auth(credentials: HTTPAuthorizationCredentials = Depends(security)) -> TokenData:
    if not credentials:
        raise HTTPException(status_code=401, detail="Authentication required")
    token_data = decode_token(credentials.credentials)
    if not token_data:
        raise HTTPException(status_code=401, detail="Invalid or expired token")
    return token_data

async def require_admin(token_data: TokenData = Depends(require_auth)) -> TokenData:
    if token_data.role != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Admin access required")
    return token_data

async def require_business(token_data: TokenData = Depends(require_auth)) -> TokenData:
    if token_data.role not in [UserRole.BUSINESS, UserRole.ADMIN]:
        raise HTTPException(status_code=403, detail="Business access required")
    return token_data

async def is_user_blacklisted(business_id: str, user_id: str = None, email: str = None, phone: str = None) -> bool:
    """Check if a user is blacklisted by a business using any identifier."""
    if not user_id and not email and not phone:
        return False
    conditions = []
    # Build lookup data from user record if we have user_id
    lookup_emails = set()
    lookup_phones = set()
    lookup_user_ids = set()
    if user_id:
        lookup_user_ids.add(user_id)
        user_doc = await db.users.find_one({"id": user_id}, {"_id": 0, "email": 1, "phone": 1})
        if user_doc:
            if user_doc.get("email"):
                lookup_emails.add(user_doc["email"].lower())
            if user_doc.get("phone"):
                lookup_phones.add(user_doc["phone"])
    if email:
        lookup_emails.add(email.lower())
    if phone:
        lookup_phones.add(phone)
    
    or_conditions = []
    if lookup_user_ids:
        or_conditions.append({"user_id": {"$in": list(lookup_user_ids)}})
    if lookup_emails:
        or_conditions.append({"email": {"$in": [e.lower() for e in lookup_emails]}})
    if lookup_phones:
        or_conditions.append({"phone": {"$in": list(lookup_phones)}})
    
    if not or_conditions:
        return False
    
    entry = await db.blacklist.find_one({
        "business_id": business_id,
        "$or": or_conditions
    })
    return entry is not None

def generate_slug(name: str) -> str:
    slug = name.lower()
    slug = re.sub(r'[^a-z0-9\s-]', '', slug)
    slug = re.sub(r'[\s_]+', '-', slug)
    slug = re.sub(r'-+', '-', slug)
    return slug.strip('-')

def calculate_bayesian_rating(rating_sum: float, review_count: int, global_avg: float = 3.5, min_reviews: int = 5) -> float:
    """Calculate Bayesian average rating"""
    if review_count == 0:
        return 0.0
    return (min_reviews * global_avg + rating_sum) / (min_reviews + review_count)

async def send_sms(phone: str, message: str):
    """Send SMS - uses new SMS service with rate limiting and proper error handling"""
    from services.sms import send_sms as sms_send, SMSServiceError, SMSRateLimitError, SMSNotConfiguredError
    
    try:
        success, msg_id = await sms_send(phone, message)
        return success
    except SMSRateLimitError as e:
        logger.warning(f"SMS rate limit exceeded for {phone}: {e}")
        raise HTTPException(status_code=429, detail=str(e))
    except SMSNotConfiguredError as e:
        logger.error(f"SMS not configured: {e}")
        raise HTTPException(status_code=503, detail=str(e))
    except SMSServiceError as e:
        logger.error(f"SMS error: {e}")
        return False

async def create_notification(user_id: str, title: str, message: str, notif_type: str, data: dict = None):
    """Create internal notification"""
    notification = {
        "id": generate_id(),
        "user_id": user_id,
        "title": title,
        "message": message,
        "type": notif_type,
        "read": False,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "data": data or {}
    }
    await db.notifications.insert_one(notification)
    return notification

async def create_audit_log(
    admin_id: str,
    admin_email: str,
    action: str,
    target_type: str,
    target_id: str,
    details: dict = None,
    request: Request = None
):
    """Create audit log for admin actions"""
    audit_log = {
        "id": generate_id(),
        "admin_id": admin_id,
        "admin_email": admin_email,
        "action": action,
        "target_type": target_type,
        "target_id": target_id,
        "details": details or {},
        "ip_address": request.client.host if request else None,
        "user_agent": request.headers.get("user-agent") if request else None,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.audit_logs.insert_one(audit_log)
    logger.info(f"[AUDIT] {admin_email} - {action} - {target_type}:{target_id}")
    return audit_log

# ── Business Activity Log Helper ──
async def create_business_activity(
    business_id: str,
    token_data: TokenData,
    action: str,
    target_type: str,
    target_id: str,
    details: dict = None,
):
    """Log business actions by owner or administrator for audit trail"""
    if token_data.is_manager and token_data.worker_id:
        worker = await db.workers.find_one({"id": token_data.worker_id}, {"_id": 0, "name": 1})
        actor_type = "admin"
        actor_name = worker["name"] if worker else "Administrador"
    else:
        actor_type = "owner"
        actor_name = "Dueño"

    log_entry = {
        "id": generate_id(),
        "business_id": business_id,
        "actor_type": actor_type,
        "actor_name": actor_name,
        "worker_id": token_data.worker_id,
        "action": action,
        "target_type": target_type,
        "target_id": target_id,
        "details": details or {},
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.business_activity_logs.insert_one(log_entry)
    logger.info(f"[BIZ-ACTIVITY] {actor_name} ({actor_type}) - {action} - {target_type}:{target_id}")
    return log_entry

# ========================== LEDGER HELPERS ==========================

def amount_to_cents(amount: float) -> int:
    """Convert decimal amount to cents (integer)"""
    return int(round(amount * 100))

def cents_to_amount(cents: int) -> float:
    """Convert cents to decimal amount"""
    return round(cents / 100, 2)

async def create_ledger_entry(
    transaction_id: str,
    business_id: str,
    direction: str,
    account: str,
    amount: float,
    currency: str = "MXN",
    country: str = "MX",
    description: str = None,
    booking_id: str = None,
    created_by: str = "system"
) -> dict:
    """Create a ledger entry for double-entry bookkeeping"""
    entry = {
        "id": generate_id(),
        "transaction_id": transaction_id,
        "booking_id": booking_id,
        "business_id": business_id,
        "direction": direction,
        "account": account,
        "amount_cents": amount_to_cents(amount),
        "amount": round(amount, 2),
        "currency": currency,
        "country": country,
        "entry_status": LedgerEntryStatus.POSTED,
        "description": description,
        "related_appointment_id": booking_id,
        "created_by": created_by,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.ledger_entries.insert_one(entry)
    logger.info(f"[LEDGER] {direction} {amount} {currency} - {account} - business:{business_id}")
    return entry

async def create_transaction_ledger_entries(transaction: dict, status: str):
    """Create all ledger entries for a transaction based on its status"""
    tx_id = transaction["id"]
    business_id = transaction["business_id"]
    booking_id = transaction.get("booking_id")
    amount = transaction["amount_total"]
    fee = transaction["fee_amount"]
    payout = transaction["payout_amount"]
    
    if status == TransactionStatus.PAID:
        # Payment received: CREDIT business_revenue, DEBIT platform_fee
        await create_ledger_entry(
            tx_id, business_id, LedgerDirection.CREDIT, LedgerAccount.BUSINESS_REVENUE,
            amount, description="Anticipo recibido", booking_id=booking_id
        )
        await create_ledger_entry(
            tx_id, business_id, LedgerDirection.DEBIT, LedgerAccount.PLATFORM_FEE,
            fee, description="Comisión Bookvia 8%", booking_id=booking_id
        )
    
    elif status == TransactionStatus.REFUND_PARTIAL:
        # Partial refund (>24h): DEBIT refund (92% of original)
        refund_amount = transaction.get("refund_amount", payout)
        await create_ledger_entry(
            tx_id, business_id, LedgerDirection.DEBIT, LedgerAccount.REFUND,
            refund_amount, description="Reembolso parcial cliente >24h", booking_id=booking_id
        )
    
    elif status == TransactionStatus.REFUND_FULL:
        # Full refund (business cancels): DEBIT full amount
        await create_ledger_entry(
            tx_id, business_id, LedgerDirection.DEBIT, LedgerAccount.REFUND,
            amount, description="Reembolso completo (negocio canceló)", booking_id=booking_id
        )
    
    elif status == TransactionStatus.BUSINESS_CANCEL_FEE:
        # Business cancel penalty: DEBIT penalty
        penalty = transaction.get("fee_amount", fee)
        await create_ledger_entry(
            tx_id, business_id, LedgerDirection.DEBIT, LedgerAccount.PENALTY,
            penalty, description="Penalidad por cancelación del negocio 8%", booking_id=booking_id
        )
    
    elif status == TransactionStatus.NO_SHOW_PAYOUT:
        # No-show: business keeps payout (already have PAID entries, just status change)
        pass  # Ledger entries already created when PAID

async def calculate_business_ledger_summary(business_id: str, period_start: str = None, period_end: str = None) -> dict:
    """Calculate financial summary from ledger entries"""
    filters = {"business_id": business_id, "entry_status": LedgerEntryStatus.POSTED}
    
    if period_start and period_end:
        filters["created_at"] = {"$gte": period_start, "$lte": period_end}
    
    entries = await db.ledger_entries.find(filters, {"_id": 0}).to_list(10000)
    
    # Calculate totals by account and direction
    gross_revenue = 0  # business_revenue CREDIT
    total_fees = 0     # platform_fee DEBIT
    total_refunds = 0  # refund DEBIT
    total_penalties = 0  # penalty DEBIT
    total_payouts = 0   # payout CREDIT (when settled)
    
    for entry in entries:
        amount = entry["amount"]
        account = entry["account"]
        direction = entry["direction"]
        
        if account == LedgerAccount.BUSINESS_REVENUE and direction == LedgerDirection.CREDIT:
            gross_revenue += amount
        elif account == LedgerAccount.PLATFORM_FEE and direction == LedgerDirection.DEBIT:
            total_fees += amount
        elif account == LedgerAccount.REFUND and direction == LedgerDirection.DEBIT:
            total_refunds += amount
        elif account == LedgerAccount.PENALTY and direction == LedgerDirection.DEBIT:
            total_penalties += amount
        elif account == LedgerAccount.PAYOUT and direction == LedgerDirection.CREDIT:
            total_payouts += amount
    
    net_earnings = gross_revenue - total_fees - total_refunds - total_penalties
    
    return {
        "gross_revenue": round(gross_revenue, 2),
        "total_fees": round(total_fees, 2),
        "total_refunds": round(total_refunds, 2),
        "total_penalties": round(total_penalties, 2),
        "net_earnings": round(net_earnings, 2),
        "total_payouts": round(total_payouts, 2),
        "pending_payout": round(net_earnings - total_payouts, 2)
    }


# ========================== REPORTS ROUTES ==========================

@businesses_router.get("/my/reports")
async def get_business_reports(
    period: str = "month",
    token_data: TokenData = Depends(require_business)
):
    """Get business reports/analytics. period: week, month, quarter, year"""
    business = await db.businesses.find_one({"user_id": token_data.user_id}, {"_id": 0})
    if not business:
        raise HTTPException(status_code=404, detail="Business not found")
    
    business_id = business["id"]
    now = datetime.now(timezone.utc)
    
    period_days = {"week": 7, "month": 30, "quarter": 90, "year": 365}
    days = period_days.get(period, 30)
    start_date = now - timedelta(days=days)
    prev_start = start_date - timedelta(days=days)
    
    # Current period bookings
    current_bookings = await db.bookings.find(
        {"business_id": business_id, "created_at": {"$gte": start_date.isoformat()}},
        {"_id": 0}
    ).to_list(10000)
    
    # Previous period bookings (for comparison)
    prev_bookings = await db.bookings.find(
        {"business_id": business_id, "created_at": {"$gte": prev_start.isoformat(), "$lt": start_date.isoformat()}},
        {"_id": 0}
    ).to_list(10000)
    
    # Current period transactions (paid only)
    current_transactions = await db.transactions.find(
        {"business_id": business_id, "status": "paid", "created_at": {"$gte": start_date.isoformat()}},
        {"_id": 0}
    ).to_list(10000)
    
    prev_transactions = await db.transactions.find(
        {"business_id": business_id, "status": "paid", "created_at": {"$gte": prev_start.isoformat(), "$lt": start_date.isoformat()}},
        {"_id": 0}
    ).to_list(10000)
    
    # === BOOKING STATS ===
    completed = len([b for b in current_bookings if b.get("status") == "completed"])
    confirmed = len([b for b in current_bookings if b.get("status") == "confirmed"])
    cancelled = len([b for b in current_bookings if b.get("status") == "cancelled"])
    total_bookings = len(current_bookings)
    prev_total = len(prev_bookings)
    cancel_rate = round((cancelled / total_bookings * 100) if total_bookings > 0 else 0, 1)
    
    # === REVENUE ===
    current_revenue = sum(t.get("amount_total", 0) for t in current_transactions)
    prev_revenue = sum(t.get("amount_total", 0) for t in prev_transactions)
    revenue_change = round(((current_revenue - prev_revenue) / prev_revenue * 100) if prev_revenue > 0 else 0, 1)
    bookings_change = round(((total_bookings - prev_total) / prev_total * 100) if prev_total > 0 else 0, 1)
    
    # === REVENUE BY DAY (for chart) ===
    revenue_by_day = {}
    for t in current_transactions:
        day = t.get("created_at", "")[:10]
        if day:
            revenue_by_day[day] = revenue_by_day.get(day, 0) + t.get("amount_total", 0)
    
    bookings_by_day = {}
    for b in current_bookings:
        day = b.get("date", b.get("created_at", "")[:10])
        if day:
            bookings_by_day[day] = bookings_by_day.get(day, 0) + 1
    
    # Build daily chart data
    daily_chart = []
    for i in range(min(days, 30)):
        d = (now - timedelta(days=days - 1 - i)).strftime("%Y-%m-%d")
        daily_chart.append({
            "date": d,
            "revenue": round(revenue_by_day.get(d, 0), 2),
            "bookings": bookings_by_day.get(d, 0)
        })
    
    # === TOP SERVICES ===
    service_counts = {}
    service_revenue = {}
    for b in current_bookings:
        sid = b.get("service_id")
        if sid:
            service_counts[sid] = service_counts.get(sid, 0) + 1
            if b.get("status") in ("completed", "confirmed"):
                service_revenue[sid] = service_revenue.get(sid, 0) + (b.get("total_amount") or b.get("deposit_amount") or 0)
    
    service_ids = list(set(list(service_counts.keys()) + list(service_revenue.keys())))
    services_map = {}
    if service_ids:
        svcs = await db.services.find({"id": {"$in": service_ids}}, {"_id": 0, "id": 1, "name": 1}).to_list(100)
        services_map = {s["id"]: s["name"] for s in svcs}
    
    top_services = sorted(
        [{"name": services_map.get(sid, "Servicio"), "bookings": service_counts.get(sid, 0), "revenue": round(service_revenue.get(sid, 0), 2)}
         for sid in service_ids],
        key=lambda x: x["bookings"], reverse=True
    )[:5]
    
    # === TOP CLIENTS ===
    client_visits = {}
    client_spent = {}
    for b in current_bookings:
        uid = b.get("user_id")
        if uid and b.get("status") in ("completed", "confirmed"):
            client_visits[uid] = client_visits.get(uid, 0) + 1
            client_spent[uid] = client_spent.get(uid, 0) + (b.get("total_amount") or b.get("deposit_amount") or 0)
    
    top_client_ids = sorted(client_visits.keys(), key=lambda x: client_visits[x], reverse=True)[:5]
    top_clients = []
    for uid in top_client_ids:
        u = await db.users.find_one({"id": uid}, {"_id": 0, "full_name": 1, "email": 1})
        top_clients.append({
            "name": u.get("full_name", "Cliente") if u else "Cliente",
            "email": u.get("email", "") if u else "",
            "visits": client_visits[uid],
            "total_spent": round(client_spent.get(uid, 0), 2)
        })
    
    # === PEAK HOURS ===
    hour_counts = {}
    for b in current_bookings:
        t = b.get("time", "")
        if t:
            hour = t.split(":")[0]
            hour_counts[hour] = hour_counts.get(hour, 0) + 1
    
    peak_hours = sorted(
        [{"hour": f"{h}:00", "bookings": c} for h, c in hour_counts.items()],
        key=lambda x: x["bookings"], reverse=True
    )[:6]
    
    # === PEAK DAYS ===
    day_names_es = ["Lun", "Mar", "Mie", "Jue", "Vie", "Sab", "Dom"]
    day_counts = {i: 0 for i in range(7)}
    for b in current_bookings:
        d = b.get("date", "")
        if d:
            try:
                weekday = datetime.strptime(d, "%Y-%m-%d").weekday()
                day_counts[weekday] += 1
            except ValueError:
                pass
    
    peak_days = [{"day": day_names_es[i], "bookings": day_counts[i]} for i in range(7)]
    
    # === CANCELLATION BREAKDOWN ===
    cancelled_by_user = len([b for b in current_bookings if b.get("status") == "cancelled" and b.get("cancelled_by") == "user"])
    cancelled_by_business = len([b for b in current_bookings if b.get("status") == "cancelled" and b.get("cancelled_by") == "business"])
    
    return {
        "period": period,
        "summary": {
            "total_bookings": total_bookings,
            "completed": completed,
            "confirmed": confirmed,
            "cancelled": cancelled,
            "cancel_rate": cancel_rate,
            "revenue": round(current_revenue, 2),
            "revenue_change": revenue_change,
            "bookings_change": bookings_change,
            "cancelled_by_user": cancelled_by_user,
            "cancelled_by_business": cancelled_by_business,
        },
        "daily_chart": daily_chart,
        "top_services": top_services,
        "top_clients": top_clients,
        "peak_hours": peak_hours,
        "peak_days": peak_days,
    }



@businesses_router.get("/my/client-history/{user_id}")
async def get_client_history(user_id: str, token_data: TokenData = Depends(require_business)):
    """Get client history within this business"""
    business = await db.businesses.find_one({"user_id": token_data.user_id}, {"_id": 0, "id": 1})
    if not business:
        raise HTTPException(status_code=404, detail="Business not found")
    
    bookings = await db.bookings.find(
        {"business_id": business["id"], "user_id": user_id},
        {"_id": 0, "id": 1, "date": 1, "time": 1, "status": 1, "service_id": 1, "total_amount": 1, "deposit_amount": 1, "cancelled_by": 1}
    ).sort("date", -1).to_list(50)
    
    # Get service names
    service_ids = list(set(b.get("service_id") for b in bookings if b.get("service_id")))
    services_map = {}
    if service_ids:
        for s in await db.services.find({"id": {"$in": service_ids}}, {"_id": 0, "id": 1, "name": 1}).to_list(100):
            services_map[s["id"]] = s["name"]
    
    completed = [b for b in bookings if b.get("status") in ("completed", "confirmed")]
    cancelled = [b for b in bookings if b.get("status") == "cancelled"]
    total_spent = sum(b.get("total_amount") or b.get("deposit_amount") or 0 for b in completed)
    
    history = []
    for b in bookings[:5]:
        history.append({
            "date": b.get("date", ""),
            "time": b.get("time", ""),
            "service_name": services_map.get(b.get("service_id"), ""),
            "status": b.get("status", ""),
            "amount": b.get("total_amount") or b.get("deposit_amount") or 0,
            "cancelled_by": b.get("cancelled_by"),
        })
    
    return {
        "total_visits": len(completed),
        "total_cancelled": len(cancelled),
        "total_spent": round(total_spent, 2),
        "first_visit": bookings[-1]["date"] if bookings else None,
        "last_visit": bookings[0]["date"] if bookings else None,
        "history": history,
    }


@businesses_router.get("/my/reports/export")
async def export_business_reports(
    period: str = "month",
    token_data: TokenData = Depends(require_business)
):
    """Export business reports as Excel file"""
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import io
    
    business = await db.businesses.find_one({"user_id": token_data.user_id}, {"_id": 0})
    if not business:
        raise HTTPException(status_code=404, detail="Business not found")
    
    business_id = business["id"]
    now = datetime.now(timezone.utc)
    period_days = {"week": 7, "month": 30, "quarter": 90, "year": 365}
    days = period_days.get(period, 30)
    start_date = now - timedelta(days=days)
    period_labels = {"week": "7 dias", "month": "30 dias", "quarter": "90 dias", "year": "1 ano"}
    
    bookings = await db.bookings.find(
        {"business_id": business_id, "created_at": {"$gte": start_date.isoformat()}},
        {"_id": 0}
    ).to_list(10000)
    
    transactions = await db.transactions.find(
        {"business_id": business_id, "status": "paid", "created_at": {"$gte": start_date.isoformat()}},
        {"_id": 0}
    ).to_list(10000)
    
    # Build lookup maps
    service_ids = list(set(b.get("service_id") for b in bookings if b.get("service_id")))
    user_ids = list(set(b.get("user_id") for b in bookings if b.get("user_id")))
    worker_ids = list(set(b.get("worker_id") for b in bookings if b.get("worker_id")))
    
    services_map, users_map, workers_map = {}, {}, {}
    if service_ids:
        for s in await db.services.find({"id": {"$in": service_ids}}, {"_id": 0, "id": 1, "name": 1}).to_list(100):
            services_map[s["id"]] = s["name"]
    if user_ids:
        for u in await db.users.find({"id": {"$in": user_ids}}, {"_id": 0, "id": 1, "full_name": 1, "email": 1}).to_list(500):
            users_map[u["id"]] = u
    if worker_ids:
        for w in await db.workers.find({"id": {"$in": worker_ids}}, {"_id": 0, "id": 1, "name": 1}).to_list(100):
            workers_map[w["id"]] = w["name"]
    
    # Create workbook
    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'), right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'), bottom=Side(style='thin', color='D0D0D0')
    )
    
    def style_header(ws, cols):
        for col_idx, title in enumerate(cols, 1):
            cell = ws.cell(row=1, column=col_idx, value=title)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
    
    # === Sheet 1: Resumen ===
    ws1 = wb.active
    ws1.title = "Resumen"
    completed = len([b for b in bookings if b.get("status") == "completed"])
    confirmed = len([b for b in bookings if b.get("status") == "confirmed"])
    cancelled = len([b for b in bookings if b.get("status") == "cancelled"])
    total_rev = sum(t.get("amount_total", 0) for t in transactions)
    
    summary_data = [
        ("Negocio", business["name"]),
        ("Periodo", period_labels.get(period, period)),
        ("Fecha del reporte", now.strftime("%Y-%m-%d %H:%M")),
        ("", ""),
        ("Total de citas", len(bookings)),
        ("Completadas", completed),
        ("Confirmadas", confirmed),
        ("Canceladas", cancelled),
        ("Tasa de cancelacion", f"{round((cancelled / len(bookings) * 100) if bookings else 0, 1)}%"),
        ("Ingresos totales", f"${round(total_rev, 2)}"),
    ]
    for row_idx, (label, value) in enumerate(summary_data, 1):
        ws1.cell(row=row_idx, column=1, value=label).font = Font(bold=True) if label else Font()
        ws1.cell(row=row_idx, column=2, value=value)
    ws1.column_dimensions['A'].width = 25
    ws1.column_dimensions['B'].width = 30
    
    # === Sheet 2: Citas ===
    ws2 = wb.create_sheet("Citas")
    cols = ["Fecha", "Hora", "Cliente", "Email", "Servicio", "Trabajador", "Estado", "Monto", "Cancelado por"]
    style_header(ws2, cols)
    
    status_map = {"confirmed": "Confirmada", "completed": "Completada", "cancelled": "Cancelada", "pending": "Pendiente"}
    for row_idx, b in enumerate(sorted(bookings, key=lambda x: x.get("date", ""), reverse=True), 2):
        u = users_map.get(b.get("user_id"), {})
        ws2.cell(row=row_idx, column=1, value=b.get("date", "")).border = thin_border
        ws2.cell(row=row_idx, column=2, value=b.get("time", "")).border = thin_border
        ws2.cell(row=row_idx, column=3, value=u.get("full_name", b.get("client_name", ""))).border = thin_border
        ws2.cell(row=row_idx, column=4, value=u.get("email", b.get("client_email", ""))).border = thin_border
        ws2.cell(row=row_idx, column=5, value=services_map.get(b.get("service_id"), "")).border = thin_border
        ws2.cell(row=row_idx, column=6, value=workers_map.get(b.get("worker_id"), "")).border = thin_border
        ws2.cell(row=row_idx, column=7, value=status_map.get(b.get("status"), b.get("status", ""))).border = thin_border
        ws2.cell(row=row_idx, column=8, value=b.get("total_amount") or b.get("deposit_amount") or 0).border = thin_border
        ws2.cell(row=row_idx, column=9, value=b.get("cancelled_by", "")).border = thin_border
    
    for col in ['A','B','C','D','E','F','G','H','I']:
        ws2.column_dimensions[col].width = 18
    
    # === Sheet 3: Servicios ===
    ws3 = wb.create_sheet("Servicios")
    style_header(ws3, ["Servicio", "Total citas", "Ingresos"])
    svc_counts, svc_rev = {}, {}
    for b in bookings:
        sid = b.get("service_id")
        if sid:
            svc_counts[sid] = svc_counts.get(sid, 0) + 1
            if b.get("status") in ("completed", "confirmed"):
                svc_rev[sid] = svc_rev.get(sid, 0) + (b.get("total_amount") or b.get("deposit_amount") or 0)
    for row_idx, sid in enumerate(sorted(svc_counts.keys(), key=lambda x: svc_counts[x], reverse=True), 2):
        ws3.cell(row=row_idx, column=1, value=services_map.get(sid, "Servicio")).border = thin_border
        ws3.cell(row=row_idx, column=2, value=svc_counts[sid]).border = thin_border
        ws3.cell(row=row_idx, column=3, value=round(svc_rev.get(sid, 0), 2)).border = thin_border
    for col in ['A','B','C']:
        ws3.column_dimensions[col].width = 25
    
    # === Sheet 4: Clientes ===
    ws4 = wb.create_sheet("Clientes")
    style_header(ws4, ["Cliente", "Email", "Visitas", "Total gastado"])
    client_visits, client_spent = {}, {}
    for b in bookings:
        uid = b.get("user_id")
        if uid and b.get("status") in ("completed", "confirmed"):
            client_visits[uid] = client_visits.get(uid, 0) + 1
            client_spent[uid] = client_spent.get(uid, 0) + (b.get("total_amount") or b.get("deposit_amount") or 0)
    for row_idx, uid in enumerate(sorted(client_visits.keys(), key=lambda x: client_visits[x], reverse=True), 2):
        u = users_map.get(uid, {})
        ws4.cell(row=row_idx, column=1, value=u.get("full_name", "Cliente")).border = thin_border
        ws4.cell(row=row_idx, column=2, value=u.get("email", "")).border = thin_border
        ws4.cell(row=row_idx, column=3, value=client_visits[uid]).border = thin_border
        ws4.cell(row=row_idx, column=4, value=round(client_spent.get(uid, 0), 2)).border = thin_border
    for col in ['A','B','C','D']:
        ws4.column_dimensions[col].width = 25
    
    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"Reporte_{business['name'].replace(' ', '_')}_{period}_{now.strftime('%Y%m%d')}.xlsx"
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )



# ========================== AUTH ROUTES ==========================

@auth_router.post("/register", response_model=dict)
async def register_user(user: UserCreate):
    # Check if email exists
    existing = await db.users.find_one({"email": user.email})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    user_doc = {
        "id": generate_id(),
        "email": user.email,
        "password_hash": hash_password(user.password),
        "full_name": user.full_name,
        "phone": user.phone,
        "country": user.country,
        "city": user.city,
        "phone_verified": False,
        "email_verified": False,
        "email_verification_token": generate_id(),
        "birth_date": user.birth_date,
        "gender": user.gender,
        "photo_url": user.photo_url,
        "role": UserRole.USER,
        "active_appointments_count": 0,
        "cancellation_count": 0,
        "suspended_until": None,
        "favorites": [],
        "preferred_language": user.preferred_language,
        "stripe_customer_id": None,
        "saved_cards": [],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.users.insert_one(user_doc)
    
    # Send verification email (non-blocking)
    try:
        from services.email import send_verification_email
        await send_verification_email(user.email, user.full_name, user_doc["email_verification_token"])
    except Exception as e:
        logger.warning(f"Failed to send verification email: {e}")
    
    return {"message": "Registro exitoso. Revisa tu correo para verificar tu cuenta.", "email": user.email}

@auth_router.post("/login", response_model=dict)
async def login_user(credentials: UserLogin):
    user = await db.users.find_one({"email": credentials.email})
    if not user:
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    if not verify_password(credentials.password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    # Check email verification
    if not user.get("email_verified", True):
        raise HTTPException(status_code=403, detail="email_not_verified")
    
    # Check suspension
    if user.get("suspended_until"):
        suspended_until = datetime.fromisoformat(user["suspended_until"])
        if suspended_until > datetime.now(timezone.utc):
            raise HTTPException(status_code=403, detail=f"Account suspended until {user['suspended_until']}")
    
    token = create_token(user["id"], user["role"], user["email"])
    return {"token": token, "user": UserResponse(**user).model_dump()}

@auth_router.get("/verify-email")
async def verify_email(token: str):
    """Verify user email address using token from email"""
    user = await db.users.find_one({"email_verification_token": token})
    if not user:
        raise HTTPException(status_code=400, detail="Token de verificación inválido o expirado")
    
    if user.get("email_verified"):
        return {"message": "Email ya verificado", "already_verified": True}
    
    await db.users.update_one(
        {"id": user["id"]},
        {"$set": {"email_verified": True}, "$unset": {"email_verification_token": ""}}
    )
    
    return {"message": "Email verificado exitosamente", "already_verified": False}

@auth_router.post("/resend-verification")
async def resend_verification_email(data: dict):
    """Resend verification email"""
    email = data.get("email")
    if not email:
        raise HTTPException(status_code=400, detail="Email requerido")
    
    user = await db.users.find_one({"email": email})
    if not user:
        return {"message": "Si el correo existe, recibirás un email de verificación"}
    
    if user.get("email_verified"):
        return {"message": "Email ya verificado"}
    
    new_token = generate_id()
    await db.users.update_one({"id": user["id"]}, {"$set": {"email_verification_token": new_token}})
    
    try:
        from services.email import send_verification_email
        await send_verification_email(email, user.get("full_name", ""), new_token)
    except Exception as e:
        logger.warning(f"Failed to resend verification email: {e}")
    
    return {"message": "Si el correo existe, recibirás un email de verificación"}

@auth_router.post("/forgot-password")
async def forgot_password(data: dict):
    """Send password reset email"""
    email = data.get("email", "").strip().lower()
    if not email:
        raise HTTPException(status_code=400, detail="Email requerido")
    
    user = await db.users.find_one({"email": email})
    if not user:
        return {"message": "Si el correo existe, recibirás instrucciones para restablecer tu contraseña"}
    
    reset_token = generate_id()
    reset_expires = (datetime.now(timezone.utc) + timedelta(hours=1)).isoformat()
    
    await db.users.update_one(
        {"id": user["id"]},
        {"$set": {"password_reset_token": reset_token, "password_reset_expires": reset_expires}}
    )
    
    try:
        from services.email import send_password_reset_email
        await send_password_reset_email(email, user.get("full_name", ""), reset_token)
    except Exception as e:
        logger.warning(f"Failed to send password reset email: {e}")
    
    return {"message": "Si el correo existe, recibirás instrucciones para restablecer tu contraseña"}

@auth_router.post("/reset-password")
async def reset_password(data: dict):
    """Reset password using token from email"""
    token = data.get("token", "")
    new_password = data.get("password", "")
    
    if not token or not new_password:
        raise HTTPException(status_code=400, detail="Token y contraseña requeridos")
    
    if len(new_password) < 6:
        raise HTTPException(status_code=400, detail="La contraseña debe tener al menos 6 caracteres")
    
    user = await db.users.find_one({"password_reset_token": token})
    if not user:
        raise HTTPException(status_code=400, detail="El enlace es inválido o ha expirado")
    
    expires = user.get("password_reset_expires", "")
    if expires:
        expires_dt = datetime.fromisoformat(expires)
        if datetime.now(timezone.utc) > expires_dt:
            raise HTTPException(status_code=400, detail="El enlace ha expirado. Solicita uno nuevo")
    
    new_hash = hash_password(new_password)
    await db.users.update_one(
        {"id": user["id"]},
        {"$set": {"password_hash": new_hash}, "$unset": {"password_reset_token": "", "password_reset_expires": ""}}
    )
    
    if user.get("role") == "business" and user.get("business_id"):
        await db.businesses.update_one(
            {"id": user["business_id"]},
            {"$set": {"password_hash": new_hash}}
        )
    
    return {"message": "Contraseña actualizada exitosamente"}

@auth_router.post("/phone/send-code")
async def send_phone_code(request: PhoneVerifyRequest):
    """Send phone verification code with rate limiting"""
    from services.sms import send_verification_code, SMSRateLimitError, SMSNotConfiguredError, SMSServiceError
    from core.config import IS_DEVELOPMENT
    
    try:
        code, msg_id = await send_verification_code(request.phone)
        
        response = {"message": "Code sent successfully"}
        if IS_DEVELOPMENT:
            response["dev_code"] = code
        return response
        
    except SMSRateLimitError as e:
        raise HTTPException(status_code=429, detail=str(e))
    except SMSNotConfiguredError as e:
        raise HTTPException(status_code=503, detail=str(e))
    except SMSServiceError as e:
        logger.error(f"SMS error: {e}")
        raise HTTPException(status_code=500, detail="Failed to send verification code")

@auth_router.post("/phone/verify")
async def verify_phone_code(request: PhoneVerifyConfirm, token_data: TokenData = Depends(require_auth)):
    """Verify phone code with proper expiration"""
    from services.sms import verify_code
    
    is_valid = await verify_code(request.phone, request.code)
    if not is_valid:
        raise HTTPException(status_code=400, detail="Invalid or expired code")
    
    # Mark phone as verified
    await db.users.update_one(
        {"id": token_data.user_id},
        {"$set": {"phone_verified": True, "phone": request.phone}}
    )
    
    return {"message": "Phone verified successfully"}

@auth_router.get("/me", response_model=UserResponse)
async def get_current_user_profile(token_data: TokenData = Depends(require_auth)):
    user = await db.users.find_one({"id": token_data.user_id}, {"_id": 0, "password_hash": 0})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    return UserResponse(**user)

# ========================== BUSINESS AUTH ROUTES ==========================

@auth_router.post("/business/register", response_model=dict)
async def register_business(business: BusinessCreate):
    existing = await db.businesses.find_one({"email": business.email})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    # Also create user account for business owner
    user_id = generate_id()
    business_id = generate_id()
    slug = generate_slug(business.name) + "-" + business_id[:8]
    city_slug = generate_slug(business.city)
    
    # Normalize country code
    country_code = business.country.upper()[:2] if business.country else "MX"
    
    # Calculate trial end (3 months free)
    trial_ends = (datetime.now(timezone.utc) + timedelta(days=90)).isoformat()
    
    business_doc = {
        "id": business_id,
        "user_id": user_id,
        "name": business.name,
        "email": business.email,
        "password_hash": hash_password(business.password),
        "phone": business.phone,
        "phone_verified": False,
        "description": business.description,
        "category_id": business.category_id,
        "address": business.address,
        "city": business.city,
        "city_slug": city_slug,
        "state": business.state,
        "country": business.country,
        "country_code": country_code,
        "zip_code": business.zip_code,
        "latitude": business.latitude,
        "longitude": business.longitude,
        "ine_url": business.ine_url,
        "rfc": business.rfc,
        "proof_of_address_url": business.proof_of_address_url,
        "clabe": business.clabe,
        "legal_name": business.legal_name,
        "owner_birth_date": business.owner_birth_date,
        "status": BusinessStatus.PENDING,
        "rating": 0.0,
        "rating_sum": 0.0,
        "review_count": 0,
        "completed_appointments": 0,
        "badges": ["nuevo"],
        "requires_deposit": business.requires_deposit,
        "deposit_amount": max(business.deposit_amount, 50.0),
        "cancellation_days": business.cancellation_days,
        "payout_schedule": business.payout_schedule if business.requires_deposit else None,
        "min_time_between_appointments": business.min_time_between_appointments,
        "service_radius_km": business.service_radius_km,
        "timezone": business.timezone,
        "photos": [],
        "logo_url": None,
        "slug": slug,
        "plan_type": business.plan_type,
        "stripe_account_id": None,
        "stripe_customer_id": None,
        "stripe_subscription_id": None,
        "subscription_status": "none",
        "trial_ends_at": trial_ends,
        "is_featured": False,
        "payout_hold": False,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    user_doc = {
        "id": user_id,
        "email": business.email,
        "password_hash": hash_password(business.password),
        "full_name": business.legal_name,
        "phone": business.phone,
        "phone_verified": False,
        "email_verified": False,
        "email_verification_token": generate_id(),
        "role": UserRole.BUSINESS,
        "business_id": business_id,
        "preferred_language": "es",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.businesses.insert_one(business_doc)
    await db.users.insert_one(user_doc)
    
    # Send verification email (non-blocking)
    try:
        from services.email import send_verification_email
        await send_verification_email(business.email, business.name, user_doc["email_verification_token"])
    except Exception as e:
        logger.warning(f"Failed to send verification email: {e}")
    
    return {"message": "Registro exitoso. Revisa tu correo para verificar tu cuenta.", "email": business.email}

@auth_router.post("/business/login", response_model=dict)
async def login_business(credentials: UserLogin):
    business = await db.businesses.find_one({"email": credentials.email})
    if not business:
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    if not verify_password(credentials.password, business["password_hash"]):
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    user = await db.users.find_one({"business_id": business["id"]})
    if not user:
        raise HTTPException(status_code=500, detail="User account not found")
    
    # Check email verification
    if not user.get("email_verified", True):
        raise HTTPException(status_code=403, detail="email_not_verified")
    
    # Ensure required defaults for BusinessResponse
    business.setdefault("description", "")
    business.setdefault("category_id", "")
    business.setdefault("address", "")
    business.setdefault("city", "")
    business.setdefault("state", "")
    business.setdefault("country", "MX")
    business.setdefault("zip_code", "")
    business.setdefault("subscription_status", "none")
    
    # Remove MongoDB _id
    business.pop("_id", None)
    
    token = create_token(user["id"], UserRole.BUSINESS, business["email"])
    return {"token": token, "business": BusinessResponse(**business).model_dump()}

@auth_router.post("/business/manager-login", response_model=dict)
async def login_manager(credentials: ManagerLogin):
    """Login as a manager/administrator of a business using PIN"""
    business = await db.businesses.find_one({"email": credentials.business_email})
    if not business:
        raise HTTPException(status_code=401, detail="Credenciales inválidas")
    
    worker = await db.workers.find_one({
        "id": credentials.worker_id,
        "business_id": business["id"],
        "is_manager": True
    })
    if not worker:
        raise HTTPException(status_code=401, detail="Credenciales inválidas")
    
    pin_hash = worker.get("manager_pin_hash")
    if not pin_hash:
        raise HTTPException(status_code=401, detail="Este administrador no tiene PIN configurado")
    
    if not verify_password(credentials.pin, pin_hash):
        raise HTTPException(status_code=401, detail="PIN incorrecto")
    
    user = await db.users.find_one({"business_id": business["id"]})
    if not user:
        raise HTTPException(status_code=500, detail="Cuenta de usuario no encontrada")
    
    business.setdefault("description", "")
    business.setdefault("category_id", "")
    business.setdefault("address", "")
    business.setdefault("city", "")
    business.setdefault("state", "")
    business.setdefault("country", "MX")
    business.setdefault("zip_code", "")
    business.setdefault("subscription_status", "none")
    business.pop("_id", None)
    
    token = create_token(user["id"], UserRole.BUSINESS, business["email"], worker_id=worker["id"], is_manager=True)
    
    permissions = worker.get("manager_permissions", {})
    return {
        "token": token,
        "business": BusinessResponse(**business).model_dump(),
        "manager": {
            "worker_id": worker["id"],
            "worker_name": worker["name"],
            "permissions": permissions,
            "is_manager": True
        }
    }

@auth_router.get("/business/managers")
async def get_business_managers(email: str):
    """Get list of manager workers for a business (used in login flow)"""
    business = await db.businesses.find_one({"email": email})
    if not business:
        return []
    workers = await db.workers.find(
        {"business_id": business["id"], "is_manager": True},
        {"_id": 0, "id": 1, "name": 1, "has_manager_pin": 1}
    ).to_list(100)
    # Compute has_manager_pin from manager_pin_hash
    result = []
    for w in workers:
        w_full = await db.workers.find_one({"id": w["id"]}, {"_id": 0, "manager_pin_hash": 1})
        result.append({"id": w["id"], "name": w["name"], "has_pin": bool(w_full.get("manager_pin_hash"))})
    return result

# ========================== ADMIN AUTH ROUTES ==========================

@auth_router.post("/admin/setup-2fa", response_model=dict)
async def setup_admin_2fa(setup: Admin2FASetup, token_data: TokenData = Depends(require_admin)):
    user = await db.users.find_one({"id": token_data.user_id})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    if not verify_password(setup.password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="Invalid password")
    
    # Generate TOTP secret
    secret = pyotp.random_base32()
    totp = pyotp.TOTP(secret)
    provisioning_uri = totp.provisioning_uri(name=user["email"], issuer_name="Bookvia Admin")
    
    # Generate QR code
    qr = qrcode.QRCode(version=1, box_size=10, border=5)
    qr.add_data(provisioning_uri)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    
    buffer = io.BytesIO()
    img.save(buffer, format='PNG')
    qr_base64 = base64.b64encode(buffer.getvalue()).decode('utf-8')
    
    # Generate backup codes
    backup_codes = [str(random.randint(10000000, 99999999)) for _ in range(8)]
    hashed_backups = [hash_password(code) for code in backup_codes]
    
    # Store secret (not yet verified)
    await db.users.update_one(
        {"id": token_data.user_id},
        {"$set": {
            "totp_secret_pending": secret,
            "backup_codes_pending": hashed_backups
        }}
    )
    
    return {
        "qr_code": f"data:image/png;base64,{qr_base64}",
        "secret": secret,
        "backup_codes": backup_codes,
        "message": "Scan QR code with authenticator app, then verify with a code"
    }

@auth_router.post("/admin/verify-2fa")
async def verify_admin_2fa(verify: Admin2FAVerify, token_data: TokenData = Depends(require_admin)):
    user = await db.users.find_one({"id": token_data.user_id})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    pending_secret = user.get("totp_secret_pending")
    if not pending_secret:
        raise HTTPException(status_code=400, detail="No 2FA setup in progress")
    
    totp = pyotp.TOTP(pending_secret)
    if not totp.verify(verify.code):
        raise HTTPException(status_code=400, detail="Invalid code")
    
    # Activate 2FA
    await db.users.update_one(
        {"id": token_data.user_id},
        {
            "$set": {
                "totp_secret": pending_secret,
                "backup_codes": user.get("backup_codes_pending", []),
                "totp_enabled": True
            },
            "$unset": {
                "totp_secret_pending": "",
                "backup_codes_pending": ""
            }
        }
    )
    
    return {"message": "2FA enabled successfully"}

@auth_router.post("/admin/login", response_model=dict)
async def admin_login(credentials: AdminLogin, request: Request):
    user = await db.users.find_one({"email": credentials.email, "role": UserRole.ADMIN})
    if not user:
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    if not verify_password(credentials.password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    # Check if 2FA is enabled - REQUIRED for admin
    if not user.get("totp_enabled"):
        # Return special response indicating 2FA setup required
        temp_token = create_token(user["id"], UserRole.ADMIN, user["email"])
        return {
            "requires_2fa_setup": True,
            "temp_token": temp_token,
            "message": "2FA setup required before accessing admin panel"
        }
    
    # Verify 2FA
    totp = pyotp.TOTP(user["totp_secret"])
    if not totp.verify(credentials.totp_code):
        # Check backup codes
        valid_backup = False
        for i, hashed_code in enumerate(user.get("backup_codes", [])):
            if verify_password(credentials.totp_code, hashed_code):
                valid_backup = True
                # Remove used backup code
                backup_codes = user["backup_codes"]
                backup_codes.pop(i)
                await db.users.update_one(
                    {"id": user["id"]},
                    {"$set": {"backup_codes": backup_codes}}
                )
                break
        
        if not valid_backup:
            raise HTTPException(status_code=401, detail="Invalid 2FA code")
    
    # Create audit log for admin login
    await create_audit_log(
        admin_id=user["id"],
        admin_email=user["email"],
        action=AuditAction.ADMIN_LOGIN,
        target_type="admin",
        target_id=user["id"],
        details={"login_successful": True},
        request=request
    )
    
    token = create_token(user["id"], UserRole.ADMIN, user["email"])
    return {"token": token, "user": UserResponse(**user).model_dump(), "totp_enabled": True}

# ========================== USER ROUTES ==========================

@users_router.put("/me", response_model=UserResponse)
async def update_user_profile(update: UserUpdate, token_data: TokenData = Depends(require_auth)):
    update_data = {k: v for k, v in update.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="No data to update")
    
    await db.users.update_one({"id": token_data.user_id}, {"$set": update_data})
    user = await db.users.find_one({"id": token_data.user_id}, {"_id": 0, "password_hash": 0})
    return UserResponse(**user)

@users_router.post("/favorites/{business_id}")
async def add_favorite(business_id: str, token_data: TokenData = Depends(require_auth)):
    business = await db.businesses.find_one({"id": business_id})
    if not business:
        raise HTTPException(status_code=404, detail="Business not found")
    
    await db.users.update_one(
        {"id": token_data.user_id},
        {"$addToSet": {"favorites": business_id}}
    )
    return {"message": "Added to favorites"}

@users_router.delete("/favorites/{business_id}")
async def remove_favorite(business_id: str, token_data: TokenData = Depends(require_auth)):
    await db.users.update_one(
        {"id": token_data.user_id},
        {"$pull": {"favorites": business_id}}
    )
    return {"message": "Removed from favorites"}

@users_router.get("/favorites", response_model=List[BusinessResponse])
async def get_favorites(token_data: TokenData = Depends(require_auth)):
    user = await db.users.find_one({"id": token_data.user_id})
    if not user or not user.get("favorites"):
        return []
    
    businesses = await db.businesses.find(
        {"id": {"$in": user["favorites"]}, **VISIBLE_BUSINESS_FILTER},
        {"_id": 0, "password_hash": 0}
    ).to_list(100)
    
    return [BusinessResponse(**b) for b in businesses]

# ========================== CATEGORY ROUTES ==========================

@categories_router.get("", response_model=List[CategoryResponse])
async def get_categories(city: Optional[str] = None, country_code: Optional[str] = None):
    categories = await db.categories.find({}, {"_id": 0}).to_list(100)
    
    # Build business filter - optionally scoped to city+country
    biz_filter = {**VISIBLE_BUSINESS_FILTER}
    if city:
        biz_filter["city"] = city
    if country_code:
        biz_filter["country_code"] = country_code.upper()
    
    # Add business count for each category
    for cat in categories:
        count = await db.businesses.count_documents({"category_id": cat["id"], **biz_filter})
        cat["business_count"] = count
    
    # If filtering by city, only return categories that actually have businesses
    if city:
        categories = [c for c in categories if c["business_count"] > 0]
    
    return [CategoryResponse(**c) for c in categories]

@categories_router.get("/{slug}", response_model=CategoryResponse)
async def get_category_by_slug(slug: str):
    category = await db.categories.find_one({"slug": slug}, {"_id": 0})
    if not category:
        raise HTTPException(status_code=404, detail="Category not found")
    
    count = await db.businesses.count_documents({"category_id": category["id"], **VISIBLE_BUSINESS_FILTER})
    category["business_count"] = count
    
    return CategoryResponse(**category)

@categories_router.post("/", response_model=CategoryResponse)
async def create_category(category: CategoryCreate, token_data: TokenData = Depends(require_admin)):
    category_doc = {
        "id": generate_id(),
        **category.model_dump()
    }
    await db.categories.insert_one(category_doc)
    return CategoryResponse(**category_doc)

# ========================== BUSINESS ROUTES ==========================

@businesses_router.get("", response_model=List[BusinessResponse])
async def search_businesses(
    request: Request,
    query: Optional[str] = None,
    category_id: Optional[str] = None,
    city: Optional[str] = None,
    country_code: Optional[str] = None,
    min_rating: Optional[float] = None,
    is_home_service: Optional[bool] = None,
    include_pending: bool = False,
    user_lat: Optional[float] = None,
    user_lng: Optional[float] = None,
    sort: Optional[str] = None,
    page: int = 1,
    limit: int = 20,
    current_user: Optional[TokenData] = Depends(get_current_user)
):
    # By default only show approved businesses with active subscription
    # If include_pending=True, also show PENDING (for profile viewing, but no bookings)
    if include_pending:
        filters = {"status": {"$in": [BusinessStatus.APPROVED, BusinessStatus.PENDING]}}
    else:
        filters = {
            "status": BusinessStatus.APPROVED,
            "$or": [
                {"subscription_status": {"$in": ["active", "trialing"]}},
                {"subscription_status": {"$exists": False}},
                {"subscription_status": None},
                {"subscription_status": "none"},
            ]
        }
    
    if country_code:
        filters["country_code"] = country_code.upper()
    if query and not category_id:
        # Also search by matching category names
        matching_cats = await db.categories.find(
            {"$or": [
                {"name_es": {"$regex": query, "$options": "i"}},
                {"name_en": {"$regex": query, "$options": "i"}}
            ]},
            {"_id": 0, "id": 1}
        ).to_list(50)
        matching_cat_ids = [c["id"] for c in matching_cats]
        
        or_conditions = [
            {"name": {"$regex": query, "$options": "i"}},
            {"description": {"$regex": query, "$options": "i"}}
        ]
        if matching_cat_ids:
            or_conditions.append({"category_id": {"$in": matching_cat_ids}})
        filters["$or"] = or_conditions
    elif query:
        filters["$or"] = [
            {"name": {"$regex": query, "$options": "i"}},
            {"description": {"$regex": query, "$options": "i"}}
        ]
    if category_id:
        filters["category_id"] = category_id
    if city:
        filters["city"] = {"$regex": city, "$options": "i"}
    if min_rating:
        filters["rating"] = {"$gte": min_rating}
    if is_home_service is not None:
        filters["service_radius_km"] = {"$exists": True, "$ne": None} if is_home_service else {"$in": [None, 0]}
    
    skip = (page - 1) * limit
    businesses = await db.businesses.find(
        filters,
        {"_id": 0, "password_hash": 0, "clabe": 0, "rfc": 0, "ine_url": 0, "proof_of_address_url": 0}
    ).sort("rating", -1).skip(skip).limit(limit).to_list(limit)
    
    # Add category names and booking availability
    for b in businesses:
        if b.get("category_id"):
            cat = await db.categories.find_one({"id": b["category_id"]})
            if cat:
                b["category_name"] = cat.get("name_es", "")
        # Mark if business can accept bookings (must be approved + subscription active/trialing or legacy)
        sub_status = b.get("subscription_status", "none")
        b["can_accept_bookings"] = (
            b.get("status") == BusinessStatus.APPROVED 
            and (sub_status in ("active", "trialing", "none", None) or sub_status is None)
        )
    
    # Filter out businesses where the current user is blacklisted
    if current_user:
        blacklisted_biz_ids = set()
        for b in businesses:
            if await is_user_blacklisted(b["id"], user_id=current_user.user_id):
                blacklisted_biz_ids.add(b["id"])
        businesses = [b for b in businesses if b["id"] not in blacklisted_biz_ids]
    
    # Calculate distance and sort by proximity if user location provided
    if user_lat is not None and user_lng is not None:
        import math
        def haversine(lat1, lng1, lat2, lng2):
            R = 6371  # km
            dlat = math.radians(lat2 - lat1)
            dlng = math.radians(lng2 - lng1)
            a = math.sin(dlat/2)**2 + math.cos(math.radians(lat1)) * math.cos(math.radians(lat2)) * math.sin(dlng/2)**2
            return R * 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
        
        for b in businesses:
            if b.get("latitude") and b.get("longitude"):
                b["distance_km"] = round(haversine(user_lat, user_lng, b["latitude"], b["longitude"]), 1)
            else:
                b["distance_km"] = None
        
        if sort == "nearest":
            businesses.sort(key=lambda x: x.get("distance_km") if x.get("distance_km") is not None else 99999)
    
    # Compute next available text based on worker schedules
    biz_ids = [b["id"] for b in businesses]
    if biz_ids:
        all_workers = await db.workers.find(
            {"business_id": {"$in": biz_ids}, "active": True},
            {"_id": 0, "business_id": 1, "schedule": 1}
        ).to_list(500)
        biz_workers_map = {}
        for w in all_workers:
            biz_workers_map.setdefault(w["business_id"], []).append(w)
        
        day_names_es = ["Lun", "Mar", "Mie", "Jue", "Vie", "Sab", "Dom"]
        try:
            now = datetime.now(pytz.timezone("America/Mexico_City"))
        except Exception:
            now = datetime.now(timezone.utc)
        current_weekday = now.weekday()
        current_time = now.strftime("%H:%M")
        
        for b in businesses:
            workers_list = biz_workers_map.get(b["id"], [])
            if not workers_list:
                continue
            found = False
            for day_offset in range(7):
                check_day = (current_weekday + day_offset) % 7
                day_str = str(check_day)
                for worker in workers_list:
                    schedule = worker.get("schedule", {})
                    day_schedule = schedule.get(day_str, {})
                    if day_schedule.get("is_available") and day_schedule.get("blocks"):
                        for block in day_schedule["blocks"]:
                            if day_offset == 0:
                                if block["end_time"] > current_time:
                                    b["next_available_text"] = "Hoy disponible"
                                    found = True
                                    break
                            else:
                                if day_offset == 1:
                                    b["next_available_text"] = f"Manana {block['start_time']}"
                                else:
                                    target_day = (current_weekday + day_offset) % 7
                                    b["next_available_text"] = f"{day_names_es[target_day]} {block['start_time']}"
                                found = True
                                break
                        if found:
                            break
                    if found:
                        break
                if found:
                    break
    
    return [BusinessResponse(**b) for b in businesses]

@businesses_router.get("/featured", response_model=List[BusinessResponse])
async def get_featured_businesses(limit: int = 8, country_code: Optional[str] = None, current_user: Optional[TokenData] = Depends(get_current_user)):
    base_filter = {"status": BusinessStatus.APPROVED, "is_featured": True, "$or": [{"subscription_status": {"$in": ["active", "trialing"]}}, {"subscription_status": {"$exists": False}}, {"subscription_status": None}, {"subscription_status": "none"}]}
    if country_code:
        base_filter["country_code"] = country_code.upper()
    businesses = await db.businesses.find(
        base_filter,
        {"_id": 0, "password_hash": 0, "clabe": 0, "rfc": 0}
    ).sort("rating", -1).limit(limit).to_list(limit)
    
    # If not enough featured, add top rated
    if len(businesses) < limit:
        existing_ids = [b["id"] for b in businesses]
        more_filter = {"status": BusinessStatus.APPROVED, "id": {"$nin": existing_ids}, "$or": [{"subscription_status": {"$in": ["active", "trialing"]}}, {"subscription_status": {"$exists": False}}, {"subscription_status": None}, {"subscription_status": "none"}]}
        if country_code:
            more_filter["country_code"] = country_code.upper()
        more = await db.businesses.find(
            more_filter,
            {"_id": 0, "password_hash": 0, "clabe": 0, "rfc": 0}
        ).sort("rating", -1).limit(limit - len(businesses)).to_list(limit - len(businesses))
        businesses.extend(more)
    
    # Filter out businesses where the current user is blacklisted
    if current_user:
        blacklisted_biz_ids = set()
        for b in businesses:
            if await is_user_blacklisted(b["id"], user_id=current_user.user_id):
                blacklisted_biz_ids.add(b["id"])
        businesses = [b for b in businesses if b["id"] not in blacklisted_biz_ids]
    
    # Compute next available text for featured businesses
    biz_ids = [b["id"] for b in businesses]
    if biz_ids:
        all_workers = await db.workers.find(
            {"business_id": {"$in": biz_ids}, "active": True},
            {"_id": 0, "business_id": 1, "schedule": 1}
        ).to_list(500)
        biz_workers_map = {}
        for w in all_workers:
            biz_workers_map.setdefault(w["business_id"], []).append(w)
        day_names_es = ["Lun", "Mar", "Mie", "Jue", "Vie", "Sab", "Dom"]
        try:
            now_ft = datetime.now(pytz.timezone("America/Mexico_City"))
        except Exception:
            now_ft = datetime.now(timezone.utc)
        current_wd = now_ft.weekday()
        current_tm = now_ft.strftime("%H:%M")
        for b in businesses:
            wl = biz_workers_map.get(b["id"], [])
            if not wl:
                continue
            found = False
            for doff in range(7):
                cd = (current_wd + doff) % 7
                for wk in wl:
                    ds = wk.get("schedule", {}).get(str(cd), {})
                    if ds.get("is_available") and ds.get("blocks"):
                        for blk in ds["blocks"]:
                            if doff == 0 and blk["end_time"] > current_tm:
                                b["next_available_text"] = "Hoy disponible"
                                found = True
                                break
                            elif doff > 0:
                                b["next_available_text"] = f"{'Manana' if doff == 1 else day_names_es[(current_wd + doff) % 7]} {blk['start_time']}"
                                found = True
                                break
                        if found:
                            break
                    if found:
                        break
                if found:
                    break
    
    return [BusinessResponse(**b) for b in businesses]

@businesses_router.get("/slug/{slug}", response_model=BusinessResponse)
async def get_business_by_slug(slug: str, current_user: Optional[TokenData] = Depends(get_current_user)):
    """Get business by slug or ID (fallback)."""
    business = await db.businesses.find_one(
        {"slug": slug},
        {"_id": 0, "password_hash": 0, "clabe": 0, "rfc": 0, "ine_url": 0, "proof_of_address_url": 0}
    )
    if not business:
        # Fallback: try finding by ID
        business = await db.businesses.find_one(
            {"id": slug},
            {"_id": 0, "password_hash": 0, "clabe": 0, "rfc": 0, "ine_url": 0, "proof_of_address_url": 0}
        )
    if not business:
        raise HTTPException(status_code=404, detail="Business not found")
    
    # Check blacklist
    if current_user and await is_user_blacklisted(business["id"], user_id=current_user.user_id):
        raise HTTPException(status_code=404, detail="Business not found")
    
    if business.get("category_id"):
        cat = await db.categories.find_one({"id": business["category_id"]})
        if cat:
            business["category_name"] = cat.get("name_es", "")
    
    return BusinessResponse(**business)

@businesses_router.get("/{business_id}", response_model=BusinessResponse)
async def get_business(business_id: str, current_user: Optional[TokenData] = Depends(get_current_user)):
    business = await db.businesses.find_one(
        {"id": business_id},
        {"_id": 0, "password_hash": 0, "clabe": 0, "rfc": 0, "ine_url": 0, "proof_of_address_url": 0}
    )
    if not business:
        raise HTTPException(status_code=404, detail="Business not found")
    # Check blacklist
    if current_user and await is_user_blacklisted(business["id"], user_id=current_user.user_id):
        raise HTTPException(status_code=404, detail="Business not found")
    return BusinessResponse(**business)

@businesses_router.put("/me", response_model=BusinessResponse)
async def update_my_business(update: BusinessUpdate, token_data: TokenData = Depends(require_business)):
    user = await db.users.find_one({"id": token_data.user_id})
    if not user or not user.get("business_id"):
        raise HTTPException(status_code=404, detail="Business not found")
    
    update_data = {k: v for k, v in update.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="No data to update")
    
    if "deposit_amount" in update_data:
        update_data["deposit_amount"] = max(update_data["deposit_amount"], 50.0)
    
    await db.businesses.update_one({"id": user["business_id"]}, {"$set": update_data})
    business = await db.businesses.find_one({"id": user["business_id"]}, {"_id": 0, "password_hash": 0})
    return BusinessResponse(**business)

@businesses_router.get("/me/dashboard")
async def get_business_dashboard(token_data: TokenData = Depends(require_business)):
    user = await db.users.find_one({"id": token_data.user_id})
    if not user or not user.get("business_id"):
        raise HTTPException(status_code=404, detail="Business not found")
    
    business = await db.businesses.find_one({"id": user["business_id"]}, {"_id": 0, "password_hash": 0})
    if not business:
        raise HTTPException(status_code=404, detail="Business document not found")
    
    # Ensure slug exists
    if not business.get("slug"):
        slug = f"{business['name'].lower().replace(' ', '-')}-{business['id'][:8]}"
        business["slug"] = slug
        await db.businesses.update_one({"id": business["id"]}, {"$set": {"slug": slug}})
    
    # Ensure required fields have defaults to prevent Pydantic validation errors
    business.setdefault("description", "")
    business.setdefault("category_id", "")
    business.setdefault("address", "")
    business.setdefault("city", "")
    business.setdefault("state", "")
    business.setdefault("country", "MX")
    business.setdefault("zip_code", "")
    business.setdefault("subscription_status", "none")
    
    # Get stats
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    bid = business["id"]
    
    today_appointments = await db.bookings.count_documents({
        "business_id": bid,
        "date": today,
        "status": AppointmentStatus.CONFIRMED
    })
    
    pending_appointments = await db.bookings.count_documents({
        "business_id": bid,
        "status": AppointmentStatus.CONFIRMED
    })
    
    # This month revenue
    first_of_month = datetime.now(timezone.utc).replace(day=1).strftime("%Y-%m-%d")
    month_bookings = await db.bookings.find({
        "business_id": bid,
        "date": {"$gte": first_of_month},
        "status": AppointmentStatus.COMPLETED
    }).to_list(1000)
    
    month_revenue = sum(b.get("total_amount", 0) for b in month_bookings)
    
    total_appointments = await db.bookings.count_documents({
        "business_id": bid,
        "status": {"$in": [AppointmentStatus.CONFIRMED, AppointmentStatus.COMPLETED, AppointmentStatus.NO_SHOW]}
    })
    
    try:
        biz_response = BusinessResponse(**business).model_dump()
    except Exception as e:
        logger.error(f"BusinessResponse validation error for {bid}: {e}")
        logger.error(f"Business fields: {list(business.keys())}")
        raise HTTPException(status_code=500, detail=f"Error loading business data: {str(e)}")
    
    return {
        "business": biz_response,
        "stats": {
            "today_appointments": today_appointments,
            "pending_appointments": pending_appointments,
            "month_revenue": month_revenue,
            "total_appointments": total_appointments,
            "total_reviews": business.get("review_count", 0),
            "rating": business.get("rating", 0)
        },
        "subscription": {
            "status": business.get("subscription_status", "none"),
            "subscription_id": business.get("stripe_subscription_id"),
            "customer_id": business.get("stripe_customer_id"),
            "started_at": business.get("subscription_started_at"),
            "cancel_requested": business.get("subscription_cancel_requested", False)
        }
    }

# ========================== BLACKLIST ROUTES ==========================

@businesses_router.get("/me/blacklist", response_model=List[BlacklistResponse])
async def get_blacklist(token_data: TokenData = Depends(require_business)):
    user = await db.users.find_one({"id": token_data.user_id})
    if not user or not user.get("business_id"):
        raise HTTPException(status_code=404, detail="Business not found")
    entries = await db.blacklist.find(
        {"business_id": user["business_id"]}, {"_id": 0}
    ).sort("created_at", -1).to_list(500)
    return [BlacklistResponse(**e) for e in entries]

@businesses_router.post("/me/blacklist", response_model=BlacklistResponse)
async def add_to_blacklist(entry: BlacklistEntry, token_data: TokenData = Depends(require_business)):
    user = await db.users.find_one({"id": token_data.user_id})
    if not user or not user.get("business_id"):
        raise HTTPException(status_code=404, detail="Business not found")
    if not entry.email and not entry.phone and not entry.user_id:
        raise HTTPException(status_code=400, detail="At least one identifier (email, phone, or user_id) is required")
    
    # Normalize email
    email_normalized = entry.email.strip().lower() if entry.email else None
    
    # Check if already blacklisted
    or_conditions = []
    if email_normalized:
        or_conditions.append({"email": email_normalized})
    if entry.phone:
        or_conditions.append({"phone": entry.phone})
    if entry.user_id:
        or_conditions.append({"user_id": entry.user_id})
    
    existing = await db.blacklist.find_one({
        "business_id": user["business_id"],
        "$or": or_conditions
    })
    if existing:
        raise HTTPException(status_code=409, detail="This user is already blacklisted")
    
    doc = {
        "id": generate_id(),
        "business_id": user["business_id"],
        "email": email_normalized,
        "phone": entry.phone,
        "user_id": entry.user_id,
        "reason": entry.reason,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.blacklist.insert_one(doc)
    return BlacklistResponse(**{k: v for k, v in doc.items() if k != "_id"})

@businesses_router.delete("/me/blacklist/{entry_id}")
async def remove_from_blacklist(entry_id: str, token_data: TokenData = Depends(require_business)):
    user = await db.users.find_one({"id": token_data.user_id})
    if not user or not user.get("business_id"):
        raise HTTPException(status_code=404, detail="Business not found")
    result = await db.blacklist.delete_one({"id": entry_id, "business_id": user["business_id"]})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Entry not found")
    return {"message": "Removed from blacklist"}

# ========================== STATS DETAIL ENDPOINT ==========================

@bookings_router.get("/business/stats-detail")
async def get_business_stats_detail(
    stat_type: str,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    token_data: TokenData = Depends(require_business)
):
    """Get detailed bookings for a specific stat card."""
    user = await db.users.find_one({"id": token_data.user_id})
    if not user or not user.get("business_id"):
        raise HTTPException(status_code=404, detail="Business not found")
    
    business_id = user["business_id"]
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    first_of_month = datetime.now(timezone.utc).replace(day=1).strftime("%Y-%m-%d")
    
    filters = {"business_id": business_id}
    
    if stat_type == "today":
        filters["date"] = today
        filters["status"] = AppointmentStatus.CONFIRMED
    elif stat_type == "pending":
        filters["status"] = AppointmentStatus.CONFIRMED
    elif stat_type == "revenue":
        filters["status"] = AppointmentStatus.COMPLETED
        if date_from and date_to:
            filters["date"] = {"$gte": date_from, "$lte": date_to}
        else:
            filters["date"] = {"$gte": first_of_month}
    elif stat_type == "total":
        filters["status"] = {"$in": [
            AppointmentStatus.CONFIRMED, 
            AppointmentStatus.COMPLETED, AppointmentStatus.NO_SHOW
        ]}
        if date_from and date_to:
            filters["date"] = {"$gte": date_from, "$lte": date_to}
    else:
        raise HTTPException(status_code=400, detail="Invalid stat_type")
    
    bookings = await db.bookings.find(filters, {"_id": 0}).sort("date", -1).limit(200).to_list(200)
    
    # Populate names
    for b in bookings:
        user_doc = await db.users.find_one({"id": b.get("user_id")}, {"_id": 0, "full_name": 1})
        service_doc = await db.services.find_one({"id": b.get("service_id")}, {"_id": 0, "name": 1})
        worker_doc = await db.workers.find_one({"id": b.get("worker_id")}, {"_id": 0, "name": 1})
        b["user_name"] = user_doc.get("full_name") if user_doc else None
        b["service_name"] = service_doc.get("name") if service_doc else None
        b["worker_name"] = worker_doc.get("name") if worker_doc else None
    
    total_revenue = sum(b.get("total_amount", 0) for b in bookings) if stat_type == "revenue" else None
    
    return {
        "bookings": bookings,
        "count": len(bookings),
        "total_revenue": total_revenue
    }

# ========================== WORKER ROUTES ==========================

def validate_schedule_blocks(schedule: Dict[str, DaySchedule]) -> None:
    """Validate that schedule blocks don't overlap within each day"""
    for day, day_schedule in schedule.items():
        if not day_schedule.is_available or not day_schedule.blocks:
            continue
        
        # Sort blocks by start time
        blocks = sorted(day_schedule.blocks, key=lambda b: b.start_time)
        
        for i in range(len(blocks) - 1):
            current_end = datetime.strptime(blocks[i].end_time, "%H:%M")
            next_start = datetime.strptime(blocks[i + 1].start_time, "%H:%M")
            
            if current_end > next_start:
                raise HTTPException(
                    status_code=400,
                    detail=f"Schedule blocks overlap on day {day}: {blocks[i].end_time} > {blocks[i+1].start_time}"
                )
            
            # Validate start < end for each block
            current_start = datetime.strptime(blocks[i].start_time, "%H:%M")
            if current_start >= current_end:
                raise HTTPException(
                    status_code=400, 
                    detail=f"Invalid block on day {day}: start_time must be before end_time"
                )
        
        # Validate last block
        if blocks:
            last_start = datetime.strptime(blocks[-1].start_time, "%H:%M")
            last_end = datetime.strptime(blocks[-1].end_time, "%H:%M")
            if last_start >= last_end:
                raise HTTPException(
                    status_code=400,
                    detail=f"Invalid block on day {day}: start_time must be before end_time"
                )

@businesses_router.post("/my/workers", response_model=WorkerResponse)
async def create_worker(worker: WorkerCreate, token_data: TokenData = Depends(require_business)):
    """Create a new worker for the business"""
    user = await db.users.find_one({"id": token_data.user_id})
    if not user or not user.get("business_id"):
        raise HTTPException(status_code=404, detail="Business not found")
    
    # Default schedule: Mon-Fri 9:00-18:00
    default_schedule = {}
    for day in range(5):  # Mon-Fri
        default_schedule[str(day)] = {
            "is_available": True,
            "blocks": [{"start_time": "09:00", "end_time": "18:00"}]
        }
    for day in range(5, 7):  # Sat-Sun
        default_schedule[str(day)] = {"is_available": False, "blocks": []}
    
    worker_doc = {
        "id": generate_id(),
        "business_id": user["business_id"],
        "name": worker.name,
        "email": worker.email,
        "phone": worker.phone,
        "photo_url": worker.photo_url,
        "bio": worker.bio,
        "service_ids": worker.service_ids,
        "schedule": default_schedule,
        "exceptions": [],  # New: list of WorkerException
        "active": True,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "deactivated_at": None
    }
    
    await db.workers.insert_one(worker_doc)
    return WorkerResponse(**worker_doc)

@businesses_router.get("/my/workers", response_model=List[WorkerResponse])
async def get_business_workers(
    include_inactive: bool = False,
    token_data: TokenData = Depends(require_auth)
):
    """Get all workers for the authenticated business"""
    user = await db.users.find_one({"id": token_data.user_id})
    if not user or not user.get("business_id"):
        raise HTTPException(status_code=400, detail="Business ID required")
    business_id = user["business_id"]
    
    filters = {"business_id": business_id}
    if not include_inactive:
        filters["active"] = True
    
    workers = await db.workers.find(filters, {"_id": 0}).to_list(100)
    for w in workers:
        w["has_manager_pin"] = bool(w.get("manager_pin_hash"))
    return [WorkerResponse(**w) for w in workers]

@businesses_router.get("/{business_id}/workers", response_model=List[WorkerResponse])
async def get_workers_by_business(business_id: str, service_id: Optional[str] = None, include_inactive: bool = False):
    """Get workers for a specific business (public endpoint), optionally filtered by service."""
    filters = {"business_id": business_id}
    if not include_inactive:
        filters["active"] = True
    workers = await db.workers.find(filters, {"_id": 0}).to_list(100)
    # Filter by service_id: include workers who have the service in their service_ids or have no service_ids set (legacy)
    if service_id:
        workers = [w for w in workers if not w.get("service_ids") or service_id in w.get("service_ids", [])]
    for w in workers:
        w["has_manager_pin"] = bool(w.get("manager_pin_hash"))
    return [WorkerResponse(**w) for w in workers]

@businesses_router.get("/my/workers/{worker_id}", response_model=WorkerResponse)
async def get_worker(worker_id: str, token_data: TokenData = Depends(require_business)):
    """Get a specific worker"""
    user = await db.users.find_one({"id": token_data.user_id})
    worker = await db.workers.find_one({"id": worker_id, "business_id": user.get("business_id")}, {"_id": 0})
    if not worker:
        raise HTTPException(status_code=404, detail="Worker not found")
    worker["has_manager_pin"] = bool(worker.get("manager_pin_hash"))
    return WorkerResponse(**worker)

@businesses_router.put("/my/workers/{worker_id}/services")
async def update_worker_services(worker_id: str, request: Request, token_data: TokenData = Depends(require_business)):
    """Update which services a worker can perform."""
    user = await db.users.find_one({"id": token_data.user_id})
    if not user or not user.get("business_id"):
        raise HTTPException(status_code=404, detail="Business not found")
    
    worker = await db.workers.find_one({"id": worker_id, "business_id": user["business_id"]})
    if not worker:
        raise HTTPException(status_code=404, detail="Worker not found")
    
    body = await request.json()
    service_ids = body.get("service_ids", [])
    
    await db.workers.update_one(
        {"id": worker_id},
        {"$set": {"service_ids": service_ids}}
    )
    
    return {"message": "Worker services updated", "service_ids": service_ids}


@businesses_router.put("/my/workers/{worker_id}", response_model=WorkerResponse)
async def update_worker(
    worker_id: str,
    update: WorkerUpdate,
    token_data: TokenData = Depends(require_business)
):
    """Update worker basic info"""
    user = await db.users.find_one({"id": token_data.user_id})
    worker = await db.workers.find_one({"id": worker_id, "business_id": user.get("business_id")})
    if not worker:
        raise HTTPException(status_code=404, detail="Worker not found")
    
    update_data = {k: v for k, v in update.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="No data to update")
    
    await db.workers.update_one({"id": worker_id}, {"$set": update_data})
    updated = await db.workers.find_one({"id": worker_id}, {"_id": 0})
    return WorkerResponse(**updated)

@businesses_router.delete("/my/workers/{worker_id}")
async def delete_worker(worker_id: str, token_data: TokenData = Depends(require_business)):
    """Soft delete a worker (marks as inactive, preserves history)"""
    user = await db.users.find_one({"id": token_data.user_id})
    worker = await db.workers.find_one({"id": worker_id, "business_id": user.get("business_id")})
    if not worker:
        raise HTTPException(status_code=404, detail="Worker not found")
    
    # Soft delete: mark as inactive but preserve all data
    await db.workers.update_one(
        {"id": worker_id},
        {"$set": {
            "active": False,
            "deactivated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    return {"message": "Worker deactivated", "worker_id": worker_id}


@businesses_router.post("/my/workers/{worker_id}/photo")
async def upload_worker_photo(worker_id: str, file: UploadFile = File(...), token_data: TokenData = Depends(require_business)):
    """Upload or replace a worker's profile photo."""
    user = await db.users.find_one({"id": token_data.user_id})
    if not user or not user.get("business_id"):
        raise HTTPException(status_code=404, detail="Business not found")

    worker = await db.workers.find_one({"id": worker_id, "business_id": user["business_id"]})
    if not worker:
        raise HTTPException(status_code=404, detail="Worker not found")

    data = await file.read()
    ok, err = validate_image(file.filename, file.content_type, len(data))
    if not ok:
        raise HTTPException(status_code=400, detail=err)

    if cloudinary_configured():
        old_pid = worker.get("photo_public_id")
        if old_pid:
            cloudinary_delete(old_pid)
        try:
            result = upload_image(data, "business_gallery", f"workers/{worker_id}")
            photo_url = result["secure_url"]
            public_id = result["public_id"]
        except Exception as e:
            logger.error(f"Worker photo upload failed: {e}")
            raise HTTPException(status_code=500, detail="Failed to upload photo")
    else:
        path = generate_upload_path(user["business_id"], f"worker_{worker_id}_{file.filename}")
        content_type = "image/jpeg" if file.filename.lower().endswith(("jfif", "jpg", "jpeg")) else file.content_type
        try:
            result = put_object(path, data, content_type)
            base_url = os.environ.get("BASE_URL", "")
            photo_url = f"{base_url}/api/files/{result['path']}"
            public_id = path
        except Exception as e:
            logger.error(f"Worker photo upload failed: {e}")
            raise HTTPException(status_code=500, detail="Failed to upload photo")

    await db.workers.update_one(
        {"id": worker_id},
        {"$set": {"photo_url": photo_url, "photo_public_id": public_id}}
    )

    return {"secure_url": photo_url, "public_id": public_id}



async def reactivate_worker(worker_id: str, token_data: TokenData = Depends(require_business)):
    """Reactivate a previously deactivated worker"""
    user = await db.users.find_one({"id": token_data.user_id})
    worker = await db.workers.find_one({"id": worker_id, "business_id": user.get("business_id")})
    if not worker:
        raise HTTPException(status_code=404, detail="Worker not found")
    
    await db.workers.update_one(
        {"id": worker_id},
        {"$set": {"active": True}, "$unset": {"deactivated_at": ""}}
    )
    return {"message": "Worker reactivated", "worker_id": worker_id}

@businesses_router.put("/my/workers/{worker_id}/schedule")
async def update_worker_schedule(
    worker_id: str,
    schedule_update: WorkerScheduleUpdate,
    token_data: TokenData = Depends(require_business)
):
    """Update worker schedule with validation for overlapping blocks"""
    user = await db.users.find_one({"id": token_data.user_id})
    worker = await db.workers.find_one({"id": worker_id, "business_id": user.get("business_id")})
    if not worker:
        raise HTTPException(status_code=404, detail="Worker not found")
    
    # Validate no overlapping blocks
    validate_schedule_blocks(schedule_update.schedule)
    
    # Convert to dict for storage
    schedule_dict = {}
    for day, day_schedule in schedule_update.schedule.items():
        schedule_dict[day] = {
            "is_available": day_schedule.is_available,
            "blocks": [{"start_time": b.start_time, "end_time": b.end_time} for b in day_schedule.blocks]
        }
    
    await db.workers.update_one({"id": worker_id}, {"$set": {"schedule": schedule_dict}})
    updated = await db.workers.find_one({"id": worker_id}, {"_id": 0})
    return WorkerResponse(**updated)

@businesses_router.post("/my/workers/{worker_id}/exceptions")
async def add_worker_exception(
    worker_id: str,
    exception_data: WorkerExceptionAdd,
    token_data: TokenData = Depends(require_business)
):
    """Add an exception (vacation/block) to a worker's schedule"""
    user = await db.users.find_one({"id": token_data.user_id})
    worker = await db.workers.find_one({"id": worker_id, "business_id": user.get("business_id")})
    if not worker:
        raise HTTPException(status_code=404, detail="Worker not found")
    
    exception = exception_data.exception
    
    # Validate dates
    try:
        start = datetime.strptime(exception.start_date, "%Y-%m-%d")
        end = datetime.strptime(exception.end_date, "%Y-%m-%d")
        if end < start:
            raise HTTPException(status_code=400, detail="end_date cannot be before start_date")
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")
    
    # Validate times if provided
    if exception.start_time and exception.end_time:
        try:
            st = datetime.strptime(exception.start_time, "%H:%M")
            et = datetime.strptime(exception.end_time, "%H:%M")
            if et <= st:
                raise HTTPException(status_code=400, detail="end_time must be after start_time")
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid time format. Use HH:MM")
    
    exception_dict = {
        "id": generate_id(),
        "start_date": exception.start_date,
        "end_date": exception.end_date,
        "start_time": exception.start_time,
        "end_time": exception.end_time,
        "reason": exception.reason,
        "exception_type": exception.exception_type,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.workers.update_one(
        {"id": worker_id},
        {"$push": {"exceptions": exception_dict}}
    )
    
    return {"message": "Exception added", "exception_id": exception_dict["id"]}

@businesses_router.delete("/my/workers/{worker_id}/exceptions/{exception_id}")
async def remove_worker_exception(
    worker_id: str,
    exception_id: str,
    token_data: TokenData = Depends(require_business)
):
    """Remove an exception from a worker's schedule"""
    user = await db.users.find_one({"id": token_data.user_id})
    worker = await db.workers.find_one({"id": worker_id, "business_id": user.get("business_id")})
    if not worker:
        raise HTTPException(status_code=404, detail="Worker not found")
    
    await db.workers.update_one(
        {"id": worker_id},
        {"$pull": {"exceptions": {"id": exception_id}}}
    )
    
    return {"message": "Exception removed"}


# ========================== OWNER PIN & MANAGER ROUTES ==========================

@businesses_router.post("/me/pin")
async def set_owner_pin(data: PinCreate, token_data: TokenData = Depends(require_business)):
    """Set or update the owner's security PIN (4-6 digits)"""
    if not data.pin.isdigit() or not (4 <= len(data.pin) <= 6):
        raise HTTPException(status_code=400, detail="PIN must be 4-6 digits")
    
    user = await db.users.find_one({"id": token_data.user_id})
    pin_hash = hash_password(data.pin)
    await db.businesses.update_one(
        {"id": user["business_id"]},
        {"$set": {"owner_pin_hash": pin_hash}}
    )
    return {"message": "PIN created successfully"}

@businesses_router.post("/me/pin/verify")
async def verify_owner_pin(data: PinVerify, token_data: TokenData = Depends(require_business)):
    """Verify the owner's PIN"""
    user = await db.users.find_one({"id": token_data.user_id})
    business = await db.businesses.find_one({"id": user["business_id"]})
    pin_hash = business.get("owner_pin_hash")
    if not pin_hash:
        raise HTTPException(status_code=400, detail="No PIN configured")
    if not verify_password(data.pin, pin_hash):
        raise HTTPException(status_code=401, detail="Incorrect PIN")
    return {"verified": True}

@businesses_router.get("/me/pin/status")
async def get_pin_status(token_data: TokenData = Depends(require_business)):
    """Check if owner has a PIN configured"""
    user = await db.users.find_one({"id": token_data.user_id})
    business = await db.businesses.find_one({"id": user["business_id"]})
    return {"has_pin": bool(business.get("owner_pin_hash"))}

@businesses_router.put("/my/workers/{worker_id}/manager")
async def designate_manager(worker_id: str, data: ManagerDesignate, token_data: TokenData = Depends(require_business)):
    """Designate a worker as manager with custom permissions"""
    user = await db.users.find_one({"id": token_data.user_id})
    business_id = user.get("business_id")
    
    worker = await db.workers.find_one({"id": worker_id, "business_id": business_id})
    if not worker:
        raise HTTPException(status_code=404, detail="Worker not found")
    
    # Check max managers (limit 2)
    current_managers = await db.workers.count_documents({"business_id": business_id, "is_manager": True, "id": {"$ne": worker_id}})
    if current_managers >= 2:
        raise HTTPException(status_code=400, detail="Maximum 2 managers allowed")
    
    # Merge provided permissions with defaults
    permissions = {**DEFAULT_MANAGER_PERMISSIONS, **data.permissions}
    
    await db.workers.update_one(
        {"id": worker_id},
        {"$set": {
            "is_manager": True,
            "manager_permissions": permissions,
            "manager_designated_at": datetime.now(timezone.utc).isoformat(),
            "manager_designated_by": token_data.user_id,
        }}
    )
    
    updated = await db.workers.find_one({"id": worker_id}, {"_id": 0})
    updated["has_manager_pin"] = bool(updated.get("manager_pin_hash"))
    
    # Log activity
    await create_business_activity(
        business_id, token_data, "designate_admin", "worker", worker_id,
        {"worker_name": worker.get("name", ""), "permissions": permissions}
    )
    
    return WorkerResponse(**updated)

@businesses_router.put("/my/workers/{worker_id}/manager/permissions")
async def update_manager_permissions(worker_id: str, data: ManagerPermissionsUpdate, token_data: TokenData = Depends(require_business)):
    """Update a manager's permissions"""
    user = await db.users.find_one({"id": token_data.user_id})
    worker = await db.workers.find_one({"id": worker_id, "business_id": user.get("business_id"), "is_manager": True})
    if not worker:
        raise HTTPException(status_code=404, detail="Manager not found")
    
    permissions = {**worker.get("manager_permissions", DEFAULT_MANAGER_PERMISSIONS), **data.permissions}
    await db.workers.update_one({"id": worker_id}, {"$set": {"manager_permissions": permissions}})
    
    # Log activity
    await create_business_activity(
        user.get("business_id"), token_data, "update_permissions", "worker", worker_id,
        {"worker_name": worker.get("name", ""), "permissions": permissions}
    )
    
    return {"message": "Permissions updated"}

@businesses_router.delete("/my/workers/{worker_id}/manager")
async def remove_manager(worker_id: str, token_data: TokenData = Depends(require_business)):
    """Remove manager role from a worker"""
    user = await db.users.find_one({"id": token_data.user_id})
    worker = await db.workers.find_one({"id": worker_id, "business_id": user.get("business_id")})
    if not worker:
        raise HTTPException(status_code=404, detail="Worker not found")
    
    await db.workers.update_one(
        {"id": worker_id},
        {"$set": {"is_manager": False}, "$unset": {"manager_pin_hash": "", "manager_permissions": "", "manager_designated_at": "", "manager_designated_by": ""}}
    )
    
    # Log activity
    await create_business_activity(
        user.get("business_id"), token_data, "remove_admin", "worker", worker_id,
        {"worker_name": worker.get("name", "")}
    )
    
    return {"message": "Manager role removed"}

@businesses_router.post("/my/workers/{worker_id}/manager/pin")
async def set_manager_pin(worker_id: str, data: PinCreate, token_data: TokenData = Depends(require_business)):
    """Set or reset a manager's PIN (can be called by owner or manager themselves)"""
    if not data.pin.isdigit() or not (4 <= len(data.pin) <= 6):
        raise HTTPException(status_code=400, detail="PIN must be 4-6 digits")
    
    user = await db.users.find_one({"id": token_data.user_id})
    worker = await db.workers.find_one({"id": worker_id, "business_id": user.get("business_id"), "is_manager": True})
    if not worker:
        raise HTTPException(status_code=404, detail="Manager not found")
    
    pin_hash = hash_password(data.pin)
    await db.workers.update_one({"id": worker_id}, {"$set": {"manager_pin_hash": pin_hash}})
    return {"message": "Manager PIN set successfully"}

@businesses_router.get("/my/activity-log")
async def get_business_activity_log(
    page: int = 1,
    limit: int = 30,
    actor_type: Optional[str] = None,
    action: Optional[str] = None,
    token_data: TokenData = Depends(require_business)
):
    """Get business activity log (owner only)"""
    if token_data.is_manager:
        raise HTTPException(status_code=403, detail="Solo el dueño puede ver el historial de actividad")
    
    user = await db.users.find_one({"id": token_data.user_id})
    business_id = user.get("business_id")
    
    query = {"business_id": business_id}
    if actor_type:
        query["actor_type"] = actor_type
    if action:
        query["action"] = action
    
    total = await db.business_activity_logs.count_documents(query)
    skip = (page - 1) * limit
    
    logs = await db.business_activity_logs.find(
        query, {"_id": 0}
    ).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    
    return {"logs": logs, "total": total, "page": page, "pages": (total + limit - 1) // limit if total > 0 else 1}


# ========================== SERVICE ROUTES ==========================

@services_router.post("", response_model=ServiceResponse)
async def create_service(service: ServiceCreate, token_data: TokenData = Depends(require_business)):
    user = await db.users.find_one({"id": token_data.user_id})
    if not user or not user.get("business_id"):
        raise HTTPException(status_code=404, detail="Business not found")
    
    service_doc = {
        "id": generate_id(),
        "business_id": user["business_id"],
        **service.model_dump(),
        "active": True,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.services.insert_one(service_doc)
    return ServiceResponse(**service_doc)

@services_router.get("/business/{business_id}", response_model=List[ServiceResponse])
async def get_business_services(business_id: str):
    services = await db.services.find({"business_id": business_id, "active": True}, {"_id": 0}).to_list(100)
    return [ServiceResponse(**s) for s in services]

@services_router.put("/{service_id}", response_model=ServiceResponse)
async def update_service(service_id: str, update: ServiceUpdate, token_data: TokenData = Depends(require_business)):
    user = await db.users.find_one({"id": token_data.user_id})
    service = await db.services.find_one({"id": service_id, "business_id": user.get("business_id")})
    if not service:
        raise HTTPException(status_code=404, detail="Service not found")
    
    update_data = {k: v for k, v in update.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="No data to update")
    
    await db.services.update_one({"id": service_id}, {"$set": update_data})
    updated = await db.services.find_one({"id": service_id}, {"_id": 0})
    return ServiceResponse(**updated)

@services_router.delete("/{service_id}")
async def delete_service(service_id: str, token_data: TokenData = Depends(require_business)):
    user = await db.users.find_one({"id": token_data.user_id})
    service = await db.services.find_one({"id": service_id, "business_id": user.get("business_id")})
    if not service:
        raise HTTPException(status_code=404, detail="Service not found")
    
    await db.services.update_one({"id": service_id}, {"$set": {"active": False}})
    return {"message": "Service deleted"}

# ========================== BOOKING ROUTES ==========================

# ========================== AVAILABILITY ENGINE ==========================

class SlotStatus(str, Enum):
    AVAILABLE = "available"
    BOOKED = "booked"
    HOLD = "hold"
    EXCEPTION = "exception"
    OUTSIDE_SCHEDULE = "outside_schedule"
    BUFFER = "buffer"

class AvailabilitySlot(BaseModel):
    time: str
    end_time: str
    status: str
    reason: Optional[str] = None
    worker_id: Optional[str] = None
    worker_name: Optional[str] = None

class AvailabilityResponse(BaseModel):
    date: str
    business_timezone: str
    slots: List[AvailabilitySlot]
    available_count: int
    total_workers: int

def is_exception_blocking(exception: dict, date_str: str, slot_start: datetime, slot_end: datetime) -> tuple:
    """Check if an exception blocks a slot. Returns (is_blocking, reason)"""
    exc_start_date = datetime.strptime(exception["start_date"], "%Y-%m-%d").date()
    exc_end_date = datetime.strptime(exception["end_date"], "%Y-%m-%d").date()
    slot_date = datetime.strptime(date_str, "%Y-%m-%d").date()
    
    # Check if date is within exception range
    if not (exc_start_date <= slot_date <= exc_end_date):
        return False, None
    
    # If no times specified, it's a full day block
    if not exception.get("start_time") or not exception.get("end_time"):
        reason = exception.get("reason") or exception.get("exception_type", "Bloqueo")
        return True, f"{reason} (día completo)"
    
    # Check time overlap
    exc_start_time = datetime.strptime(exception["start_time"], "%H:%M")
    exc_end_time = datetime.strptime(exception["end_time"], "%H:%M")
    
    # Convert slot times to same format for comparison
    slot_start_time = datetime.strptime(slot_start.strftime("%H:%M"), "%H:%M")
    slot_end_time = datetime.strptime(slot_end.strftime("%H:%M"), "%H:%M")
    
    # Check overlap
    if not (slot_end_time <= exc_start_time or slot_start_time >= exc_end_time):
        reason = exception.get("reason") or exception.get("exception_type", "Bloqueo")
        return True, f"{reason} ({exception['start_time']}-{exception['end_time']})"
    
    return False, None

@bookings_router.get("/availability/{business_id}", response_model=AvailabilityResponse)
async def get_availability(
    business_id: str,
    date: str,
    service_id: Optional[str] = None,
    worker_id: Optional[str] = None,
    include_unavailable: bool = False
):
    """
    Get available time slots for a business on a specific date.
    
    Args:
        business_id: The business ID
        date: Date in YYYY-MM-DD format
        service_id: Optional service ID to filter workers who can perform it
        worker_id: Optional worker ID to get availability for specific worker
        include_unavailable: If True, returns all slots with reasons for unavailability
    
    Returns:
        AvailabilityResponse with slots showing availability status and reasons
    """
    business = await db.businesses.find_one({"id": business_id})
    if not business:
        raise HTTPException(status_code=404, detail="Business not found")
    
    business_tz_name = business.get("timezone", "America/Mexico_City")
    try:
        business_tz = pytz.timezone(business_tz_name)
    except pytz.UnknownTimeZoneError:
        business_tz = pytz.timezone("America/Mexico_City")
    
    # Get buffer time between appointments
    buffer_minutes = business.get("min_time_between_appointments", 0)
    
    # Check if business is closed on this date
    closure = await db.business_closures.find_one({
        "business_id": business_id, "date": date, "is_deleted": False
    })
    if closure:
        return AvailabilityResponse(
            date=date,
            business_timezone=business_tz_name,
            slots=[],
            available_count=0,
            total_workers=0
        )
    
    # Get service info
    service = None
    duration = 60
    allowed_workers = None
    
    if service_id:
        service = await db.services.find_one({"id": service_id})
        if service:
            duration = service.get("duration_minutes", 60)
            if service.get("allowed_worker_ids"):
                allowed_workers = service["allowed_worker_ids"]
    
    # Build worker filter
    worker_filter = {"business_id": business_id, "active": True}
    if worker_id:
        worker_filter["id"] = worker_id
    
    workers = await db.workers.find(worker_filter, {"_id": 0}).to_list(100)
    
    # Filter workers by service: check worker.service_ids contains the service
    if service_id:
        workers = [w for w in workers if not w.get("service_ids") or service_id in w.get("service_ids", [])]
    
    # Also filter by allowed_worker_ids from service model
    if allowed_workers:
        workers = [w for w in workers if w["id"] in allowed_workers]
    
    if not workers:
        return AvailabilityResponse(
            date=date,
            business_timezone=business_tz_name,
            slots=[],
            available_count=0,
            total_workers=0
        )
    
    # Get existing bookings for the date (HOLD and CONFIRMED)
    existing_bookings = await db.bookings.find({
        "business_id": business_id,
        "date": date,
        "status": {"$in": [AppointmentStatus.HOLD, AppointmentStatus.CONFIRMED]}
    }, {"_id": 0}).to_list(1000)
    
    # Parse date
    date_obj = datetime.strptime(date, "%Y-%m-%d")
    day_of_week = str(date_obj.weekday())
    
    # Get current time in business timezone for filtering past slots
    now_utc = datetime.now(timezone.utc)
    now_local = now_utc.astimezone(business_tz)
    is_today = date_obj.date() == now_local.date()
    
    slots_dict = {}  # time -> best slot info
    
    for worker in workers:
        schedule = worker.get("schedule", {}).get(day_of_week)
        
        # Check if worker is available this day
        if not schedule or not schedule.get("is_available", True):
            continue
        
        # Get all time blocks for this day
        blocks = schedule.get("blocks", [])
        if not blocks:
            # Legacy support: convert old format
            if schedule.get("start_time") and schedule.get("end_time"):
                blocks = [{"start_time": schedule["start_time"], "end_time": schedule["end_time"]}]
            else:
                continue
        
        for block in blocks:
            block_start = datetime.strptime(block["start_time"], "%H:%M")
            block_end = datetime.strptime(block["end_time"], "%H:%M")
            
            current_time = block_start
            while current_time + timedelta(minutes=duration) <= block_end:
                time_str = current_time.strftime("%H:%M")
                slot_end = current_time + timedelta(minutes=duration)
                end_time_str = slot_end.strftime("%H:%M")
                
                # Skip past slots if today
                if is_today:
                    slot_datetime = now_local.replace(
                        hour=current_time.hour,
                        minute=current_time.minute,
                        second=0,
                        microsecond=0
                    )
                    if slot_datetime <= now_local:
                        current_time += timedelta(minutes=30)
                        continue
                
                status = SlotStatus.AVAILABLE
                reason = None
                
                # Check exceptions (vacations/blocks)
                for exc in worker.get("exceptions", []):
                    is_blocking, exc_reason = is_exception_blocking(exc, date, current_time, slot_end)
                    if is_blocking:
                        status = SlotStatus.EXCEPTION
                        reason = exc_reason
                        break
                
                # Check existing bookings
                if status == SlotStatus.AVAILABLE:
                    for booking in existing_bookings:
                        if booking["worker_id"] == worker["id"]:
                            booking_start = datetime.strptime(booking["time"], "%H:%M")
                            booking_end = datetime.strptime(booking["end_time"], "%H:%M")
                            
                            # Add buffer to booking end
                            booking_end_with_buffer = booking_end + timedelta(minutes=buffer_minutes)
                            
                            # Check overlap (including buffer)
                            if not (slot_end <= booking_start or current_time >= booking_end_with_buffer):
                                if booking["status"] == AppointmentStatus.HOLD:
                                    status = SlotStatus.HOLD
                                    reason = "Reserva en proceso de pago"
                                else:
                                    status = SlotStatus.BOOKED
                                    reason = "Ocupado"
                                break
                            
                            # Check buffer before booking
                            if buffer_minutes > 0:
                                buffer_start = booking_start - timedelta(minutes=buffer_minutes)
                                if buffer_start < slot_end <= booking_start and current_time < booking_start:
                                    status = SlotStatus.BUFFER
                                    reason = f"Buffer de {buffer_minutes} min antes de cita"
                                    break
                
                # Store best slot for this time (prefer available)
                if time_str not in slots_dict:
                    slots_dict[time_str] = {
                        "time": time_str,
                        "end_time": end_time_str,
                        "status": status,
                        "reason": reason,
                        "worker_id": worker["id"] if status == SlotStatus.AVAILABLE else None,
                        "worker_name": worker["name"] if status == SlotStatus.AVAILABLE else None
                    }
                elif status == SlotStatus.AVAILABLE and slots_dict[time_str]["status"] != SlotStatus.AVAILABLE:
                    # This worker is available for this slot
                    slots_dict[time_str] = {
                        "time": time_str,
                        "end_time": end_time_str,
                        "status": status,
                        "reason": None,
                        "worker_id": worker["id"],
                        "worker_name": worker["name"]
                    }
                
                current_time += timedelta(minutes=30)  # 30 min intervals
    
    # Convert to list and sort
    all_slots = sorted(slots_dict.values(), key=lambda s: s["time"])
    
    # Filter if not including unavailable
    if not include_unavailable:
        all_slots = [s for s in all_slots if s["status"] == SlotStatus.AVAILABLE]
    
    available_count = len([s for s in all_slots if s["status"] == SlotStatus.AVAILABLE])
    
    return AvailabilityResponse(
        date=date,
        business_timezone=business_tz_name,
        slots=[AvailabilitySlot(**s) for s in all_slots],
        available_count=available_count,
        total_workers=len(workers)
    )

@bookings_router.post("/", response_model=BookingResponse)
async def create_booking(booking: BookingCreate, token_data: TokenData = Depends(require_auth)):
    # Check user limits (skip for business users)
    user = await db.users.find_one({"id": token_data.user_id})
    if token_data.role != UserRole.BUSINESS and user.get("active_appointments_count", 0) >= 5:
        raise HTTPException(status_code=400, detail="Maximum 5 active appointments allowed")
    
    # Check suspension
    if user.get("suspended_until"):
        suspended_until = datetime.fromisoformat(user["suspended_until"])
        if suspended_until > datetime.now(timezone.utc):
            raise HTTPException(status_code=403, detail="Account suspended")
    
    # Get business
    business = await db.businesses.find_one({"id": booking.business_id})
    if not business:
        raise HTTPException(status_code=404, detail="Business not found")
    
    if business["status"] != BusinessStatus.APPROVED:
        raise HTTPException(status_code=400, detail="Business is not accepting bookings")
    
    if business.get("subscription_status") in ("canceled", "past_due", "unpaid"):
        raise HTTPException(status_code=400, detail="Business subscription is not active")
    
    # Check blacklist (skip for business users)
    if token_data.role != UserRole.BUSINESS and await is_user_blacklisted(booking.business_id, user_id=token_data.user_id):
        raise HTTPException(status_code=404, detail="Business not found")
    
    # Get service
    service = await db.services.find_one({"id": booking.service_id})
    if not service:
        raise HTTPException(status_code=404, detail="Service not found")
    
    # Get allowed workers for this service
    allowed_worker_ids = service.get("allowed_worker_ids", [])
    
    # Determine worker
    worker_id = booking.worker_id
    if not worker_id:
        # Auto-assign to worker with least appointments
        worker_filter = {"business_id": booking.business_id, "active": True}
        
        # If service has specific workers, filter by them
        if allowed_worker_ids:
            worker_filter["id"] = {"$in": allowed_worker_ids}
        
        workers = await db.workers.find(worker_filter).to_list(100)
        
        # Filter workers by service_ids (worker must offer this service)
        if service_id:
            workers = [w for w in workers if not w.get("service_ids") or service_id in w.get("service_ids", [])]
        
        if not workers:
            raise HTTPException(status_code=400, detail="No workers available for this service")
        
        # Parse booking date and time
        booking_date_str = booking.date
        booking_time_str = booking.time
        date_obj = datetime.strptime(booking_date_str, "%Y-%m-%d")
        day_of_week = str(date_obj.weekday())
        slot_start = datetime.strptime(booking_time_str, "%H:%M")
        slot_end = slot_start + timedelta(minutes=service["duration_minutes"])
        
        # Filter workers by availability on this specific slot
        eligible_workers = []
        for w in workers:
            # Check schedule
            schedule = w.get("schedule", {}).get(day_of_week)
            if not schedule or not schedule.get("is_available", True):
                continue
            
            # Check if slot falls within any block
            blocks = schedule.get("blocks", [])
            if not blocks:
                # Legacy support
                if schedule.get("start_time") and schedule.get("end_time"):
                    blocks = [{"start_time": schedule["start_time"], "end_time": schedule["end_time"]}]
            
            slot_in_schedule = False
            for block in blocks:
                block_start = datetime.strptime(block["start_time"], "%H:%M")
                block_end = datetime.strptime(block["end_time"], "%H:%M")
                if block_start <= slot_start and slot_end <= block_end:
                    slot_in_schedule = True
                    break
            
            if not slot_in_schedule:
                continue
            
            # Check exceptions
            is_blocked = False
            for exc in w.get("exceptions", []):
                blocking, _ = is_exception_blocking(exc, booking_date_str, slot_start, slot_end)
                if blocking:
                    is_blocked = True
                    break
            
            if is_blocked:
                continue
            
            eligible_workers.append(w)
        
        if not eligible_workers:
            raise HTTPException(status_code=400, detail="No workers available for this time slot")
        
        # Count appointments per eligible worker for this date (include HOLD status)
        worker_counts = {}
        for w in eligible_workers:
            count = await db.bookings.count_documents({
                "worker_id": w["id"],
                "date": booking.date,
                "status": {"$in": [AppointmentStatus.HOLD, AppointmentStatus.CONFIRMED]}
            })
            worker_counts[w["id"]] = count
        
        # Choose worker with least load
        worker_id = min(worker_counts, key=worker_counts.get)
    else:
        # Validate specific worker can perform this service
        if allowed_worker_ids and worker_id not in allowed_worker_ids:
            raise HTTPException(status_code=400, detail="Selected worker cannot perform this service")
    
    worker = await db.workers.find_one({"id": worker_id})
    if not worker:
        raise HTTPException(status_code=404, detail="Worker not found")
    
    # Also check worker's own service_ids
    if worker.get("service_ids") and booking.service_id not in worker.get("service_ids", []):
        raise HTTPException(status_code=400, detail="Selected worker does not offer this service")
    
    if not worker.get("active", True):
        raise HTTPException(status_code=400, detail="Worker is not active")
    
    # Check if slot is already taken (including HOLD) with buffer consideration
    buffer_minutes = business.get("min_time_between_appointments", 0)
    start_time_dt = datetime.strptime(booking.time, "%H:%M")
    end_time_dt = start_time_dt + timedelta(minutes=service["duration_minutes"])
    
    # Get all bookings for this worker on this date
    existing_bookings = await db.bookings.find({
        "worker_id": worker_id,
        "date": booking.date,
        "status": {"$in": [AppointmentStatus.HOLD, AppointmentStatus.CONFIRMED]}
    }).to_list(100)
    
    for existing in existing_bookings:
        ex_start = datetime.strptime(existing["time"], "%H:%M")
        ex_end = datetime.strptime(existing["end_time"], "%H:%M")
        ex_end_with_buffer = ex_end + timedelta(minutes=buffer_minutes)
        
        # Check overlap with buffer
        if not (end_time_dt <= ex_start or start_time_dt >= ex_end_with_buffer):
            raise HTTPException(status_code=409, detail="Slot conflicts with existing booking")
    
    # Calculate deposit amount (minimum 50 MXN)
    deposit_amount = max(
        business.get("deposit_amount", MIN_DEPOSIT_AMOUNT) if business.get("requires_deposit") else MIN_DEPOSIT_AMOUNT,
        MIN_DEPOSIT_AMOUNT
    )
    
    # Determine if this is a business-created booking (skip payment)
    is_biz_booking = booking.skip_payment and token_data.role == UserRole.BUSINESS
    now_iso = datetime.now(timezone.utc).isoformat()
    
    if is_biz_booking:
        # Business creates booking: confirmed directly, no deposit needed
        booking_doc = {
            "id": generate_id(),
            "user_id": token_data.user_id,
            "business_id": booking.business_id,
            "service_id": booking.service_id,
            "worker_id": worker_id,
            "date": booking.date,
            "time": booking.time,
            "end_time": end_time_dt.strftime("%H:%M"),
            "duration_minutes": service["duration_minutes"],
            "status": AppointmentStatus.CONFIRMED,
            "notes": booking.notes,
            "is_home_service": booking.is_home_service,
            "address": booking.address,
            "deposit_amount": 0,
            "deposit_paid": True,
            "total_amount": service["price"],
            "transaction_id": None,
            "stripe_session_id": None,
            "hold_expires_at": None,
            "created_at": now_iso,
            "confirmed_at": now_iso,
            "cancelled_at": None,
            "cancelled_by": None,
            "cancellation_reason": None,
            "client_name": booking.client_name,
            "client_email": booking.client_email,
            "client_phone": booking.client_phone,
            "client_info": booking.client_info,
        }
    else:
        # Regular user: hold status, pending payment
        hold_expires_at = datetime.now(timezone.utc) + timedelta(minutes=HOLD_EXPIRATION_MINUTES)
        booking_doc = {
            "id": generate_id(),
            "user_id": token_data.user_id,
            "business_id": booking.business_id,
            "service_id": booking.service_id,
            "worker_id": worker_id,
            "date": booking.date,
            "time": booking.time,
            "end_time": end_time_dt.strftime("%H:%M"),
            "duration_minutes": service["duration_minutes"],
            "status": AppointmentStatus.HOLD,
            "notes": booking.notes,
            "is_home_service": booking.is_home_service,
            "address": booking.address,
            "deposit_amount": deposit_amount,
            "deposit_paid": False,
            "total_amount": service["price"],
            "transaction_id": None,
            "stripe_session_id": None,
            "hold_expires_at": hold_expires_at.isoformat(),
            "created_at": now_iso,
            "confirmed_at": None,
            "cancelled_at": None,
            "cancelled_by": None,
            "cancellation_reason": None,
        }
    
    await db.bookings.insert_one(booking_doc)
    
    # Update user active appointments count
    await db.users.update_one(
        {"id": token_data.user_id},
        {"$inc": {"active_appointments_count": 1}}
    )
    
    # Notifications
    if is_biz_booking:
        await create_notification(
            business["user_id"],
            "Cita Registrada",
            f"Cita registrada para {booking.client_name or 'Cliente'} - {service['name']} el {booking.date} a las {booking.time}",
            "booking",
            {"booking_id": booking_doc["id"]}
        )
    else:
        await create_notification(
            business["user_id"],
            "Nueva Reserva (Pendiente de Pago)",
            f"Nueva reserva de {user['full_name']} para {service['name']} - Esperando pago del anticipo",
            "booking",
            {"booking_id": booking_doc["id"]}
        )
    
    booking_doc["business_name"] = business["name"]
    booking_doc["service_name"] = service["name"]
    booking_doc["worker_name"] = worker["name"]
    
    return BookingResponse(**booking_doc)

@bookings_router.get("/my", response_model=List[BookingResponse])
async def get_my_bookings(
    status: Optional[str] = None,
    upcoming: bool = True,
    token_data: TokenData = Depends(require_auth)
):
    filters = {"user_id": token_data.user_id}
    
    if status:
        filters["status"] = status
    
    if upcoming:
        today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
        filters["date"] = {"$gte": today}
    else:
        today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
        filters["date"] = {"$lt": today}
    
    bookings = await db.bookings.find(filters, {"_id": 0}).sort("date", 1).to_list(100)
    
    now = datetime.now(timezone.utc)
    
    # Populate names and calculate fields
    for b in bookings:
        business = await db.businesses.find_one({"id": b["business_id"]})
        service = await db.services.find_one({"id": b["service_id"]})
        worker = await db.workers.find_one({"id": b["worker_id"]})
        
        b["business_name"] = business["name"] if business else None
        b["business_slug"] = business.get("slug") if business else None
        b["service_name"] = service["name"] if service else None
        b["worker_name"] = worker["name"] if worker else None
        
        # Check if already reviewed
        existing_review = await db.reviews.find_one({"booking_id": b["id"]})
        b["has_review"] = existing_review is not None
        
        # Calculate hours until appointment
        try:
            appointment_dt = datetime.strptime(f"{b['date']} {b['time']}", "%Y-%m-%d %H:%M")
            appointment_dt = appointment_dt.replace(tzinfo=timezone.utc)
            hours_diff = (appointment_dt - now).total_seconds() / 3600
            b["hours_until_appointment"] = round(hours_diff, 2)
            # Can cancel only if > 24h (for policy) or if still in HOLD
            b["can_cancel"] = hours_diff > 0 and (b["status"] in [AppointmentStatus.HOLD, AppointmentStatus.CONFIRMED])
        except:
            b["hours_until_appointment"] = None
            b["can_cancel"] = False
    
    return [BookingResponse(**b) for b in bookings]

@bookings_router.get("/business", response_model=List[BookingResponse])
async def get_business_bookings(
    date: Optional[str] = None,
    status: Optional[str] = None,
    token_data: TokenData = Depends(require_business)
):
    user = await db.users.find_one({"id": token_data.user_id})
    if not user or not user.get("business_id"):
        raise HTTPException(status_code=404, detail="Business not found")
    
    filters = {"business_id": user["business_id"]}
    
    if date:
        filters["date"] = date
    if status:
        filters["status"] = status
    
    bookings = await db.bookings.find(filters, {"_id": 0}).sort("date", 1).to_list(1000)
    
    now = datetime.now(timezone.utc)
    
    # Populate names
    for b in bookings:
        booking_user = await db.users.find_one({"id": b["user_id"]})
        service = await db.services.find_one({"id": b["service_id"]})
        worker = await db.workers.find_one({"id": b["worker_id"]})
        
        b["user_name"] = booking_user["full_name"] if booking_user else None
        b["user_email"] = booking_user.get("email") if booking_user else b.get("client_email")
        b["user_phone"] = booking_user.get("phone") if booking_user else b.get("client_phone")
        b["service_name"] = service["name"] if service else None
        b["worker_name"] = worker["name"] if worker else None
        
        # Calculate hours until appointment
        try:
            appointment_dt = datetime.strptime(f"{b['date']} {b['time']}", "%Y-%m-%d %H:%M")
            appointment_dt = appointment_dt.replace(tzinfo=timezone.utc)
            hours_diff = (appointment_dt - now).total_seconds() / 3600
            b["hours_until_appointment"] = round(hours_diff, 2)
            b["can_cancel"] = hours_diff > 0
        except:
            b["hours_until_appointment"] = None
            b["can_cancel"] = False
    
    return [BookingResponse(**b) for b in bookings]

@bookings_router.put("/{booking_id}/reschedule")
async def reschedule_booking(
    booking_id: str,
    new_date: str,
    new_time: str,
    token_data: TokenData = Depends(require_auth)
):
    booking = await db.bookings.find_one({"id": booking_id, "user_id": token_data.user_id})
    if not booking:
        raise HTTPException(status_code=404, detail="Booking not found")
    
    # Check if >24h before
    booking_datetime = datetime.strptime(f"{booking['date']} {booking['time']}", "%Y-%m-%d %H:%M")
    booking_datetime = booking_datetime.replace(tzinfo=timezone.utc)
    hours_until = (booking_datetime - datetime.now(timezone.utc)).total_seconds() / 3600
    
    if hours_until <= 24:
        raise HTTPException(status_code=400, detail="Cannot reschedule less than 24 hours before appointment")
    
    # Calculate new end time
    service = await db.services.find_one({"id": booking["service_id"]})
    start_time = datetime.strptime(new_time, "%H:%M")
    end_time = start_time + timedelta(minutes=service["duration_minutes"])
    
    await db.bookings.update_one(
        {"id": booking_id},
        {"$set": {
            "date": new_date,
            "time": new_time,
            "end_time": end_time.strftime("%H:%M"),
            "rescheduled_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {"message": "Booking rescheduled"}


class RescheduleByBusinessRequest(BaseModel):
    new_date: str
    new_time: str
    new_worker_id: Optional[str] = None

@bookings_router.put("/{booking_id}/reschedule/business")
async def reschedule_booking_by_business(
    booking_id: str,
    req: RescheduleByBusinessRequest,
    token_data: TokenData = Depends(require_business)
):
    """Reschedule a booking by business owner - no 24h restriction, frees old slot"""
    booking = await db.bookings.find_one({"id": booking_id})
    if not booking:
        raise HTTPException(status_code=404, detail="Booking not found")
    
    user = await db.users.find_one({"id": token_data.user_id})
    if not user or user.get("business_id") != booking["business_id"]:
        raise HTTPException(status_code=403, detail="Not your business booking")
    
    if booking["status"] not in [AppointmentStatus.CONFIRMED, AppointmentStatus.HOLD]:
        raise HTTPException(status_code=400, detail="Only confirmed or hold bookings can be rescheduled")
    
    # Get service duration
    service = await db.services.find_one({"id": booking["service_id"]})
    duration = service["duration_minutes"] if service else booking.get("duration_minutes", 60)
    
    # Calculate new end time
    start_dt = datetime.strptime(req.new_time, "%H:%M")
    end_dt = start_dt + timedelta(minutes=duration)
    new_end_time = end_dt.strftime("%H:%M")
    
    worker_id = req.new_worker_id or booking["worker_id"]
    
    # Check the new slot is available (exclude current booking)
    conflict = await db.bookings.find_one({
        "business_id": booking["business_id"],
        "worker_id": worker_id,
        "date": req.new_date,
        "status": {"$in": [AppointmentStatus.HOLD, AppointmentStatus.CONFIRMED]},
        "id": {"$ne": booking_id},
        "$or": [
            {"time": {"$lt": new_end_time}, "end_time": {"$gt": req.new_time}},
        ]
    })
    
    if conflict:
        raise HTTPException(status_code=409, detail="El horario seleccionado ya está ocupado")
    
    now = datetime.now(timezone.utc).isoformat()
    old_date = booking["date"]
    old_time = booking["time"]
    
    update_fields = {
        "date": req.new_date,
        "time": req.new_time,
        "end_time": new_end_time,
        "rescheduled_at": now,
        "rescheduled_by": "business",
    }
    if req.new_worker_id:
        update_fields["worker_id"] = req.new_worker_id
    
    await db.bookings.update_one(
        {"id": booking_id},
        {"$set": update_fields}
    )
    
    # Notify client
    try:
        worker = await db.workers.find_one({"id": worker_id})
        worker_name = worker["name"] if worker else ""
        await create_notification(
            booking["user_id"],
            "Cita Reagendada",
            f"Tu cita del {old_date} a las {old_time} fue reagendada al {req.new_date} a las {req.new_time} con {worker_name}",
            "booking",
            {"booking_id": booking_id}
        )
    except Exception as e:
        logger.error(f"Reschedule notification error: {e}")
    
    logger.info(f"Business rescheduled booking {booking_id}: {old_date} {old_time} -> {req.new_date} {req.new_time}")
    
    # Log activity
    await create_business_activity(
        booking["business_id"], token_data, "reschedule_booking", "booking", booking_id,
        {"client_name": booking.get("client_name", ""), "service_name": booking.get("service_name", ""), "old_date": old_date, "old_time": old_time, "new_date": req.new_date, "new_time": req.new_time}
    )
    
    return {
        "message": "Cita reagendada exitosamente",
        "new_date": req.new_date,
        "new_time": req.new_time,
        "new_end_time": new_end_time
    }


@bookings_router.put("/{booking_id}/confirm")
async def confirm_booking(booking_id: str, token_data: TokenData = Depends(require_business)):
    user = await db.users.find_one({"id": token_data.user_id})
    booking = await db.bookings.find_one({"id": booking_id, "business_id": user.get("business_id")})
    if not booking:
        raise HTTPException(status_code=404, detail="Booking not found")
    
    await db.bookings.update_one(
        {"id": booking_id},
        {"$set": {"status": AppointmentStatus.CONFIRMED}}
    )
    
    # Notify user
    await create_notification(
        booking["user_id"],
        "Reserva Confirmada",
        "Tu reserva ha sido confirmada",
        "booking",
        {"booking_id": booking_id}
    )
    
    return {"message": "Booking confirmed"}

@bookings_router.put("/{booking_id}/complete")
async def complete_booking(booking_id: str, token_data: TokenData = Depends(require_business)):
    user = await db.users.find_one({"id": token_data.user_id})
    booking = await db.bookings.find_one({"id": booking_id, "business_id": user.get("business_id")})
    if not booking:
        raise HTTPException(status_code=404, detail="Booking not found")
    
    await db.bookings.update_one(
        {"id": booking_id},
        {"$set": {"status": AppointmentStatus.COMPLETED, "completed_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    # Update business completed appointments
    await db.businesses.update_one(
        {"id": booking["business_id"]},
        {"$inc": {"completed_appointments": 1}}
    )
    
    # Update user active count
    await db.users.update_one(
        {"id": booking["user_id"]},
        {"$inc": {"active_appointments_count": -1}}
    )
    
    # Notify user to leave review
    await create_notification(
        booking["user_id"],
        "Deja tu opinion",
        "Tu cita ha sido completada. Dejanos tu opinion!",
        "review",
        {"booking_id": booking_id, "business_id": booking["business_id"]}
    )
    
    # Log activity
    await create_business_activity(
        booking["business_id"], token_data, "complete_booking", "booking", booking_id,
        {"client_name": booking.get("client_name", ""), "service_name": booking.get("service_name", ""), "date": booking.get("date", ""), "time": booking.get("time", "")}
    )
    
    return {"message": "Booking completed"}

# ========================== REVIEW ROUTES ==========================

@reviews_router.post("/", response_model=ReviewResponse)
async def create_review(review: ReviewCreate, token_data: TokenData = Depends(require_auth)):
    # Verify booking was completed (or confirmed with past date)
    booking = await db.bookings.find_one({
        "id": review.booking_id,
        "user_id": token_data.user_id,
    })
    
    if not booking:
        raise HTTPException(status_code=400, detail="Booking not found")
    
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    is_past = booking.get("date", "") <= today
    allowed_statuses = [AppointmentStatus.COMPLETED]
    if is_past:
        allowed_statuses.append(AppointmentStatus.CONFIRMED)
    
    if booking["status"] not in allowed_statuses:
        raise HTTPException(status_code=400, detail="Solo puedes calificar citas completadas o pasadas")
    
    # Check if already reviewed
    existing = await db.reviews.find_one({"booking_id": review.booking_id})
    if existing:
        raise HTTPException(status_code=400, detail="Already reviewed this booking")
    
    user = await db.users.find_one({"id": token_data.user_id})
    
    review_doc = {
        "id": generate_id(),
        "user_id": token_data.user_id,
        "business_id": review.business_id,
        "booking_id": review.booking_id,
        "rating": min(max(review.rating, 1), 5),
        "comment": review.comment,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "user_name": user["full_name"],
        "user_photo": user.get("photo_url")
    }
    
    await db.reviews.insert_one(review_doc)
    
    # Update business rating (Bayesian)
    business = await db.businesses.find_one({"id": review.business_id})
    new_rating_sum = business.get("rating_sum", 0) + review.rating
    new_review_count = business.get("review_count", 0) + 1
    new_rating = calculate_bayesian_rating(new_rating_sum, new_review_count)
    
    await db.businesses.update_one(
        {"id": review.business_id},
        {"$set": {"rating": round(new_rating, 2), "rating_sum": new_rating_sum, "review_count": new_review_count}}
    )
    
    return ReviewResponse(**review_doc)

@reviews_router.get("/business/{business_id}", response_model=List[ReviewResponse])
async def get_business_reviews(business_id: str, page: int = 1, limit: int = 20):
    skip = (page - 1) * limit
    reviews = await db.reviews.find(
        {"business_id": business_id},
        {"_id": 0}
    ).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    
    return [ReviewResponse(**r) for r in reviews]

# ========================== PAYMENT ROUTES ==========================

# Helper to calculate fees
def calculate_fees(amount: float) -> dict:
    """Calculate platform fee and payout amount"""
    fee_amount = round(amount * PLATFORM_FEE_PERCENT, 2)
    payout_amount = round(amount - fee_amount, 2)
    return {"fee_amount": fee_amount, "payout_amount": payout_amount}

# Create deposit checkout session
@payments_router.post("/deposit/checkout")
async def create_deposit_checkout(
    request: Request, 
    checkout_req: DepositCheckoutRequest, 
    token_data: TokenData = Depends(require_auth)
):
    """Create Stripe Checkout session for booking deposit"""
    # Get booking
    booking = await db.bookings.find_one({"id": checkout_req.booking_id})
    if not booking:
        raise HTTPException(status_code=404, detail="Booking not found")
    
    # Verify ownership
    if booking["user_id"] != token_data.user_id:
        raise HTTPException(status_code=403, detail="Not your booking")
    
    # Check status
    if booking["status"] != AppointmentStatus.HOLD:
        raise HTTPException(status_code=400, detail=f"Booking status is {booking['status']}, cannot pay")
    
    # Check expiration
    hold_expires = datetime.fromisoformat(booking["hold_expires_at"].replace('Z', '+00:00'))
    if hold_expires < datetime.now(timezone.utc):
        # Mark as expired
        await db.bookings.update_one(
            {"id": checkout_req.booking_id},
            {"$set": {"status": AppointmentStatus.EXPIRED}}
        )
        await db.users.update_one(
            {"id": booking["user_id"]},
            {"$inc": {"active_appointments_count": -1}}
        )
        raise HTTPException(status_code=400, detail="Hold expired. Please create a new booking.")
    
    # Check for existing transaction (idempotency)
    existing_tx = await db.transactions.find_one({
        "booking_id": checkout_req.booking_id,
        "status": {"$in": [TransactionStatus.CREATED, TransactionStatus.PAID]}
    })
    if existing_tx and existing_tx.get("stripe_session_id"):
        try:
            existing_session = stripe_lib.checkout.Session.retrieve(existing_tx["stripe_session_id"])
            if existing_session.status != "expired":
                return {"url": existing_session.url, 
                        "session_id": existing_tx["stripe_session_id"],
                        "existing": True}
        except Exception:
            pass
    
    # Get business
    business = await db.businesses.find_one({"id": booking["business_id"]})
    service = await db.services.find_one({"id": booking["service_id"]})
    
    deposit_amount = max(booking.get("deposit_amount", MIN_DEPOSIT_AMOUNT), MIN_DEPOSIT_AMOUNT)
    fees = calculate_fees(deposit_amount)
    
    # Create transaction record
    transaction_id = generate_id()
    transaction_doc = {
        "id": transaction_id,
        "booking_id": checkout_req.booking_id,
        "user_id": token_data.user_id,
        "business_id": booking["business_id"],
        "stripe_session_id": None,
        "stripe_payment_intent_id": None,
        "amount_total": deposit_amount,
        "fee_amount": fees["fee_amount"],
        "payout_amount": fees["payout_amount"],
        "currency": "MXN",
        "status": TransactionStatus.CREATED,
        "refund_amount": None,
        "refund_reason": None,
        "cancelled_by": None,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": None,
        "paid_at": None
    }
    
    await db.transactions.insert_one(transaction_doc)
    
    origin = request.headers.get('origin', str(request.base_url).rstrip('/'))
    success_url = f"{origin}/payment/success?session_id={{CHECKOUT_SESSION_ID}}&booking_id={checkout_req.booking_id}"
    cancel_url = f"{origin}/payment/cancel?booking_id={checkout_req.booking_id}"
    
    try:
        session = stripe_lib.checkout.Session.create(
            payment_method_types=["card"],
            mode="payment",
            line_items=[{
                "price_data": {
                    "currency": "mxn",
                    "unit_amount": int(deposit_amount * 100),
                    "product_data": {
                        "name": f"Anticipo - {service['name'] if service else 'Reserva'}" if service else "Anticipo de reserva",
                        "description": f"Anticipo para {business['name']}" if business else "Anticipo de reserva"
                    }
                },
                "quantity": 1
            }],
            success_url=success_url,
            cancel_url=cancel_url,
            metadata={
                "transaction_id": transaction_id,
                "booking_id": checkout_req.booking_id,
                "user_id": token_data.user_id,
                "business_id": booking["business_id"],
                "type": "deposit"
            }
        )
    except Exception as e:
        logger.error(f"Stripe checkout error: {e}")
        raise HTTPException(status_code=500, detail=f"Error creating payment session: {str(e)}")
    
    # Update transaction with session ID
    await db.transactions.update_one(
        {"id": transaction_id},
        {"$set": {"stripe_session_id": session.id}}
    )
    
    # Update booking with transaction reference
    await db.bookings.update_one(
        {"id": checkout_req.booking_id},
        {"$set": {"transaction_id": transaction_id, "stripe_session_id": session.id}}
    )
    
    logger.info(f"Created checkout session {session.id} for booking {checkout_req.booking_id}")
    
    return {
        "url": session.url, 
        "session_id": session.id,
        "transaction_id": transaction_id,
        "amount": deposit_amount,
        "fee": fees["fee_amount"]
    }

@payments_router.get("/checkout/status/{session_id}")
async def get_checkout_status(session_id: str, request: Request):
    """Check checkout session status - also acts as webhook fallback to confirm payment"""
    try:
        session = stripe_lib.checkout.Session.retrieve(session_id)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not retrieve session: {str(e)}")
    
    # Get transaction
    transaction = await db.transactions.find_one({"stripe_session_id": session_id})
    
    # Fallback: if Stripe says paid but our DB hasn't been updated (webhook didn't fire)
    if transaction and session.payment_status == "paid" and transaction["status"] != TransactionStatus.PAID:
        now = datetime.now(timezone.utc).isoformat()
        logger.info(f"Checkout status fallback: confirming payment for transaction {transaction['id']} (webhook may not have fired)")
        
        # Update transaction
        await db.transactions.update_one(
            {"id": transaction["id"]},
            {"$set": {
                "status": TransactionStatus.PAID,
                "stripe_payment_intent_id": session.payment_intent if hasattr(session, 'payment_intent') else None,
                "paid_at": now,
                "updated_at": now
            }}
        )
        
        # Create ledger entries
        try:
            await create_transaction_ledger_entries(transaction, TransactionStatus.PAID)
        except Exception as e:
            logger.error(f"Fallback ledger error: {e}")
        
        # Update booking to CONFIRMED
        await db.bookings.update_one(
            {"id": transaction["booking_id"]},
            {"$set": {
                "status": AppointmentStatus.CONFIRMED,
                "deposit_paid": True,
                "confirmed_at": now
            }}
        )
        
        # Update business pending balance
        await db.businesses.update_one(
            {"id": transaction["business_id"]},
            {"$inc": {"pending_balance": transaction["payout_amount"]}}
        )
        
        # Send notifications (best-effort)
        try:
            booking = await db.bookings.find_one({"id": transaction["booking_id"]})
            business = await db.businesses.find_one({"id": transaction["business_id"]})
            user = await db.users.find_one({"id": transaction["user_id"]})
            service = await db.services.find_one({"id": booking["service_id"]}) if booking else None
            
            if user:
                await create_notification(
                    user["id"],
                    "Pago Confirmado",
                    f"Tu anticipo de ${transaction['amount_total']} MXN ha sido confirmado para {service['name'] if service else 'tu cita'}",
                    "system",
                    {"booking_id": transaction["booking_id"], "transaction_id": transaction["id"]}
                )
            if business:
                await create_notification(
                    business["user_id"],
                    "Reserva Confirmada",
                    f"Nueva reserva confirmada - Anticipo recibido",
                    "booking",
                    {"booking_id": transaction["booking_id"]}
                )
            # Send confirmation email to client (fallback)
            if user and booking and service:
                from services.email import send_booking_confirmation
                try:
                    worker_name = ""
                    if booking.get("worker_id"):
                        w = await db.workers.find_one({"id": booking["worker_id"]}, {"_id": 0, "name": 1})
                        worker_name = w["name"] if w else ""
                    await send_booking_confirmation(
                        user_email=user["email"],
                        user_name=user.get("full_name", "Cliente"),
                        business_name=business["name"] if business else "Negocio",
                        service_name=service["name"],
                        date=booking["date"],
                        time=booking["time"],
                        worker_name=worker_name
                    )
                except Exception as e:
                    logger.error(f"Fallback confirmation email error: {e}")
        except Exception as e:
            logger.error(f"Fallback notification error: {e}")
        
        logger.info(f"Fallback: Payment confirmed for booking {transaction['booking_id']}")
        
        # Refresh transaction data after update
        transaction = await db.transactions.find_one({"stripe_session_id": session_id})
    
    return {
        "status": session.status,
        "payment_status": session.payment_status,
        "amount_total": session.amount_total,
        "currency": session.currency,
        "transaction_id": transaction["id"] if transaction else None,
        "booking_id": transaction["booking_id"] if transaction else None
    }

@payments_router.get("/transaction/{transaction_id}", response_model=TransactionResponse)
async def get_transaction(transaction_id: str, token_data: TokenData = Depends(require_auth)):
    """Get transaction details"""
    transaction = await db.transactions.find_one({"id": transaction_id}, {"_id": 0})
    if not transaction:
        raise HTTPException(status_code=404, detail="Transaction not found")
    
    # Verify ownership
    if transaction["user_id"] != token_data.user_id and token_data.role != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Access denied")
    
    return TransactionResponse(**transaction)

# Stripe Webhook (Source of Truth)
@api_router.post("/webhook/stripe")
async def stripe_webhook(request: Request):
    """Handle Stripe webhook events - SOURCE OF TRUTH for payments"""
    body = await request.body()
    signature = request.headers.get("Stripe-Signature", "")
    
    try:
        # Try to verify with webhook secret if available
        webhook_secret = os.environ.get("STRIPE_WEBHOOK_SECRET")
        if webhook_secret:
            event = stripe_lib.Webhook.construct_event(body, signature, webhook_secret)
        else:
            event = stripe_lib.Event.construct_from(
                stripe_lib.util.convert_to_stripe_object(
                    __import__('json').loads(body)
                ),
                stripe_lib.api_key
            )
        
        if event.type == "checkout.session.completed":
            session = event.data.object
            session_id = session.id
            payment_status = "paid" if session.payment_status == "paid" else "unpaid"
            
            logger.info(f"Webhook received: session={session_id}, status={payment_status}")
            
            # Idempotency check
            transaction = await db.transactions.find_one({"stripe_session_id": session_id})
            if not transaction:
                logger.warning(f"No transaction found for session {session_id}")
                return {"status": "ignored", "reason": "no_transaction"}
            
            # Already processed?
            if transaction["status"] == TransactionStatus.PAID:
                logger.info(f"Transaction {transaction['id']} already paid, ignoring duplicate webhook")
                return {"status": "already_processed"}
            
            if payment_status == "paid":
                now = datetime.now(timezone.utc).isoformat()
                
                # Update transaction
                await db.transactions.update_one(
                    {"id": transaction["id"]},
                    {"$set": {
                        "status": TransactionStatus.PAID,
                        "stripe_payment_intent_id": session.payment_intent if hasattr(session, 'payment_intent') else None,
                        "paid_at": now,
                        "updated_at": now
                    }}
                )
                
                # Create ledger entries for payment
                await create_transaction_ledger_entries(transaction, TransactionStatus.PAID)
                
                # Update booking to CONFIRMED
                await db.bookings.update_one(
                    {"id": transaction["booking_id"]},
                    {"$set": {
                        "status": AppointmentStatus.CONFIRMED,
                        "deposit_paid": True,
                        "confirmed_at": now
                    }}
                )
                
                # Get booking details for notification
                booking = await db.bookings.find_one({"id": transaction["booking_id"]})
                business = await db.businesses.find_one({"id": transaction["business_id"]})
                user = await db.users.find_one({"id": transaction["user_id"]})
                service = await db.services.find_one({"id": booking["service_id"]}) if booking else None
                
                # Notify user
                if user:
                    await create_notification(
                        user["id"],
                        "Pago Confirmado",
                        f"Tu anticipo de ${transaction['amount_total']} MXN ha sido confirmado para {service['name'] if service else 'tu cita'}",
                        "system",
                        {"booking_id": transaction["booking_id"], "transaction_id": transaction["id"]}
                    )
                
                # Notify business
                if business:
                    await create_notification(
                        business["user_id"],
                        "Reserva Confirmada",
                        f"Nueva reserva confirmada de {user['full_name'] if user else 'cliente'} - Anticipo recibido",
                        "booking",
                        {"booking_id": transaction["booking_id"]}
                    )
                
                # Notify worker (email + internal notification)
                if booking and booking.get("worker_id"):
                    worker = await db.workers.find_one({"id": booking["worker_id"]})
                    if worker:
                        worker_user = await db.users.find_one({"email": worker.get("email")}) if worker.get("email") else None
                        if worker_user:
                            await create_notification(
                                worker_user["id"],
                                "Nueva cita asignada",
                                f"Se te ha asignado una cita: {service['name'] if service else 'servicio'} el {booking['date']} a las {booking['time']}",
                                "worker_assignment",
                                {"booking_id": booking["id"]}
                            )
                        
                        if worker.get("email"):
                            from services.email import send_worker_assignment
                            try:
                                await send_worker_assignment(
                                    worker_email=worker["email"],
                                    worker_name=worker["name"],
                                    business_name=business["name"] if business else "Bookvia",
                                    service_name=service["name"] if service else "Servicio",
                                    client_name=user["full_name"] if user else "Cliente",
                                    date=booking["date"],
                                    time=booking["time"],
                                    notes=booking.get("notes")
                                )
                                logger.info(f"Worker notification sent to {worker['email']} for booking {booking['id']}")
                            except Exception as e:
                                logger.error(f"Error sending worker notification: {e}")
                
                # Send confirmation email to client
                if user and booking and service:
                    from services.email import send_booking_confirmation
                    try:
                        worker_name = ""
                        if booking.get("worker_id"):
                            w = await db.workers.find_one({"id": booking["worker_id"]}, {"_id": 0, "name": 1})
                            worker_name = w["name"] if w else ""
                        await send_booking_confirmation(
                            user_email=user["email"],
                            user_name=user.get("full_name", "Cliente"),
                            business_name=business["name"] if business else "Negocio",
                            service_name=service["name"],
                            date=booking["date"],
                            time=booking["time"],
                            worker_name=worker_name
                        )
                        logger.info(f"Confirmation email sent to {user['email']} for booking {booking['id']}")
                    except Exception as e:
                        logger.error(f"Error sending confirmation email: {e}")
                
                # Update business balance (pending payout)
                await db.businesses.update_one(
                    {"id": transaction["business_id"]},
                    {"$inc": {"pending_balance": transaction["payout_amount"]}}
                )
                
                logger.info(f"Payment confirmed for booking {transaction['booking_id']}")
        
        return {"status": "success"}
        
    except Exception as e:
        logger.error(f"Webhook error: {e}")
        return {"status": "error", "message": str(e)}

# Cancel booking with refund logic
@bookings_router.put("/{booking_id}/cancel/user")
async def cancel_booking_by_user(
    booking_id: str,
    cancel_req: CancelBookingRequest,
    token_data: TokenData = Depends(require_auth)
):
    """Cancel booking by user - applies cancellation policy"""
    booking = await db.bookings.find_one({"id": booking_id})
    if not booking:
        raise HTTPException(status_code=404, detail="Booking not found")
    
    if booking["user_id"] != token_data.user_id:
        raise HTTPException(status_code=403, detail="Not your booking")
    
    if booking["status"] not in [AppointmentStatus.HOLD, AppointmentStatus.CONFIRMED]:
        raise HTTPException(status_code=400, detail=f"Cannot cancel booking with status {booking['status']}")
    
    now = datetime.now(timezone.utc)
    
    # Calculate hours until appointment
    appointment_dt = datetime.strptime(f"{booking['date']} {booking['time']}", "%Y-%m-%d %H:%M")
    appointment_dt = appointment_dt.replace(tzinfo=timezone.utc)
    hours_until = (appointment_dt - now).total_seconds() / 3600
    
    # Get transaction if exists
    transaction = await db.transactions.find_one({
        "booking_id": booking_id,
        "status": TransactionStatus.PAID
    })
    
    refund_result = None
    
    if transaction:
        # Apply cancellation policy
        if hours_until > 24:
            # >24h: Refund amount - fee (8%)
            refund_amount = transaction["payout_amount"]  # amount - fee
            refund_status = TransactionStatus.REFUND_PARTIAL
            refund_reason = "client_cancel_gt_24h"
            
            logger.info(f"Partial refund: ${refund_amount} for booking {booking_id}")
        else:
            # <24h: No refund (business keeps the deposit minus fee)
            refund_amount = 0
            refund_status = TransactionStatus.REFUND_PARTIAL  # 0 refund, business gets payout
            refund_reason = "client_cancel_lt_24h"
            
            logger.info(f"No refund (<24h): booking {booking_id}")
        
        # Update transaction
        await db.transactions.update_one(
            {"id": transaction["id"]},
            {"$set": {
                "status": refund_status,
                "refund_amount": refund_amount,
                "refund_reason": refund_reason,
                "cancelled_by": "user",
                "updated_at": now.isoformat()
            }}
        )
        
        # Create ledger entries for refund (if >24h)
        if hours_until > 24 and refund_amount > 0:
            updated_tx = {**transaction, "refund_amount": refund_amount}
            await create_transaction_ledger_entries(updated_tx, TransactionStatus.REFUND_PARTIAL)
        
        refund_result = {
            "refund_amount": refund_amount,
            "policy_applied": ">24h partial refund" if hours_until > 24 else "<24h no refund"
        }
    
    # Update booking
    await db.bookings.update_one(
        {"id": booking_id},
        {"$set": {
            "status": AppointmentStatus.CANCELLED,
            "cancelled_at": now.isoformat(),
            "cancelled_by": "user",
            "cancellation_reason": cancel_req.reason
        }}
    )
    
    # Update user stats
    await db.users.update_one(
        {"id": token_data.user_id},
        {"$inc": {"active_appointments_count": -1, "cancellation_count": 1}}
    )
    
    # Check suspension threshold (4 cancellations)
    user = await db.users.find_one({"id": token_data.user_id})
    if user.get("cancellation_count", 0) >= 4:
        suspended_until = (now + timedelta(days=15)).isoformat()
        await db.users.update_one(
            {"id": token_data.user_id},
            {"$set": {"suspended_until": suspended_until}}
        )
        logger.warning(f"User {token_data.user_id} suspended for excessive cancellations")
    
    # Notify business
    business = await db.businesses.find_one({"id": booking["business_id"]})
    if business:
        await create_notification(
            business["user_id"],
            "Reserva Cancelada",
            f"El cliente canceló su cita del {booking['date']} a las {booking['time']}",
            "booking",
            {"booking_id": booking_id}
        )
    
    # Send cancellation email to business owner
    try:
        if business:
            service = await db.services.find_one({"id": booking.get("service_id")})
            from services.email import send_booking_cancelled
            await send_booking_cancelled(
                user_email=business["email"],
                user_name=business["name"],
                business_name=business["name"],
                service_name=service["name"] if service else "Servicio",
                date=booking["date"],
                time=booking["time"],
                reason=f"Cancelada por el cliente. Motivo: {cancel_req.reason or 'No especificado'}"
            )
    except Exception as e:
        logger.error(f"Error sending cancellation email to business: {e}")
    
    return {
        "message": "Booking cancelled",
        "status": AppointmentStatus.CANCELLED,
        "refund": refund_result
    }

@bookings_router.put("/{booking_id}/cancel/business")
async def cancel_booking_by_business(
    booking_id: str,
    cancel_req: CancelBookingRequest,
    token_data: TokenData = Depends(require_business)
):
    """Cancel booking by business - full refund to client, fee charged to business"""
    booking = await db.bookings.find_one({"id": booking_id})
    if not booking:
        raise HTTPException(status_code=404, detail="Booking not found")
    
    # Verify business ownership
    user = await db.users.find_one({"id": token_data.user_id})
    if not user or user.get("business_id") != booking["business_id"]:
        raise HTTPException(status_code=403, detail="Not your business booking")
    
    if booking["status"] not in [AppointmentStatus.HOLD, AppointmentStatus.CONFIRMED]:
        raise HTTPException(status_code=400, detail=f"Cannot cancel booking with status {booking['status']}")
    
    now = datetime.now(timezone.utc)
    
    # Get transaction if exists
    transaction = await db.transactions.find_one({
        "booking_id": booking_id,
        "status": TransactionStatus.PAID
    })
    
    refund_result = None
    
    if transaction:
        # Business cancels: Full refund to client, fee charged to business
        refund_amount = transaction["amount_total"]  # 100% refund
        fee_penalty = transaction["fee_amount"]  # 8% fee charged to business
        
        # Update transaction
        await db.transactions.update_one(
            {"id": transaction["id"]},
            {"$set": {
                "status": TransactionStatus.REFUND_FULL,
                "refund_amount": refund_amount,
                "refund_reason": "business_cancelled",
                "cancelled_by": "business",
                "updated_at": now.isoformat()
            }}
        )
        
        # Create ledger entries for full refund
        updated_tx = {**transaction, "refund_amount": refund_amount}
        await create_transaction_ledger_entries(updated_tx, TransactionStatus.REFUND_FULL)
        
        # Create fee penalty transaction for business
        penalty_tx = {
            "id": generate_id(),
            "booking_id": booking_id,
            "user_id": booking["user_id"],
            "business_id": booking["business_id"],
            "stripe_session_id": None,
            "stripe_payment_intent_id": None,
            "amount_total": fee_penalty,
            "fee_amount": fee_penalty,
            "payout_amount": -fee_penalty,  # Negative = charge to business
            "currency": "MXN",
            "status": TransactionStatus.BUSINESS_CANCEL_FEE,
            "refund_amount": None,
            "refund_reason": "business_cancellation_penalty",
            "cancelled_by": "business",
            "created_at": now.isoformat(),
            "updated_at": now.isoformat(),
            "paid_at": None
        }
        await db.transactions.insert_one(penalty_tx)
        
        # Create ledger entries for penalty
        await create_transaction_ledger_entries(penalty_tx, TransactionStatus.BUSINESS_CANCEL_FEE)
        
        # Update business balance (deduct pending and add penalty)
        await db.businesses.update_one(
            {"id": booking["business_id"]},
            {"$inc": {
                "pending_balance": -transaction["payout_amount"],
                "penalty_balance": fee_penalty
            }}
        )
        
        refund_result = {
            "refund_amount": refund_amount,
            "fee_penalty": fee_penalty,
            "policy_applied": "business_cancel_full_refund_fee_penalty"
        }
        
        logger.info(f"Business cancelled booking {booking_id}: full refund ${refund_amount}, penalty ${fee_penalty}")
    
    # Update booking
    await db.bookings.update_one(
        {"id": booking_id},
        {"$set": {
            "status": AppointmentStatus.CANCELLED,
            "cancelled_at": now.isoformat(),
            "cancelled_by": "business",
            "cancellation_reason": cancel_req.reason
        }}
    )
    
    # Update user stats (don't penalize user for business cancellation)
    await db.users.update_one(
        {"id": booking["user_id"]},
        {"$inc": {"active_appointments_count": -1}}
    )
    
    # Notify user
    await create_notification(
        booking["user_id"],
        "Reserva Cancelada por el Negocio",
        f"Tu cita del {booking['date']} a las {booking['time']} fue cancelada. Se procesará un reembolso completo.",
        "system",
        {"booking_id": booking_id}
    )
    
    # Send cancellation email to client
    try:
        client_user = await db.users.find_one({"id": booking["user_id"]})
        business = await db.businesses.find_one({"id": booking["business_id"]})
        service = await db.services.find_one({"id": booking.get("service_id")})
        if client_user and business:
            from services.email import send_booking_cancelled
            refund_msg = "Se procesará un reembolso completo a tu método de pago." if refund_result else None
            await send_booking_cancelled(
                user_email=client_user["email"],
                user_name=client_user.get("full_name", "Cliente"),
                business_name=business["name"],
                service_name=service["name"] if service else "Servicio",
                date=booking["date"],
                time=booking["time"],
                reason=f"Cancelada por el negocio. Motivo: {cancel_req.reason or 'No especificado'}",
                refund_info=refund_msg
            )
    except Exception as e:
        logger.error(f"Error sending cancellation email to client: {e}")
    
    # Log activity
    await create_business_activity(
        booking["business_id"], token_data, "cancel_booking", "booking", booking_id,
        {"client_name": booking.get("client_name", ""), "service_name": booking.get("service_name", ""), "date": booking.get("date", ""), "time": booking.get("time", ""), "reason": cancel_req.reason}
    )
    
    return {
        "message": "Booking cancelled by business",
        "status": AppointmentStatus.CANCELLED,
        "refund": refund_result
    }

@bookings_router.put("/{booking_id}/no-show")
async def mark_no_show(booking_id: str, token_data: TokenData = Depends(require_business)):
    """Mark booking as no-show - business keeps deposit minus fee"""
    booking = await db.bookings.find_one({"id": booking_id})
    if not booking:
        raise HTTPException(status_code=404, detail="Booking not found")
    
    # Verify business ownership
    user = await db.users.find_one({"id": token_data.user_id})
    if not user or user.get("business_id") != booking["business_id"]:
        raise HTTPException(status_code=403, detail="Not your business booking")
    
    if booking["status"] != AppointmentStatus.CONFIRMED:
        raise HTTPException(status_code=400, detail="Can only mark confirmed bookings as no-show")
    
    now = datetime.now(timezone.utc)
    
    # Get transaction
    transaction = await db.transactions.find_one({
        "booking_id": booking_id,
        "status": TransactionStatus.PAID
    })
    
    payout_result = None
    
    if transaction:
        # No-show: Business receives payout (deposit - fee)
        await db.transactions.update_one(
            {"id": transaction["id"]},
            {"$set": {
                "status": TransactionStatus.NO_SHOW_PAYOUT,
                "updated_at": now.isoformat()
            }}
        )
        
        payout_result = {
            "payout_amount": transaction["payout_amount"],
            "fee_retained": transaction["fee_amount"]
        }
        
        logger.info(f"No-show for booking {booking_id}: business payout ${transaction['payout_amount']}")
    
    # Update booking
    await db.bookings.update_one(
        {"id": booking_id},
        {"$set": {"status": AppointmentStatus.NO_SHOW}}
    )
    
    # Update user stats (no-show counts as cancellation)
    await db.users.update_one(
        {"id": booking["user_id"]},
        {"$inc": {"active_appointments_count": -1, "cancellation_count": 1, "no_show_count": 1}}
    )
    
    # Check suspension threshold
    booking_user = await db.users.find_one({"id": booking["user_id"]})
    if booking_user.get("cancellation_count", 0) >= 4:
        suspended_until = (now + timedelta(days=15)).isoformat()
        await db.users.update_one(
            {"id": booking["user_id"]},
            {"$set": {"suspended_until": suspended_until}}
        )
    
    return {
        "message": "Marked as no-show",
        "payout": payout_result
    }

# Background task: Expire holds
async def expire_holds_task():
    """Background task to expire holds that weren't paid"""
    now = datetime.now(timezone.utc)
    
    expired_bookings = await db.bookings.find({
        "status": AppointmentStatus.HOLD,
        "hold_expires_at": {"$lt": now.isoformat()}
    }).to_list(100)
    
    for booking in expired_bookings:
        await db.bookings.update_one(
            {"id": booking["id"]},
            {"$set": {"status": AppointmentStatus.EXPIRED}}
        )
        
        await db.users.update_one(
            {"id": booking["user_id"]},
            {"$inc": {"active_appointments_count": -1}}
        )
        
        # Update transaction if exists
        await db.transactions.update_one(
            {"booking_id": booking["id"], "status": TransactionStatus.CREATED},
            {"$set": {"status": TransactionStatus.EXPIRED, "updated_at": now.isoformat()}}
        )
        
        logger.info(f"Expired hold for booking {booking['id']}")
    
    return len(expired_bookings)

# Manual endpoint to trigger expiration (for testing/cron)
@payments_router.post("/expire-holds")
async def trigger_expire_holds(token_data: TokenData = Depends(require_admin)):
    """Admin endpoint to manually trigger hold expiration"""
    count = await expire_holds_task()
    return {"expired_count": count}

# Get user transactions
@payments_router.get("/my-transactions", response_model=List[TransactionResponse])
async def get_my_transactions(
    status: Optional[str] = None,
    token_data: TokenData = Depends(require_auth)
):
    """Get user's payment transactions"""
    filters = {"user_id": token_data.user_id}
    if status:
        filters["status"] = status
    
    transactions = await db.transactions.find(filters, {"_id": 0}).sort("created_at", -1).to_list(100)
    return [TransactionResponse(**t) for t in transactions]

# Get business transactions
@payments_router.get("/business-transactions", response_model=List[TransactionResponse])
async def get_business_transactions(
    status: Optional[str] = None,
    token_data: TokenData = Depends(require_business)
):
    """Get business payment transactions"""
    user = await db.users.find_one({"id": token_data.user_id})
    if not user or not user.get("business_id"):
        raise HTTPException(status_code=404, detail="Business not found")
    
    filters = {"business_id": user["business_id"]}
    if status:
        filters["status"] = status
    
    transactions = await db.transactions.find(filters, {"_id": 0}).sort("created_at", -1).to_list(100)
    return [TransactionResponse(**t) for t in transactions]

# Legacy checkout endpoint (for subscriptions)
@payments_router.post("/checkout/session")
async def create_checkout_session(request: Request, payment: PaymentCreate, token_data: TokenData = Depends(require_auth)):
    origin = request.headers.get('origin', str(request.base_url).rstrip('/'))
    success_url = f"{origin}/payment/success?session_id={{CHECKOUT_SESSION_ID}}"
    cancel_url = f"{origin}/payment/cancel"
    
    try:
        session = stripe_lib.checkout.Session.create(
            payment_method_types=["card"],
            mode="payment",
            line_items=[{
                "price_data": {
                    "currency": payment.currency.lower(),
                    "unit_amount": int(float(payment.amount) * 100),
                    "product_data": {
                        "name": "Pago Bookvia"
                    }
                },
                "quantity": 1
            }],
            success_url=success_url,
            cancel_url=cancel_url,
            metadata={
                "user_id": token_data.user_id,
                "booking_id": payment.booking_id or "",
                "subscription_type": payment.subscription_type or ""
            }
        )
        return {"url": session.url, "session_id": session.id}
    except Exception as e:
        logger.error(f"Stripe checkout error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# ========================== NOTIFICATION ROUTES ==========================

@notifications_router.get("", response_model=List[NotificationResponse])
async def get_notifications(
    unread_only: bool = False,
    token_data: TokenData = Depends(require_auth)
):
    filters = {"user_id": token_data.user_id}
    if unread_only:
        filters["read"] = False
    
    notifications = await db.notifications.find(filters, {"_id": 0}).sort("created_at", -1).limit(50).to_list(50)
    return [NotificationResponse(**n) for n in notifications]

@notifications_router.put("/{notification_id}/read")
async def mark_notification_read(notification_id: str, token_data: TokenData = Depends(require_auth)):
    await db.notifications.update_one(
        {"id": notification_id, "user_id": token_data.user_id},
        {"$set": {"read": True}}
    )
    return {"message": "Marked as read"}

@notifications_router.put("/read-all")
async def mark_all_read(token_data: TokenData = Depends(require_auth)):
    await db.notifications.update_many(
        {"user_id": token_data.user_id},
        {"$set": {"read": True}}
    )
    return {"message": "All marked as read"}

@notifications_router.get("/unread-count")
async def get_unread_count(token_data: TokenData = Depends(require_auth)):
    count = await db.notifications.count_documents({"user_id": token_data.user_id, "read": False})
    return {"count": count}


# ========================== BUSINESS FINANCE ROUTES ==========================

@finance_router.get("/summary", response_model=BusinessFinanceSummary)
async def get_finance_summary(token_data: TokenData = Depends(require_business)):
    """Get financial summary for the business"""
    user = await db.users.find_one({"id": token_data.user_id})
    if not user or not user.get("business_id"):
        raise HTTPException(status_code=404, detail="Business not found")
    
    business_id = user["business_id"]
    
    # Calculate from ledger
    summary = await calculate_business_ledger_summary(business_id)
    
    # Get settlement totals
    paid_settlements = await db.settlements.find({
        "business_id": business_id,
        "status": SettlementStatus.PAID
    }).to_list(1000)
    paid_payout = sum(s.get("net_payout", 0) for s in paid_settlements)
    
    held_settlements = await db.settlements.find({
        "business_id": business_id,
        "status": SettlementStatus.HELD
    }).to_list(100)
    held_payout = sum(s.get("net_payout", 0) for s in held_settlements)
    
    # Calculate next settlement date (1st of next month)
    now = datetime.now(timezone.utc)
    if now.month == 12:
        next_settlement = datetime(now.year + 1, 1, 1, tzinfo=timezone.utc)
    else:
        next_settlement = datetime(now.year, now.month + 1, 1, tzinfo=timezone.utc)
    
    return BusinessFinanceSummary(
        gross_revenue=summary["gross_revenue"],
        total_fees=summary["total_fees"],
        total_refunds=summary["total_refunds"],
        total_penalties=summary["total_penalties"],
        net_earnings=summary["net_earnings"],
        pending_payout=summary["pending_payout"],
        paid_payout=round(paid_payout, 2),
        held_payout=round(held_payout, 2),
        next_settlement_date=next_settlement.isoformat()
    )

@finance_router.get("/transactions", response_model=List[TransactionResponse])
async def get_finance_transactions(
    status: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    page: int = 1,
    limit: int = 50,
    token_data: TokenData = Depends(require_business)
):
    """Get business transactions with filters"""
    user = await db.users.find_one({"id": token_data.user_id})
    if not user or not user.get("business_id"):
        raise HTTPException(status_code=404, detail="Business not found")
    
    filters = {"business_id": user["business_id"]}
    
    if status:
        filters["status"] = status
    if start_date and end_date:
        filters["created_at"] = {"$gte": start_date, "$lte": end_date}
    
    skip = (page - 1) * limit
    transactions = await db.transactions.find(filters, {"_id": 0}).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    
    return [TransactionResponse(**t) for t in transactions]

@finance_router.get("/ledger", response_model=List[LedgerEntryResponse])
async def get_finance_ledger(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    account: Optional[str] = None,
    page: int = 1,
    limit: int = 100,
    token_data: TokenData = Depends(require_business)
):
    """Get ledger entries for the business"""
    user = await db.users.find_one({"id": token_data.user_id})
    if not user or not user.get("business_id"):
        raise HTTPException(status_code=404, detail="Business not found")
    
    filters = {"business_id": user["business_id"]}
    
    if account:
        filters["account"] = account
    if start_date and end_date:
        filters["created_at"] = {"$gte": start_date, "$lte": end_date}
    
    skip = (page - 1) * limit
    entries = await db.ledger_entries.find(filters, {"_id": 0}).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    
    return [LedgerEntryResponse(**e) for e in entries]

@finance_router.get("/settlements", response_model=List[SettlementResponse])
async def get_finance_settlements(
    status: Optional[str] = None,
    token_data: TokenData = Depends(require_business)
):
    """Get settlement history for the business"""
    user = await db.users.find_one({"id": token_data.user_id})
    if not user or not user.get("business_id"):
        raise HTTPException(status_code=404, detail="Business not found")
    
    filters = {"business_id": user["business_id"]}
    if status:
        filters["status"] = status
    
    settlements = await db.settlements.find(filters, {"_id": 0}).sort("period_start", -1).to_list(100)
    
    # Add business name
    business = await db.businesses.find_one({"id": user["business_id"]})
    for s in settlements:
        s["business_name"] = business["name"] if business else None
    
    return [SettlementResponse(**s) for s in settlements]

# ========================== SETTLEMENT GENERATION ==========================

async def generate_monthly_settlements(
    year: int, 
    month: int, 
    idempotency_key: str,
    admin_id: str = None,
    request: Request = None
) -> dict:
    """Generate monthly settlements for all businesses - IDEMPOTENT"""
    
    period_start = datetime(year, month, 1, tzinfo=timezone.utc)
    if month == 12:
        period_end = datetime(year + 1, 1, 1, tzinfo=timezone.utc) - timedelta(seconds=1)
    else:
        period_end = datetime(year, month + 1, 1, tzinfo=timezone.utc) - timedelta(seconds=1)
    
    period_key = f"MX-{year}-{str(month).zfill(2)}"
    
    # Get all approved businesses
    businesses = await db.businesses.find({"status": BusinessStatus.APPROVED}).to_list(1000)
    
    results = {
        "period_key": period_key,
        "processed": 0,
        "skipped": 0,
        "held": 0,
        "errors": []
    }
    
    for business in businesses:
        business_id = business["id"]
        
        # Idempotency check - skip if settlement already exists for this period
        existing = await db.settlements.find_one({
            "business_id": business_id,
            "period_key": period_key
        })
        if existing:
            results["skipped"] += 1
            continue
        
        try:
            # Calculate from ledger for this period
            summary = await calculate_business_ledger_summary(
                business_id,
                period_start.isoformat(),
                period_end.isoformat()
            )
            
            # Check if business is on payout hold
            is_held = business.get("payout_hold", False)
            held_reason = business.get("payout_hold_reason") if is_held else None
            
            settlement = {
                "id": generate_id(),
                "business_id": business_id,
                "period_key": period_key,
                "period_start": period_start.isoformat(),
                "period_end": period_end.isoformat(),
                "gross_paid": summary["gross_revenue"],
                "total_fees": summary["total_fees"],
                "total_refunds": summary["total_refunds"],
                "total_penalties": summary["total_penalties"],
                "net_payout": summary["pending_payout"],
                "currency": "MXN",
                "country": "MX",
                "status": SettlementStatus.HELD if is_held else SettlementStatus.PENDING,
                "held_reason": held_reason,
                "idempotency_key": idempotency_key,
                "payout_reference": None,
                "paid_at": None,
                "created_at": datetime.now(timezone.utc).isoformat(),
                "updated_at": None
            }
            
            await db.settlements.insert_one(settlement)
            
            if is_held:
                results["held"] += 1
            else:
                results["processed"] += 1
            
            logger.info(f"Settlement created for business {business_id}: ${summary['pending_payout']} ({settlement['status']})")
            
        except Exception as e:
            results["errors"].append({"business_id": business_id, "error": str(e)})
            logger.error(f"Error creating settlement for {business_id}: {e}")
    
    # Create audit log
    if admin_id:
        await create_audit_log(
            admin_id=admin_id,
            admin_email="system",
            action=AuditAction.SETTLEMENT_GENERATE,
            target_type="settlement",
            target_id=period_key,
            details=results,
            request=request
        )
    
    return results

# ========================== ADMIN SETTLEMENTS ==========================

@admin_router.post("/settlements/generate")
async def admin_generate_settlements(
    request: Request,
    year: int,
    month: int,
    token_data: TokenData = Depends(require_admin)
):
    """Admin endpoint to generate monthly settlements"""
    # Generate idempotency key based on job run
    idempotency_key = f"job-{year}-{month}-{datetime.now(timezone.utc).strftime('%Y%m%d%H%M')}"
    
    results = await generate_monthly_settlements(
        year, month, idempotency_key,
        admin_id=token_data.user_id,
        request=request
    )
    
    return results

@admin_router.get("/settlements", response_model=List[SettlementResponse])
async def admin_get_settlements(
    status: Optional[str] = None,
    period_key: Optional[str] = None,
    business_id: Optional[str] = None,
    page: int = 1,
    limit: int = 50,
    token_data: TokenData = Depends(require_admin)
):
    """Admin: Get all settlements with filters"""
    filters = {}
    if status:
        filters["status"] = status
    if period_key:
        filters["period_key"] = period_key
    if business_id:
        filters["business_id"] = business_id
    
    skip = (page - 1) * limit
    settlements = await db.settlements.find(filters, {"_id": 0}).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    
    # Add business names
    for s in settlements:
        business = await db.businesses.find_one({"id": s["business_id"]})
        s["business_name"] = business["name"] if business else None
    
    return [SettlementResponse(**s) for s in settlements]

@admin_router.put("/settlements/{settlement_id}/pay")
async def admin_mark_settlement_paid(
    settlement_id: str,
    request: Request,
    pay_req: SettlementMarkPaidRequest,
    token_data: TokenData = Depends(require_admin)
):
    """Admin: Mark settlement as paid (manual)"""
    settlement = await db.settlements.find_one({"id": settlement_id})
    if not settlement:
        raise HTTPException(status_code=404, detail="Settlement not found")
    
    if settlement["status"] not in [SettlementStatus.PENDING, SettlementStatus.HELD]:
        raise HTTPException(status_code=400, detail=f"Cannot mark settlement with status {settlement['status']} as paid")
    
    now = datetime.now(timezone.utc)
    
    await db.settlements.update_one(
        {"id": settlement_id},
        {"$set": {
            "status": SettlementStatus.PAID,
            "payout_reference": pay_req.payout_reference,
            "paid_at": now.isoformat(),
            "updated_at": now.isoformat()
        }}
    )
    
    # Create payout ledger entry
    await create_ledger_entry(
        transaction_id=settlement_id,
        business_id=settlement["business_id"],
        direction=LedgerDirection.CREDIT,
        account=LedgerAccount.PAYOUT,
        amount=settlement["net_payout"],
        description=f"Payout {settlement['period_key']} - {pay_req.payout_reference}",
        created_by="admin"
    )
    
    # Create audit log
    await create_audit_log(
        admin_id=token_data.user_id,
        admin_email=token_data.email,
        action=AuditAction.SETTLEMENT_MARK_PAID,
        target_type="settlement",
        target_id=settlement_id,
        details={
            "business_id": settlement["business_id"],
            "net_payout": settlement["net_payout"],
            "payout_reference": pay_req.payout_reference
        },
        request=request
    )
    
    return {"message": "Settlement marked as paid", "payout_reference": pay_req.payout_reference}

@admin_router.put("/businesses/{business_id}/payout-hold")
async def admin_toggle_payout_hold(
    business_id: str,
    request: Request,
    hold_req: PayoutHoldRequest,
    token_data: TokenData = Depends(require_admin)
):
    """Admin: Toggle payout hold for a business"""
    business = await db.businesses.find_one({"id": business_id})
    if not business:
        raise HTTPException(status_code=404, detail="Business not found")
    
    await db.businesses.update_one(
        {"id": business_id},
        {"$set": {
            "payout_hold": hold_req.hold,
            "payout_hold_reason": hold_req.reason if hold_req.hold else None
        }}
    )
    
    # If releasing hold, update any HELD settlements to PENDING
    if not hold_req.hold:
        await db.settlements.update_many(
            {"business_id": business_id, "status": SettlementStatus.HELD},
            {"$set": {"status": SettlementStatus.PENDING, "held_reason": None}}
        )
    
    # Create audit log
    await create_audit_log(
        admin_id=token_data.user_id,
        admin_email=token_data.email,
        action=AuditAction.PAYOUT_HOLD if hold_req.hold else AuditAction.PAYOUT_RELEASE,
        target_type="business",
        target_id=business_id,
        details={"hold": hold_req.hold, "reason": hold_req.reason},
        request=request
    )
    
    return {"message": f"Payout {'held' if hold_req.hold else 'released'} for business {business_id}"}

# ========================== ADMIN EXPORT (CSV) ==========================

@admin_router.get("/export/transactions")
async def admin_export_transactions(
    year: int,
    month: int,
    token_data: TokenData = Depends(require_admin)
):
    """Admin: Export transactions as CSV"""
    import csv
    import io
    from fastapi.responses import StreamingResponse
    
    period_start = datetime(year, month, 1, tzinfo=timezone.utc)
    if month == 12:
        period_end = datetime(year + 1, 1, 1, tzinfo=timezone.utc)
    else:
        period_end = datetime(year, month + 1, 1, tzinfo=timezone.utc)
    
    transactions = await db.transactions.find({
        "created_at": {"$gte": period_start.isoformat(), "$lt": period_end.isoformat()}
    }, {"_id": 0}).to_list(10000)
    
    # Create CSV
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Header
    writer.writerow([
        "ID", "Booking ID", "User ID", "Business ID",
        "Amount Total", "Fee Amount", "Payout Amount", "Currency",
        "Status", "Refund Amount", "Refund Reason", "Cancelled By",
        "Created At", "Paid At"
    ])
    
    # Data
    for t in transactions:
        writer.writerow([
            t.get("id"), t.get("booking_id"), t.get("user_id"), t.get("business_id"),
            t.get("amount_total"), t.get("fee_amount"), t.get("payout_amount"), t.get("currency"),
            t.get("status"), t.get("refund_amount"), t.get("refund_reason"), t.get("cancelled_by"),
            t.get("created_at"), t.get("paid_at")
        ])
    
    output.seek(0)
    
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=transactions_{year}_{month:02d}.csv"}
    )

@admin_router.get("/export/settlements")
async def admin_export_settlements(
    year: int,
    month: int,
    token_data: TokenData = Depends(require_admin)
):
    """Admin: Export settlements as CSV"""
    import csv
    import io
    from fastapi.responses import StreamingResponse
    
    period_key = f"MX-{year}-{str(month).zfill(2)}"
    
    settlements = await db.settlements.find({
        "period_key": period_key
    }, {"_id": 0}).to_list(10000)
    
    # Add business names
    for s in settlements:
        business = await db.businesses.find_one({"id": s["business_id"]})
        s["business_name"] = business["name"] if business else "Unknown"
    
    # Create CSV
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Header
    writer.writerow([
        "ID", "Business ID", "Business Name", "Period Key",
        "Gross Paid", "Total Fees", "Total Refunds", "Total Penalties", "Net Payout",
        "Currency", "Status", "Held Reason", "Payout Reference", "Paid At", "Created At"
    ])
    
    # Data
    for s in settlements:
        writer.writerow([
            s.get("id"), s.get("business_id"), s.get("business_name"), s.get("period_key"),
            s.get("gross_paid"), s.get("total_fees"), s.get("total_refunds"), s.get("total_penalties"), s.get("net_payout"),
            s.get("currency"), s.get("status"), s.get("held_reason"), s.get("payout_reference"), s.get("paid_at"), s.get("created_at")
        ])
    
    output.seek(0)
    
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=settlements_{year}_{month:02d}.csv"}
    )

# ========================== ADMIN ROUTES ==========================

@admin_router.get("/stats")
async def get_admin_stats(token_data: TokenData = Depends(require_admin)):
    total_users = await db.users.count_documents({"role": UserRole.USER})
    total_businesses = await db.businesses.count_documents({})
    approved_businesses = await db.businesses.count_documents({"status": BusinessStatus.APPROVED})
    pending_businesses = await db.businesses.count_documents({"status": BusinessStatus.PENDING})
    
    total_bookings = await db.bookings.count_documents({})
    completed_bookings = await db.bookings.count_documents({"status": AppointmentStatus.COMPLETED})
    
    # This month stats
    first_of_month = datetime.now(timezone.utc).replace(day=1).strftime("%Y-%m-%d")
    month_bookings = await db.bookings.count_documents({"date": {"$gte": first_of_month}})
    month_revenue = 0
    
    payments = await db.payment_transactions.find({
        "status": PaymentStatus.COMPLETED,
        "created_at": {"$gte": first_of_month}
    }).to_list(10000)
    month_revenue = sum(p.get("amount", 0) for p in payments)
    
    return {
        "users": {
            "total": total_users
        },
        "businesses": {
            "total": total_businesses,
            "approved": approved_businesses,
            "pending": pending_businesses
        },
        "bookings": {
            "total": total_bookings,
            "completed": completed_bookings,
            "this_month": month_bookings
        },
        "revenue": {
            "this_month": month_revenue
        }
    }

@admin_router.get("/businesses/pending", response_model=List[BusinessResponse])
async def get_pending_businesses(token_data: TokenData = Depends(require_admin)):
    # Only show businesses whose owner has verified their email
    businesses = await db.businesses.find(
        {"status": BusinessStatus.PENDING},
        {"_id": 0, "password_hash": 0}
    ).to_list(100)
    
    verified_businesses = []
    for b in businesses:
        user = await db.users.find_one({"business_id": b["id"]}, {"_id": 0, "email_verified": 1})
        if user and user.get("email_verified", False):
            verified_businesses.append(b)
    
    return [BusinessResponse(**b) for b in verified_businesses]

@admin_router.put("/businesses/{business_id}/approve")
async def approve_business(business_id: str, request: Request, token_data: TokenData = Depends(require_admin)):
    business = await db.businesses.find_one({"id": business_id})
    if not business:
        raise HTTPException(status_code=404, detail="Business not found")
    
    result = await db.businesses.update_one(
        {"id": business_id},
        {"$set": {"status": BusinessStatus.APPROVED}}
    )
    
    # Create audit log
    await create_audit_log(
        admin_id=token_data.user_id,
        admin_email=token_data.email,
        action=AuditAction.BUSINESS_APPROVE,
        target_type="business",
        target_id=business_id,
        details={"business_name": business.get("name")},
        request=request
    )
    
    # Notify business owner
    if business.get("user_id"):
        await create_notification(
            business["user_id"],
            "Negocio Aprobado",
            f"¡Tu negocio {business['name']} ha sido aprobado! Ya puedes recibir reservas.",
            "system",
            {"business_id": business_id}
        )
    
    return {"message": "Business approved"}

@admin_router.put("/businesses/{business_id}/reject")
async def reject_business(business_id: str, request: Request, reason: str = "", token_data: TokenData = Depends(require_admin)):
    business = await db.businesses.find_one({"id": business_id})
    if not business:
        raise HTTPException(status_code=404, detail="Business not found")
    
    result = await db.businesses.update_one(
        {"id": business_id},
        {"$set": {"status": BusinessStatus.REJECTED, "rejection_reason": reason}}
    )
    
    # Create audit log
    await create_audit_log(
        admin_id=token_data.user_id,
        admin_email=token_data.email,
        action=AuditAction.BUSINESS_REJECT,
        target_type="business",
        target_id=business_id,
        details={"business_name": business.get("name"), "reason": reason},
        request=request
    )
    
    # Notify business owner
    if business.get("user_id"):
        await create_notification(
            business["user_id"],
            "Negocio Rechazado",
            f"Tu solicitud de negocio ha sido rechazada. Razón: {reason or 'No especificada'}",
            "system",
            {"business_id": business_id}
        )
    
    return {"message": "Business rejected"}

@admin_router.put("/businesses/{business_id}/suspend")
async def suspend_business(business_id: str, request: Request, reason: str = "", token_data: TokenData = Depends(require_admin)):
    business = await db.businesses.find_one({"id": business_id})
    if not business:
        raise HTTPException(status_code=404, detail="Business not found")
    
    result = await db.businesses.update_one(
        {"id": business_id},
        {"$set": {"status": BusinessStatus.SUSPENDED, "suspension_reason": reason}}
    )
    
    # Create audit log
    await create_audit_log(
        admin_id=token_data.user_id,
        admin_email=token_data.email,
        action=AuditAction.BUSINESS_SUSPEND,
        target_type="business",
        target_id=business_id,
        details={"business_name": business.get("name"), "reason": reason},
        request=request
    )
    
    # Notify business owner
    if business.get("user_id"):
        await create_notification(
            business["user_id"],
            "Negocio Suspendido",
            f"Tu negocio ha sido suspendido. Razón: {reason or 'No especificada'}",
            "system",
            {"business_id": business_id}
        )
    
    return {"message": "Business suspended"}

@admin_router.put("/users/{user_id}/suspend")
async def suspend_user(user_id: str, request: Request, days: int = 15, reason: str = "", token_data: TokenData = Depends(require_admin)):
    user = await db.users.find_one({"id": user_id})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    suspended_until = (datetime.now(timezone.utc) + timedelta(days=days)).isoformat()
    
    result = await db.users.update_one(
        {"id": user_id},
        {"$set": {"suspended_until": suspended_until, "suspension_reason": reason}}
    )
    
    # Create audit log
    await create_audit_log(
        admin_id=token_data.user_id,
        admin_email=token_data.email,
        action=AuditAction.USER_SUSPEND,
        target_type="user",
        target_id=user_id,
        details={"user_email": user.get("email"), "days": days, "reason": reason},
        request=request
    )
    
    return {"message": f"User suspended for {days} days"}

@admin_router.delete("/reviews/{review_id}")
async def delete_review(review_id: str, request: Request, reason: str = "", token_data: TokenData = Depends(require_admin)):
    review = await db.reviews.find_one({"id": review_id})
    if not review:
        raise HTTPException(status_code=404, detail="Review not found")
    
    # Update business rating
    business = await db.businesses.find_one({"id": review["business_id"]})
    if business:
        new_rating_sum = business.get("rating_sum", 0) - review["rating"]
        new_review_count = max(business.get("review_count", 0) - 1, 0)
        new_rating = calculate_bayesian_rating(new_rating_sum, new_review_count) if new_review_count > 0 else 0
        
        await db.businesses.update_one(
            {"id": review["business_id"]},
            {"$set": {"rating": round(new_rating, 2), "rating_sum": new_rating_sum, "review_count": new_review_count}}
        )
    
    await db.reviews.delete_one({"id": review_id})
    
    # Create audit log
    await create_audit_log(
        admin_id=token_data.user_id,
        admin_email=token_data.email,
        action=AuditAction.REVIEW_DELETE,
        target_type="review",
        target_id=review_id,
        details={
            "business_id": review.get("business_id"),
            "user_id": review.get("user_id"),
            "rating": review.get("rating"),
            "reason": reason
        },
        request=request
    )
    
    return {"message": "Review deleted"}

@admin_router.get("/audit-logs", response_model=List[AuditLogResponse])
async def get_audit_logs(
    page: int = 1,
    limit: int = 50,
    action: Optional[str] = None,
    admin_id: Optional[str] = None,
    token_data: TokenData = Depends(require_admin)
):
    filters = {}
    if action:
        filters["action"] = action
    if admin_id:
        filters["admin_id"] = admin_id
    
    skip = (page - 1) * limit
    logs = await db.audit_logs.find(filters, {"_id": 0}).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    return [AuditLogResponse(**log) for log in logs]

@admin_router.put("/businesses/{business_id}/feature")
async def toggle_featured(business_id: str, request: Request, featured: bool = True, token_data: TokenData = Depends(require_admin)):
    business = await db.businesses.find_one({"id": business_id})
    if not business:
        raise HTTPException(status_code=404, detail="Business not found")
    
    await db.businesses.update_one(
        {"id": business_id},
        {"$set": {"is_featured": featured}}
    )
    
    # Create audit log
    await create_audit_log(
        admin_id=token_data.user_id,
        admin_email=token_data.email,
        action=AuditAction.BUSINESS_FEATURE,
        target_type="business",
        target_id=business_id,
        details={"business_name": business.get("name"), "featured": featured},
        request=request
    )
    
    return {"message": f"Business {'featured' if featured else 'unfeatured'}"}

# Admin: View sent emails
@admin_router.get("/emails")
async def get_sent_emails(
    limit: int = 50,
    status: Optional[str] = None,
    to: Optional[str] = None,
    token_data: TokenData = Depends(require_admin)
):
    """Admin: Get sent emails from the system"""
    from services.email import get_sent_emails as fetch_emails
    emails = await fetch_emails(limit=limit, status=status, to=to)
    return emails

# Payment hold/release endpoints for admin
@admin_router.put("/payments/{payment_id}/hold")
async def hold_payment(payment_id: str, request: Request, reason: str = "", token_data: TokenData = Depends(require_admin)):
    """Put a payment on hold - prevents payout to business"""
    payment = await db.payment_transactions.find_one({"id": payment_id})
    if not payment:
        raise HTTPException(status_code=404, detail="Payment not found")
    
    await db.payment_transactions.update_one(
        {"id": payment_id},
        {"$set": {"on_hold": True, "hold_reason": reason, "held_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    # Create audit log
    await create_audit_log(
        admin_id=token_data.user_id,
        admin_email=token_data.email,
        action=AuditAction.PAYMENT_HOLD,
        target_type="payment",
        target_id=payment_id,
        details={"amount": payment.get("amount"), "reason": reason},
        request=request
    )
    
    return {"message": "Payment put on hold"}

@admin_router.put("/payments/{payment_id}/release")
async def release_payment(payment_id: str, request: Request, token_data: TokenData = Depends(require_admin)):
    """Release a payment from hold"""
    payment = await db.payment_transactions.find_one({"id": payment_id})
    if not payment:
        raise HTTPException(status_code=404, detail="Payment not found")
    
    await db.payment_transactions.update_one(
        {"id": payment_id},
        {"$set": {"on_hold": False, "released_at": datetime.now(timezone.utc).isoformat()},
         "$unset": {"hold_reason": ""}}
    )
    
    # Create audit log
    await create_audit_log(
        admin_id=token_data.user_id,
        admin_email=token_data.email,
        action=AuditAction.PAYMENT_RELEASE,
        target_type="payment",
        target_id=payment_id,
        details={"amount": payment.get("amount")},
        request=request
    )
    
    return {"message": "Payment released"}

@admin_router.get("/payments/held")
async def get_held_payments(page: int = 1, limit: int = 50, token_data: TokenData = Depends(require_admin)):
    """Get all payments currently on hold"""
    skip = (page - 1) * limit
    payments = await db.payment_transactions.find(
        {"on_hold": True},
        {"_id": 0}
    ).sort("held_at", -1).skip(skip).limit(limit).to_list(limit)
    return payments

# ========================== ADMIN CREATION ==========================

async def ensure_admin_exists():
    """Create admin user if configured in environment and doesn't exist"""
    if not ADMIN_EMAIL or not ADMIN_INITIAL_PASSWORD:
        logger.info("ADMIN_EMAIL or ADMIN_INITIAL_PASSWORD not set - skipping admin creation")
        return False
    
    # Check if admin already exists
    existing_admin = await db.users.find_one({"email": ADMIN_EMAIL})
    if existing_admin:
        logger.info(f"Admin user already exists: {ADMIN_EMAIL}")
        return True
    
    # Create admin
    admin_doc = {
        "id": generate_id(),
        "email": ADMIN_EMAIL,
        "password_hash": hash_password(ADMIN_INITIAL_PASSWORD),
        "full_name": "Admin Bookvia",
        "phone": "+521234567890",
        "phone_verified": True,
        "role": UserRole.ADMIN,
        "totp_enabled": False,
        "totp_secret": None,
        "backup_codes": [],
        "must_change_password": False,
        "preferred_language": "es",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.users.insert_one(admin_doc)
    logger.info(f"Admin user created: {ADMIN_EMAIL}")
    return True

# ========================== SEED DATA ==========================

@api_router.post("/seed")
async def seed_data():
    """Seed initial categories and admin user from environment variables"""
    categories = [
        {"id": generate_id(), "name_es": "Belleza y Estética", "name_en": "Beauty & Aesthetics", "slug": "belleza-estetica", "icon": "Sparkles", "image_url": "https://images.pexels.com/photos/853427/pexels-photo-853427.jpeg?auto=compress&cs=tinysrgb&dpr=2&h=650&w=940"},
        {"id": generate_id(), "name_es": "Salud", "name_en": "Health", "slug": "salud", "icon": "Heart", "image_url": "https://images.pexels.com/photos/4270095/pexels-photo-4270095.jpeg?auto=compress&cs=tinysrgb&dpr=2&h=650&w=940"},
        {"id": generate_id(), "name_es": "Fitness y Bienestar", "name_en": "Fitness & Wellness", "slug": "fitness-bienestar", "icon": "Dumbbell", "image_url": "https://images.unsplash.com/photo-1761971975724-31001b4de0bf?crop=entropy&cs=srgb&fm=jpg&ixid=M3w3NDQ2NDF8MHwxfHNlYXJjaHwzfHx5b2dhJTIwc3R1ZGlvJTIwaW50ZXJpb3IlMjBjYWxtfGVufDB8fHx8MTc3MTgwMjE1OXww&ixlib=rb-4.1.0&q=85"},
        {"id": generate_id(), "name_es": "Spa y Masajes", "name_en": "Spa & Massage", "slug": "spa-masajes", "icon": "Flower2", "image_url": "https://images.pexels.com/photos/5240677/pexels-photo-5240677.jpeg?auto=compress&cs=tinysrgb&dpr=2&h=650&w=940"},
        {"id": generate_id(), "name_es": "Servicios Legales", "name_en": "Legal Services", "slug": "servicios-legales", "icon": "Scale", "image_url": "https://images.unsplash.com/photo-1589829545856-d10d557cf95f?auto=format&fit=crop&q=80&w=2070"},
        {"id": generate_id(), "name_es": "Consultoría", "name_en": "Consulting", "slug": "consultoria", "icon": "Briefcase", "image_url": "https://images.unsplash.com/photo-1454165804606-c3d57bc86b40?auto=format&fit=crop&q=80&w=2070"},
        {"id": generate_id(), "name_es": "Automotriz", "name_en": "Automotive", "slug": "automotriz", "icon": "Car", "image_url": "https://images.unsplash.com/photo-1487754180451-c456f719a1fc?auto=format&fit=crop&q=80&w=2070"},
        {"id": generate_id(), "name_es": "Veterinaria", "name_en": "Veterinary", "slug": "veterinaria", "icon": "PawPrint", "image_url": "https://images.unsplash.com/photo-1628009368231-7bb7cfcb0def?auto=format&fit=crop&q=80&w=2070"},
        {"id": generate_id(), "name_es": "Salones, Servicios y Eventos", "name_en": "Venues, Services & Events", "slug": "salones-servicios-eventos", "icon": "PartyPopper", "image_url": "https://images.unsplash.com/photo-1519167758481-83f550bb49b3?auto=format&fit=crop&q=80&w=2070"},
    ]
    
    # Upsert categories by slug (idempotent - adds new ones, updates existing)
    for cat in categories:
        await db.categories.update_one(
            {"slug": cat["slug"]},
            {"$setOnInsert": cat},
            upsert=True
        )
    
    # Create admin user from environment variables - NEVER hardcode credentials
    admin_email = ADMIN_EMAIL
    admin_password = ADMIN_INITIAL_PASSWORD
    
    if not admin_email or not admin_password:
        logger.warning("ADMIN_EMAIL or ADMIN_INITIAL_PASSWORD not set in environment variables")
        return {
            "message": "Categories seeded. Admin not created - set ADMIN_EMAIL and ADMIN_INITIAL_PASSWORD in environment",
            "admin_created": False
        }
    
    # Check if admin already exists
    existing_admin = await db.users.find_one({"email": admin_email})
    if existing_admin:
        return {
            "message": "Categories seeded. Admin already exists",
            "admin_created": False,
            "admin_email": admin_email
        }
    
    admin_doc = {
        "id": generate_id(),
        "email": admin_email,
        "password_hash": hash_password(admin_password),
        "full_name": "Admin Bookvia",
        "phone": "+521234567890",
        "phone_verified": True,
        "role": UserRole.ADMIN,
        "totp_enabled": False,  # Must be enabled on first login
        "totp_secret": None,
        "backup_codes": [],
        "must_change_password": False,  # Set to True if using temp password
        "preferred_language": "es",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.users.insert_one(admin_doc)
    
    logger.info(f"Admin user created with email: {admin_email}")
    
    return {
        "message": "Seed data created successfully",
        "admin_created": True,
        "admin_email": admin_email,
        "note": "2FA setup required on first admin login"
    }

# ========================== CONTACT ==========================

class ContactMessage(BaseModel):
    name: str
    email: str
    subject: Optional[str] = None
    category: Optional[str] = None
    message: str

@api_router.post("/contact")
async def submit_contact_form(contact: ContactMessage):
    """Submit a contact form message"""
    contact_doc = {
        "id": generate_id(),
        "name": contact.name,
        "email": contact.email,
        "subject": contact.subject or "Sin asunto",
        "category": contact.category or "general",
        "message": contact.message,
        "status": "pending",
        "created_at": datetime.now(timezone.utc).isoformat(),
        "responded_at": None,
        "response": None
    }
    
    await db.contact_messages.insert_one(contact_doc)
    
    # Log for admin notification (in production, send email)
    logger.info(f"New contact message from {contact.email}: {contact.subject}")
    
    return {"success": True, "message": "Contact message received"}

@api_router.get("/admin/contact-messages")
async def get_contact_messages(
    status: Optional[str] = None,
    token_data: TokenData = Depends(require_admin)
):
    """Get all contact messages (admin only)"""
    filters = {}
    if status:
        filters["status"] = status
    
    messages = await db.contact_messages.find(
        filters, {"_id": 0}
    ).sort("created_at", -1).to_list(100)
    
    return messages

@api_router.put("/admin/contact-messages/{message_id}")
async def update_contact_message(
    message_id: str,
    response: str = None,
    status: str = "responded",
    token_data: TokenData = Depends(require_admin)
):
    """Update a contact message (admin only)"""
    update_data = {
        "status": status,
        "responded_at": datetime.now(timezone.utc).isoformat()
    }
    if response:
        update_data["response"] = response
    
    result = await db.contact_messages.update_one(
        {"id": message_id},
        {"$set": update_data}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Message not found")
    
    return {"success": True}

# ========================== CITIES & COUNTRIES ==========================

@api_router.get("/cities")
async def get_cities(country_code: str = "MX", with_businesses: bool = False, q: Optional[str] = None):
    """Get list of cities for a country. If with_businesses=True, only returns cities that have approved businesses, sorted by count."""
    country_upper = country_code.upper()

    if with_businesses:
        # Aggregate: count approved businesses per city in this country
        pipeline = [
            {"$match": {**VISIBLE_BUSINESS_FILTER, "country_code": country_upper}},
            {"$group": {"_id": "$city", "count": {"$sum": 1}}},
            {"$match": {"_id": {"$ne": None}}},
            {"$sort": {"count": -1}},
        ]
        if q:
            pipeline[0]["$match"]["city"] = {"$regex": q, "$options": "i"}

        agg_results = await db.businesses.aggregate(pipeline).to_list(200)

        cities_out = []
        for r in agg_results:
            city_name = r["_id"]
            # Try to enrich with seeded city data (state, slug)
            seeded = await db.cities.find_one(
                {"name": city_name, "country_code": country_upper},
                {"_id": 0}
            )
            if seeded:
                seeded["business_count"] = r["count"]
                cities_out.append(seeded)
            else:
                cities_out.append({
                    "name": city_name,
                    "slug": city_name.lower().replace(" ", "-"),
                    "country_code": country_upper,
                    "business_count": r["count"],
                })
        return cities_out

    # Default: return all seeded cities for the country
    filter_q = {"country_code": country_upper, "active": True}
    if q:
        filter_q["name"] = {"$regex": q, "$options": "i"}

    cities_from_db = await db.cities.find(filter_q, {"_id": 0}).sort("name", 1).to_list(500)
    if cities_from_db:
        return cities_from_db

    # Fallback: get from businesses
    cities = await db.businesses.distinct("city", {
        **VISIBLE_BUSINESS_FILTER,
        "country_code": country_upper
    })
    return [{"name": city, "slug": city.lower().replace(" ", "-"), "country_code": country_upper} for city in cities if city]


@api_router.post("/seed/countries")
async def seed_countries():
    """Seed countries and cities for multi-country support (idempotent via upsert)"""
    from data.countries import COUNTRIES
    from data.cities import CITIES as CITIES_DATA

    upserted = 0
    for c in COUNTRIES:
        result = await db.countries.update_one(
            {"code": c["code"]},
            {"$set": c},
            upsert=True,
        )
        if result.upserted_id or result.modified_count:
            upserted += 1

    # Seed cities from master data (idempotent via upsert)
    cities_upserted = 0
    for city in CITIES_DATA:
        result = await db.cities.update_one(
            {"slug": city["slug"], "country_code": city["country_code"]},
            {"$set": city},
            upsert=True,
        )
        if result.upserted_id or result.modified_count:
            cities_upserted += 1

    # Backfill: existing businesses without country_code default to MX
    await db.businesses.update_many(
        {"country_code": {"$exists": False}},
        {"$set": {"country_code": "MX"}}
    )
    await db.ledger_entries.update_many(
        {"country_code": {"$exists": False}},
        {"$set": {"country_code": "MX"}}
    )
    await db.settlements.update_many(
        {"country_code": {"$exists": False}},
        {"$set": {"country_code": "MX"}}
    )

    total_countries = await db.countries.count_documents({})
    return {
        "message": "Countries seed completed (idempotent)",
        "countries_total": total_countries,
        "countries_upserted": upserted,
        "cities_upserted": cities_upserted,
    }


# ========================== SUBSCRIPTION ENDPOINTS ==========================

import stripe as stripe_lib
stripe_lib.api_key = STRIPE_API_KEY
if "sk_test_emergent" in STRIPE_API_KEY:
    stripe_lib.api_base = "https://integrations.emergentagent.com/stripe"

SUBSCRIPTION_PRICE_MXN = 39.00
SUBSCRIPTION_TRIAL_DAYS = 30

async def get_or_create_stripe_price():
    """Get or create the Bookvia monthly subscription price in Stripe."""
    # Check if we already have a price stored
    config = await db.stripe_config.find_one({"type": "subscription_price"})
    if config and config.get("price_id"):
        return config["price_id"]
    
    try:
        # Create product
        product = stripe_lib.Product.create(
            name="Bookvia Suscripción Mensual",
            description="Suscripción mensual a la plataforma Bookvia"
        )
        # Create price (39 MXN monthly)
        price = stripe_lib.Price.create(
            product=product.id,
            unit_amount=int(SUBSCRIPTION_PRICE_MXN * 100),
            currency="mxn",
            recurring={"interval": "month"}
        )
        await db.stripe_config.update_one(
            {"type": "subscription_price"},
            {"$set": {"price_id": price.id, "product_id": product.id}},
            upsert=True
        )
        return price.id
    except Exception as e:
        logger.error(f"Failed to create Stripe price: {e}")
        return None


@businesses_router.post("/me/subscribe")
async def create_subscription_checkout(request: Request, token_data: TokenData = Depends(require_business)):
    """Create a Stripe Checkout Session for the monthly subscription."""
    user = await db.users.find_one({"id": token_data.user_id})
    if not user or not user.get("business_id"):
        raise HTTPException(status_code=404, detail="Business not found")
    
    business = await db.businesses.find_one({"id": user["business_id"]})
    if not business:
        raise HTTPException(status_code=404, detail="Business not found")
    
    # Check if already subscribed
    if business.get("stripe_subscription_id"):
        raise HTTPException(status_code=400, detail="Already subscribed")
    
    price_id = await get_or_create_stripe_price()
    if not price_id:
        raise HTTPException(status_code=500, detail="Payment configuration error")
    
    body = await request.json()
    origin_url = body.get("origin_url", os.environ.get("BASE_URL", ""))
    success_url = f"{origin_url}/business/subscription/success?session_id={{CHECKOUT_SESSION_ID}}"
    cancel_url = f"{origin_url}/business/dashboard"
    
    try:
        # Create or get Stripe Customer
        customer_id = business.get("stripe_customer_id")
        if not customer_id:
            customer = stripe_lib.Customer.create(
                email=business.get("email"),
                name=business.get("name"),
                metadata={"business_id": business["id"]}
            )
            customer_id = customer.id
            await db.businesses.update_one(
                {"id": business["id"]},
                {"$set": {"stripe_customer_id": customer_id}}
            )
        
        # Create Checkout Session for subscription
        session = stripe_lib.checkout.Session.create(
            customer=customer_id,
            payment_method_types=["card"],
            mode="subscription",
            line_items=[{"price": price_id, "quantity": 1}],
            subscription_data={"trial_period_days": SUBSCRIPTION_TRIAL_DAYS},
            success_url=success_url,
            cancel_url=cancel_url,
            metadata={"business_id": business["id"]}
        )
        
        # Record the transaction
        await db.payment_transactions.insert_one({
            "id": generate_id(),
            "business_id": business["id"],
            "session_id": session.id,
            "type": "subscription",
            "amount": SUBSCRIPTION_PRICE_MXN,
            "currency": "mxn",
            "payment_status": "pending",
            "created_at": datetime.now(timezone.utc).isoformat()
        })
        
        return {"url": session.url, "session_id": session.id}
    except Exception as e:
        logger.error(f"Stripe subscription error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@businesses_router.get("/me/subscription/status")
async def get_subscription_status(session_id: str = None, token_data: TokenData = Depends(require_business)):
    """Check subscription status."""
    user = await db.users.find_one({"id": token_data.user_id})
    if not user or not user.get("business_id"):
        raise HTTPException(status_code=404, detail="Business not found")
    
    business = await db.businesses.find_one({"id": user["business_id"]})
    
    # If session_id provided, verify the checkout session
    if session_id:
        try:
            session = stripe_lib.checkout.Session.retrieve(session_id)
            if session.payment_status in ("paid", "no_payment_required") or session.status == "complete":
                subscription_id = session.subscription
                # Update business with subscription info
                await db.businesses.update_one(
                    {"id": user["business_id"]},
                    {"$set": {
                        "stripe_subscription_id": subscription_id,
                        "stripe_customer_id": session.customer,
                        "subscription_status": "trialing",
                        "subscription_started_at": datetime.now(timezone.utc).isoformat()
                    }}
                )
                # Update transaction
                await db.payment_transactions.update_one(
                    {"session_id": session_id},
                    {"$set": {"payment_status": "completed", "subscription_id": subscription_id}}
                )
                return {
                    "status": "active",
                    "subscription_status": "trialing",
                    "subscription_id": subscription_id,
                    "trial": True,
                    "message": "Subscription activated with 30-day free trial"
                }
        except Exception as e:
            logger.error(f"Stripe session check error: {e}")
    
    # Return current subscription status with details from Stripe
    sub_id = business.get("stripe_subscription_id") if business else None
    sub_status = business.get("subscription_status", "none") if business else "none"
    
    result = {
        "status": sub_status if sub_id else "none",
        "subscription_status": sub_status,
        "subscription_id": sub_id,
        "trial": sub_status == "trialing",
        "current_period_end": None,
        "cancel_at_period_end": False
    }
    
    # Fetch live details from Stripe if subscription exists
    if sub_id:
        try:
            sub = stripe_lib.Subscription.retrieve(sub_id)
            result["current_period_end"] = datetime.fromtimestamp(sub.current_period_end, tz=timezone.utc).isoformat() if sub.current_period_end else None
            result["cancel_at_period_end"] = sub.cancel_at_period_end
            # Sync status from Stripe
            stripe_status = sub.status  # trialing, active, past_due, canceled, unpaid
            if stripe_status != sub_status:
                await db.businesses.update_one(
                    {"id": user["business_id"]},
                    {"$set": {"subscription_status": stripe_status}}
                )
                result["status"] = stripe_status
                result["subscription_status"] = stripe_status
                result["trial"] = stripe_status == "trialing"
        except Exception as e:
            logger.error(f"Stripe subscription fetch error: {e}")
    
    return result


@businesses_router.post("/me/subscription/cancel")
async def cancel_subscription(token_data: TokenData = Depends(require_business)):
    """Cancel the business subscription at the end of the current period."""
    user = await db.users.find_one({"id": token_data.user_id})
    if not user or not user.get("business_id"):
        raise HTTPException(status_code=404, detail="Business not found")
    
    business = await db.businesses.find_one({"id": user["business_id"]})
    if not business or not business.get("stripe_subscription_id"):
        raise HTTPException(status_code=400, detail="No active subscription")
    
    try:
        # Cancel at end of period (not immediately)
        stripe_lib.Subscription.modify(
            business["stripe_subscription_id"],
            cancel_at_period_end=True
        )
        
        await db.businesses.update_one(
            {"id": business["id"]},
            {"$set": {"subscription_cancel_requested": True}}
        )
        
        return {"message": "Subscription will be canceled at the end of the current billing period"}
    except Exception as e:
        logger.error(f"Stripe cancel error: {e}")
        raise HTTPException(status_code=500, detail="Error canceling subscription")


# ========================== BUSINESS CLOSURES ENDPOINTS ==========================

class ClosureDateCreate(BaseModel):
    date: str  # YYYY-MM-DD
    reason: Optional[str] = None

@businesses_router.get("/me/closures")
async def get_business_closures(token_data: TokenData = Depends(require_business)):
    """Get all closure dates for the authenticated business."""
    user = await db.users.find_one({"id": token_data.user_id})
    if not user or not user.get("business_id"):
        raise HTTPException(status_code=404, detail="Business not found")
    closures = await db.business_closures.find(
        {"business_id": user["business_id"], "is_deleted": False},
        {"_id": 0}
    ).to_list(500)
    return closures

@businesses_router.post("/me/closures")
async def add_business_closure(data: ClosureDateCreate, token_data: TokenData = Depends(require_business)):
    """Mark a date as closed."""
    user = await db.users.find_one({"id": token_data.user_id})
    if not user or not user.get("business_id"):
        raise HTTPException(status_code=404, detail="Business not found")

    # Check if date already marked
    existing = await db.business_closures.find_one({
        "business_id": user["business_id"], "date": data.date, "is_deleted": False
    })
    if existing:
        raise HTTPException(status_code=400, detail="Date already marked as closed")

    closure = {
        "id": generate_id(),
        "business_id": user["business_id"],
        "date": data.date,
        "reason": data.reason,
        "is_deleted": False,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.business_closures.insert_one(closure)
    del closure["_id"]
    return closure

@businesses_router.delete("/me/closures/{date}")
async def remove_business_closure(date: str, token_data: TokenData = Depends(require_business)):
    """Remove a closure date (re-open)."""
    user = await db.users.find_one({"id": token_data.user_id})
    if not user or not user.get("business_id"):
        raise HTTPException(status_code=404, detail="Business not found")
    result = await db.business_closures.update_one(
        {"business_id": user["business_id"], "date": date, "is_deleted": False},
        {"$set": {"is_deleted": True}}
    )
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Closure not found")
    return {"message": "Closure removed"}


# ========================== PHOTO UPLOAD ENDPOINTS (Cloudinary + Emergent fallback) ==========================

from services.cloudinary_service import (
    init_cloudinary, is_configured as cloudinary_configured,
    validate_image, upload_image, delete_image as cloudinary_delete,
    ALLOWED_EXTENSIONS as CLOUDINARY_ALLOWED_EXT
)
from services.storage import init_storage, put_object, get_object, generate_upload_path, ALLOWED_IMAGE_TYPES, ALLOWED_IMAGE_EXTENSIONS, MAX_FILE_SIZE


# ── Generic upload endpoint ──────────────────────────
@api_router.post("/upload/image")
async def upload_image_endpoint(
    file: UploadFile = File(...),
    folder: str = "business_gallery",
    entity_id: str = "",
    token_data: TokenData = Depends(require_auth),
):
    """Upload an image to Cloudinary. Used by logo upload, gallery, etc."""
    data = await file.read()
    ok, err = validate_image(file.filename, file.content_type, len(data))
    if not ok:
        raise HTTPException(status_code=400, detail=err)

    if not cloudinary_configured():
        raise HTTPException(status_code=503, detail="Image storage not configured")

    try:
        result = upload_image(data, folder, entity_id)
        return {"secure_url": result["secure_url"], "public_id": result["public_id"]}
    except Exception as e:
        logger.error(f"Cloudinary upload error: {e}")
        raise HTTPException(status_code=500, detail="Failed to upload image")


# ── Business gallery photos ──────────────────────────
@businesses_router.post("/me/photos")
async def upload_business_photo(file: UploadFile = File(...), token_data: TokenData = Depends(require_business)):
    """Upload a gallery photo for the authenticated business."""
    user = await db.users.find_one({"id": token_data.user_id})
    if not user or not user.get("business_id"):
        raise HTTPException(status_code=404, detail="Business not found")

    data = await file.read()
    ok, err = validate_image(file.filename, file.content_type, len(data))
    if not ok:
        raise HTTPException(status_code=400, detail=err)

    business_id = user["business_id"]

    # Try Cloudinary first, fallback to Emergent Storage
    if cloudinary_configured():
        try:
            result = upload_image(data, "business_gallery", business_id)
            photo_url = result["secure_url"]
            public_id = result["public_id"]
        except Exception as e:
            logger.error(f"Cloudinary upload failed: {e}")
            raise HTTPException(status_code=500, detail="Failed to upload photo")
    else:
        # Fallback: Emergent Object Storage (preview env)
        ext = file.filename.split(".")[-1].lower() if "." in file.filename else "jpg"
        content_type = file.content_type
        if ext in ("jfif", "jpg", "jpeg", "pjpeg"):
            content_type = "image/jpeg"
        path = generate_upload_path(business_id, file.filename)
        try:
            result = put_object(path, data, content_type)
            base_url = os.environ.get("BASE_URL", "")
            photo_url = f"{base_url}/api/files/{result['path']}"
            public_id = path
        except Exception as e:
            logger.error(f"Storage upload failed: {e}")
            raise HTTPException(status_code=500, detail="Failed to upload photo")

    # Store reference in DB
    file_record = {
        "id": generate_id(),
        "business_id": business_id,
        "url": photo_url,
        "public_id": public_id,
        "storage": "cloudinary" if cloudinary_configured() else "emergent",
        "original_filename": file.filename,
        "content_type": file.content_type,
        "is_deleted": False,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.business_photos.insert_one(file_record)

    # Add to business photos array
    await db.businesses.update_one(
        {"id": business_id},
        {"$push": {"photos": photo_url}}
    )

    return {
        "id": file_record["id"],
        "url": photo_url,
        "public_id": public_id,
        "original_filename": file.filename
    }


@businesses_router.delete("/me/photos/{photo_id}")
async def delete_business_photo(photo_id: str, token_data: TokenData = Depends(require_business)):
    """Delete a business photo."""
    user = await db.users.find_one({"id": token_data.user_id})
    if not user or not user.get("business_id"):
        raise HTTPException(status_code=404, detail="Business not found")

    photo = await db.business_photos.find_one({"id": photo_id, "business_id": user["business_id"], "is_deleted": False})
    if not photo:
        raise HTTPException(status_code=404, detail="Photo not found")

    # Delete from cloud storage
    if photo.get("storage") == "cloudinary" and photo.get("public_id"):
        cloudinary_delete(photo["public_id"])

    await db.business_photos.update_one({"id": photo_id}, {"$set": {"is_deleted": True}})

    # Remove from business photos array - handle legacy photos with storage_path instead of url
    photo_url = photo.get("url") or photo.get("storage_path")
    if photo_url:
        # For legacy photos, construct URL from storage_path
        if not photo_url.startswith("http"):
            base_url = os.environ.get("BASE_URL", "")
            photo_url = f"{base_url}/api/files/{photo_url}"
        await db.businesses.update_one(
            {"id": user["business_id"]},
            {"$pull": {"photos": photo_url}}
        )

    return {"message": "Photo deleted"}


@businesses_router.get("/me/photos")
async def get_business_photos(token_data: TokenData = Depends(require_business)):
    """Get all photos for the authenticated business."""
    user = await db.users.find_one({"id": token_data.user_id})
    if not user or not user.get("business_id"):
        raise HTTPException(status_code=404, detail="Business not found")

    photos = await db.business_photos.find(
        {"business_id": user["business_id"], "is_deleted": False},
        {"_id": 0}
    ).to_list(100)

    # Normalize: ensure all photos have 'url' field (handle legacy photos with storage_path)
    base_url = os.environ.get("BASE_URL", "")
    for photo in photos:
        if "url" not in photo and "storage_path" in photo:
            photo["url"] = f"{base_url}/api/files/{photo['storage_path']}"

    return photos


# ── Business logo upload ──────────────────────────
@businesses_router.post("/me/logo")
async def upload_business_logo(file: UploadFile = File(...), token_data: TokenData = Depends(require_business)):
    """Upload or replace the business logo."""
    user = await db.users.find_one({"id": token_data.user_id})
    if not user or not user.get("business_id"):
        raise HTTPException(status_code=404, detail="Business not found")

    data = await file.read()
    ok, err = validate_image(file.filename, file.content_type, len(data))
    if not ok:
        raise HTTPException(status_code=400, detail=err)

    business_id = user["business_id"]
    business = await db.businesses.find_one({"id": business_id})

    if cloudinary_configured():
        # Delete old logo if exists
        old_public_id = business.get("logo_public_id") if business else None
        if old_public_id:
            cloudinary_delete(old_public_id)

        try:
            result = upload_image(data, "business_logo", business_id)
            logo_url = result["secure_url"]
            public_id = result["public_id"]
        except Exception as e:
            logger.error(f"Logo upload failed: {e}")
            raise HTTPException(status_code=500, detail="Failed to upload logo")
    else:
        # Fallback: Emergent storage
        path = generate_upload_path(business_id, f"logo_{file.filename}")
        content_type = "image/jpeg" if file.filename.lower().endswith(("jfif", "jpg", "jpeg")) else file.content_type
        try:
            result = put_object(path, data, content_type)
            base_url = os.environ.get("BASE_URL", "")
            logo_url = f"{base_url}/api/files/{result['path']}"
            public_id = path
        except Exception as e:
            logger.error(f"Logo upload failed: {e}")
            raise HTTPException(status_code=500, detail="Failed to upload logo")

    await db.businesses.update_one(
        {"id": business_id},
        {"$set": {"logo_url": logo_url, "logo_public_id": public_id}}
    )

    return {"secure_url": logo_url, "public_id": public_id}


# ── Serve files from Emergent storage (legacy/fallback) ──────────
@api_router.get("/files/{path:path}")
async def serve_file(path: str):
    """Serve uploaded files from Emergent object storage (fallback)."""
    # Check business_photos collection first
    record = await db.business_photos.find_one({"public_id": path, "is_deleted": False})
    
    # If not found in photos, check if it's a logo (stored in businesses collection)
    if not record:
        business = await db.businesses.find_one({"logo_public_id": path})
        if business:
            record = {"content_type": "image/jpeg"}  # Default for logos
    
    if not record:
        raise HTTPException(status_code=404, detail="File not found")

    try:
        data, content_type = get_object(path)
    except Exception as e:
        logger.error(f"Failed to retrieve file: {e}")
        raise HTTPException(status_code=500, detail="Failed to retrieve file")

    return Response(
        content=data,
        media_type=record.get("content_type", content_type),
        headers={"Cache-Control": "public, max-age=86400"}
    )


# ========================== INCLUDE ROUTERS ==========================

api_router.include_router(auth_router)
api_router.include_router(users_router)
api_router.include_router(businesses_router)
api_router.include_router(services_router)
api_router.include_router(bookings_router)
api_router.include_router(reviews_router)
api_router.include_router(categories_router)
api_router.include_router(payments_router)
api_router.include_router(admin_router)
api_router.include_router(notifications_router)
api_router.include_router(finance_router)

# Import and include SEO router (handles /sitemap.xml, /robots.txt, /api/seo/*)
from routers.seo import seo_router
app.include_router(seo_router)  # SEO routes at root level (no /api prefix for sitemap/robots)

app.include_router(api_router)

# ========================== MIDDLEWARE ==========================

# Rate limiting middleware
from middleware.rate_limit import RateLimitMiddleware
app.add_middleware(RateLimitMiddleware)

# HTTPS redirect middleware for proxy
from starlette.middleware.base import BaseHTTPMiddleware
from starlette.responses import RedirectResponse

class HTTPSRedirectMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request, call_next):
        response = await call_next(request)
        # If it's a redirect and we're behind HTTPS proxy, fix the Location header
        if response.status_code in (307, 308, 301, 302):
            location = response.headers.get('location', '')
            x_forwarded_proto = request.headers.get('x-forwarded-proto', 'http')
            if x_forwarded_proto == 'https' and location.startswith('http://'):
                new_location = location.replace('http://', 'https://', 1)
                return RedirectResponse(url=new_location, status_code=response.status_code)
        return response

app.add_middleware(HTTPSRedirectMiddleware)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()

@app.on_event("startup")
async def startup_event():
    try:
        init_storage()
        logger.info("Object storage initialized")
    except Exception as e:
        logger.warning(f"Object storage init failed (uploads will retry): {e}")
    # Initialize Cloudinary
    from services.cloudinary_service import init_cloudinary
    if init_cloudinary():
        logger.info("Cloudinary initialized")
    else:
        logger.warning("Cloudinary not configured - using fallback storage")
